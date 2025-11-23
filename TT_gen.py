# TT_gen.py -- Timetable generator with basket system for cross-department electives
# Run: python TT_gen.py
# Requires: pandas, openpyxl

import pandas as pd
import random
from datetime import datetime, time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
import traceback
import os
import json
import math

# ---------------------------
# Default Configuration
# ---------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")

DEFAULT_CONFIG = {
    "days": ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'],
    "LECTURE_MIN": 90,
    "LAB_MIN": 120,
    "TUTORIAL_MIN": 60,
    "SELF_STUDY_MIN": 60,
    "MORNING_BREAK_START": "10:30",
    "MORNING_BREAK_END": "10:45",
    "LUNCH_BREAK_START": "13:00",
    "LUNCH_BREAK_END": "13:45",
    "LECTURE_TUTORIAL_BREAK_START": "15:30",
    "LECTURE_TUTORIAL_BREAK_END": "15:40",
    "TIME_SLOTS": [
        ["07:30", "09:00"],
        ["09:00", "09:30"],
        ["09:30", "10:00"],
        ["10:00", "10:30"],
        ["10:30", "10:45"],
        ["10:45", "11:00"],
        ["11:00", "11:30"],
        ["11:30", "12:00"],
        ["12:00", "12:15"],
        ["12:15", "12:30"],
        ["12:30", "13:00"],
        ["13:00", "13:30"],
        ["13:30", "14:00"],
        ["14:00", "14:30"],
        ["14:30", "15:00"],
        ["15:00", "15:30"],
        ["15:30", "15:40"],
        ["15:40", "16:00"],
        ["16:00", "16:30"],
        ["16:30", "17:00"],
        ["17:00", "17:30"],
        ["17:30", "18:00"],
        ["18:00", "18:30"],
        ["18:30", "23:59"]
    ],
    "USE_CUSTOM_SLOTS": False
}

# ---------------------------
# Load Configuration
# ---------------------------
def load_configuration():
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config = json.load(f)
            print("✓ Loaded configuration from config.json")
    except Exception:
        print("Using default configuration")
        config = DEFAULT_CONFIG.copy()
    return config

# ---------------------------
# Parse time strings
# ---------------------------
def parse_time_string(time_str):
    try:
        hour, minute = map(int, time_str.split(':'))
        return time(hour, minute)
    except:
        return time(10, 30)

# ---------------------------
# Generate time slots
# ---------------------------
def generate_time_slots(config):
    time_slots_config = config.get("TIME_SLOTS", DEFAULT_CONFIG["TIME_SLOTS"])
    slots = []
    for slot_config in time_slots_config:
        try:
            start_str, end_str = slot_config
            start = parse_time_string(start_str)
            end = parse_time_string(end_str)
            slots.append((start, end))
        except:
            continue
    
    if not slots:
        slots = [
            (time(9, 0), time(10, 30)),
            (time(11, 0), time(12, 30)),
            (time(14, 0), time(15, 30)),
            (time(16, 0), time(17, 30))
        ]
    
    return slots

@dataclass
class UnscheduledComponent:
    department: str
    semester: int
    code: str
    name: str
    faculty: str
    component_type: str
    sessions: int
    section: int
    reason: str

INPUT_DIR = os.path.join(BASE_DIR, "inputs")
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")

# ---------------------------
# Load data
# ---------------------------
def load_data():
    try:
        df = pd.read_csv(os.path.join(INPUT_DIR, 'combined.csv'))
    except FileNotFoundError:
        raise SystemExit("Error: 'combined.csv' not found in working directory.")

    try:
        rooms_df = pd.read_csv(os.path.join(INPUT_DIR, 'rooms.csv'))
    except FileNotFoundError:
        rooms_df = pd.DataFrame(columns=['roomNumber', 'type', 'capacity'])
    
    # Parse room capacities and categorize rooms
    lecture_rooms = []
    computer_lab_rooms = []
    large_rooms = []
    auditorium_rooms = []
    
    # Create dictionaries to store rooms by capacity
    small_rooms = []  # < 60
    medium_rooms = []  # 60-120
    # large_rooms = []  # > 120 (already defined)
    
    for _, row in rooms_df.iterrows():
        room_number = str(row.get('roomNumber', '')).strip()
        room_type = str(row.get('type', '')).strip().upper()
        capacity = int(row.get('capacity', 60)) if pd.notna(row.get('capacity', 60)) else 60
        
        if room_type == 'LECTURE_ROOM':
            lecture_rooms.append(room_number)
            if capacity < 60:
                small_rooms.append(room_number)
            elif capacity <= 120:
                medium_rooms.append(room_number)
            else:
                large_rooms.append(room_number)
        elif room_type == 'COMPUTER_LAB':
            computer_lab_rooms.append(room_number)
        elif room_type == 'SEATER_120':
            large_rooms.append(room_number)
        elif room_type == 'SEATER_240':
            auditorium_rooms.append(room_number)
    
    return df, lecture_rooms, computer_lab_rooms, large_rooms, auditorium_rooms, small_rooms, medium_rooms

# ---------------------------
# Load electives
# ---------------------------
def load_electives():
    electives_path_csv = os.path.join(INPUT_DIR, "elective.csv")
    electives_path_xlsx = os.path.join(INPUT_DIR, "elective.xlsx")
    
    electives_df = None
    
    if os.path.exists(electives_path_csv):
        electives_df = pd.read_csv(electives_path_csv)
    elif os.path.exists(electives_path_xlsx):
        electives_df = pd.read_excel(electives_path_xlsx)
    
    if electives_df is None:
        return None
    
    electives_dict = {}
    
    for _, row in electives_df.iterrows():
        if len(row) >= 5:  # Now expecting 5 columns with count
            sem_str = str(row.iloc[0]).strip().lower()
            electives_str = str(row.iloc[1]).strip()
            faculty_str = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
            n_value = int(row.iloc[3]) if pd.notna(row.iloc[3]) and str(row.iloc[3]).isdigit() else 1
            count_str = str(row.iloc[4]).strip() if pd.notna(row.iloc[4]) else ''
        elif len(row) >= 4:
            sem_str = str(row.iloc[0]).strip().lower()
            electives_str = str(row.iloc[1]).strip()
            faculty_str = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
            n_value = int(row.iloc[3]) if pd.notna(row.iloc[3]) and str(row.iloc[3]).isdigit() else 1
            count_str = ''
        elif len(row) >= 3:
            sem_str = str(row.iloc[0]).strip().lower()
            electives_str = str(row.iloc[1]).strip()
            faculty_str = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
            n_value = 1
            count_str = ''
        else:
            sem_str = str(row.iloc[0]).strip().lower()
            electives_str = str(row.iloc[1]).strip()
            faculty_str = ''
            n_value = 1
            count_str = ''
        
        if "1st" in sem_str:
            sem = 1
            basket_label = "ELECTIVE"
        elif "3rd" in sem_str:
            sem = 3
            basket_label = "ELECTIVE"
        elif "5th(b1)" in sem_str:
            sem = 5
            basket_label = "B1"
        elif "5th(b2)" in sem_str:
            sem = 5
            basket_label = "B2"
        elif "7th(b1)" in sem_str:
            sem = 7
            basket_label = "B1"
        elif "7th(b2)" in sem_str:
            sem = 7
            basket_label = "B2"
        elif "7th(b3)" in sem_str:
            sem = 7
            basket_label = "B3"
        elif "7th(b4)" in sem_str:
            sem = 7
            basket_label = "B4"
        else:
            continue
        
        electives_list = [e.strip() for e in electives_str.split(',') if e.strip()]
        faculty_list = [f.strip() for f in faculty_str.split(',') if f.strip()] if faculty_str else []
        
        # Parse count values for each elective
        count_list = []
        if count_str:
            count_list = [c.strip() for c in count_str.split(',') if c.strip()]
        
        # Ensure we have a count for each elective (default to 60)
        while len(count_list) < len(electives_list):
            count_list.append('60')
        
        # Convert count strings to integers
        counts = []
        for count_str in count_list:
            try:
                count = int(count_str)
            except:
                count = 60  # Default to 60 if conversion fails
            counts.append(count)
        
        if sem not in electives_dict:
            electives_dict[sem] = {}
        
        electives_dict[sem][basket_label] = {
            'electives': electives_list,
            'faculty': faculty_list,
            'n_value': n_value,
            'counts': counts
        }
    
    return electives_dict

# ---------------------------
# Helper functions
# ---------------------------
def slot_minutes(slot):
    s, e = slot
    s_m = s.hour*60 + s.minute
    e_m = e.hour*60 + e.minute
    if e_m < s_m:
        e_m += 24*60
    return e_m - s_m

def overlaps(a_start, a_end, b_start, b_end):
    a_s_min = a_start.hour*60 + a_start.minute
    a_e_min = a_end.hour*60 + a_end.minute
    b_s_min = b_start.hour*60 + b_start.minute
    b_e_min = b_end.hour*60 + b_end.minute
    return (a_s_min < b_e_min) and (b_s_min < a_e_min)

def is_break_time_slot(slot, semester=None, comp_type=None, config=None):
    if config is None:
        config = load_configuration()
    
    start, end = slot
    MORNING_BREAK_START = parse_time_string(config.get("MORNING_BREAK_START", "10:30"))
    MORNING_BREAK_END = parse_time_string(config.get("MORNING_BREAK_END", "10:45"))
    LUNCH_BREAK_START = parse_time_string(config.get("LUNCH_BREAK_START", "13:00"))
    LUNCH_BREAK_END = parse_time_string(config.get("LUNCH_BREAK_END", "13:45"))
    LECTURE_TUTORIAL_BREAK_START = parse_time_string(config.get("LECTURE_TUTORIAL_BREAK_START", "15:30"))
    LECTURE_TUTORIAL_BREAK_END = parse_time_string(config.get("LECTURE_TUTORIAL_BREAK_END", "15:40"))
    
    if overlaps(start, end, MORNING_BREAK_START, MORNING_BREAK_END):
        return True
    if overlaps(start, end, LUNCH_BREAK_START, LUNCH_BREAK_END):
        return True
    if comp_type in ['LEC', 'TUT'] and overlaps(start, end, LECTURE_TUTORIAL_BREAK_START, LECTURE_TUTORIAL_BREAK_END):
        return True
    return False

def is_minor_slot(slot):
    start, end = slot
    return (start == time(7, 30) and end == time(9, 0)) or (start == time(18, 30))

def is_lecture_unfriendly_slot(slot):
    start, _ = slot
    return start >= time(17, 30)

def select_faculty_for_section(faculty_field, section_char='A'):
    if pd.isna(faculty_field) or str(faculty_field).strip().lower() in ['nan', 'none', '']:
        return "TBD"
    
    s = str(faculty_field).strip()
    
    if '&' in s:
        faculties = [f.strip() for f in s.split('&')]
        if len(faculties) >= 2:
            if section_char.upper() == 'A':
                return faculties[0]
            elif section_char.upper() == 'B':
                return faculties[1]
            else:
                return faculties[0]
        else:
            return faculties[0] if faculties else "TBD"
    else:
        for sep in ['/', ',', ';']:
            if sep in s:
                return s.split(sep)[0].strip()
        return s

def is_elective(course_row):
    name = str(course_row.get('Course Name', '')).lower()
    code = str(course_row.get('Course Code', '')).lower()
    keywords = ["elective", "oe", "open elective", "pe", "program elective"]
    return any(k in name for k in keywords) or any(k in code for k in keywords)

def get_course_priority(row):
    try:
        l = int(row.get('L', 0)) if pd.notna(row.get('L', 0)) else 0
        t = int(row.get('T', 0)) if pd.notna(row.get('T', 0)) else 0
        p = int(row.get('P', 0)) if pd.notna(row.get('P', 0)) else 0
        return -(l + t + p)
    except:
        return 0

def calculate_required_sessions(course_row, config):
    l = int(course_row['L']) if ('L' in course_row and pd.notna(course_row['L'])) else 0
    t = int(course_row['T']) if ('T' in course_row and pd.notna(course_row['T'])) else 0
    p = int(course_row['P']) if ('P' in course_row and pd.notna(course_row['P'])) else 0
    s = int(course_row['S']) if ('S' in course_row and pd.notna(course_row['S'])) else 0
    
    LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    LAB_MIN = config.get("LAB_MIN", DEFAULT_CONFIG["LAB_MIN"])
    TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    SELF_STUDY_MIN = config.get("SELF_STUDY_MIN", DEFAULT_CONFIG["SELF_STUDY_MIN"])
    
    LECTURE_HOURS = LECTURE_MIN / 60
    LAB_HOURS = LAB_MIN / 60
    TUTORIAL_HOURS = TUTORIAL_MIN / 60
    SELF_STUDY_HOURS = SELF_STUDY_MIN / 60
    
    lec_sessions = math.ceil(l / LECTURE_HOURS) if l > 0 and LECTURE_HOURS > 0 else 0
    tut_sessions = math.ceil(t / TUTORIAL_HOURS) if t > 0 and TUTORIAL_HOURS > 0 else 0
    lab_sessions = 1 if p > 0 else 0
    ss_sessions = math.ceil(s / SELF_STUDY_HOURS) if s > 0 and SELF_STUDY_HOURS > 0 else 0
    
    lab_duration = p * 60 if p > 0 else 0
    
    return (lec_sessions, tut_sessions, lab_sessions, ss_sessions, lab_duration)

def get_required_room_type(course_row):
    try:
        p = int(course_row['P']) if ('P' in course_row and pd.notna(course_row['P'])) else 0
        return 'COMPUTER_LAB' if p > 0 else 'LECTURE_ROOM'
    except:
        return 'LECTURE_ROOM'

def is_auditorium_course(course_row):
    if '240' in course_row and pd.notna(course_row['240']) and str(course_row['240']).strip().lower() == 'yes':
        return True
    return False

# ---------------------------
# Room allocation
# ---------------------------
def find_suitable_room_for_slot(course_code, room_type, day, slot_indices, room_schedule, course_room_mapping, config, lecture_rooms, computer_lab_rooms, auditorium_rooms, small_rooms, medium_rooms, large_rooms, required_capacity=60):
    if course_code in course_room_mapping:
        fixed_room = course_room_mapping[course_code]
        for si in slot_indices:
            if si in room_schedule[fixed_room][day]:
                return None
        return fixed_room

    # Select appropriate room pool based on capacity requirements
    pool = []
    if room_type == 'COMPUTER_LAB':
        pool = computer_lab_rooms
    elif room_type == 'AUDITORIUM':
        pool = auditorium_rooms
    else:
        # For lecture rooms, select based on capacity
        if required_capacity > 120:
            pool = large_rooms + auditorium_rooms
        elif required_capacity > 60:
            pool = medium_rooms + large_rooms
        else:
            pool = small_rooms + medium_rooms + large_rooms
        
    if not pool:
        # Fallback to all lecture rooms if specific capacity rooms aren't available
        pool = lecture_rooms
        
    if not pool:
        return None
    
    random.shuffle(pool)
    for room in pool:
        if room not in room_schedule:
            room_schedule[room] = {d: set() for d in range(len(config.get("days", DEFAULT_CONFIG["days"])))}
        if all(si not in room_schedule[room][day] for si in slot_indices):
            course_room_mapping[course_code] = room
            return room
    return None

def find_consecutive_slots_for_minutes(timetable, day, start_idx, required_minutes,
                                       semester, professor_schedule, faculty,
                                       room_schedule, room_type, course_code, course_room_mapping, comp_type, config, TIME_SLOTS, lecture_rooms, computer_lab_rooms, auditorium_rooms, small_rooms, medium_rooms, large_rooms, required_capacity=60):
    n = len(TIME_SLOTS)
    slot_indices = []
    i = start_idx
    accumulated = 0

    while i < n and accumulated < required_minutes:
        if is_minor_slot(TIME_SLOTS[i]):
            return None, None
        if is_break_time_slot(TIME_SLOTS[i], semester, comp_type, config):
            return None, None
        if timetable[day][i]['type'] is not None:
            return None, None
        if faculty in professor_schedule and i in professor_schedule[faculty][day]:
            return None, None
        
        if room_type == 'COMPUTER_LAB' and not computer_lab_rooms:
            return None, None
        if room_type == 'LECTURE_ROOM' and not lecture_rooms:
            return None, None
        if room_type == 'AUDITORIUM' and not auditorium_rooms:
            return None, None

        slot_indices.append(i)
        accumulated += slot_minutes(TIME_SLOTS[i])
        i += 1

    if accumulated >= required_minutes:
        room = find_suitable_room_for_slot(course_code, room_type, day, slot_indices, room_schedule, course_room_mapping, config, lecture_rooms, computer_lab_rooms, auditorium_rooms, small_rooms, medium_rooms, large_rooms, required_capacity)
        if room is not None:
            return slot_indices, room

    return None, None

def get_all_possible_start_indices_for_duration(comp_type=None, TIME_SLOTS=None):
    idxs = list(range(len(TIME_SLOTS)))
    random.shuffle(idxs)
    
    if comp_type == 'LEC':
        idxs = [i for i in idxs if not is_lecture_unfriendly_slot(TIME_SLOTS[i])]
    
    return idxs

def check_professor_availability(professor_schedule, faculty, day, start_idx, duration_slots, TIME_SLOTS):
    if faculty not in professor_schedule:
        return True
    if not professor_schedule[faculty][day]:
        return True
    new_start = TIME_SLOTS[start_idx][0]
    new_start_m = new_start.hour*60 + new_start.minute
    MIN_GAP = 180
    for s in professor_schedule[faculty][day]:
        exist_start = TIME_SLOTS[s][0]
        exist_m = exist_start.hour*60 + exist_start.minute
        if abs(exist_m - new_start_m) < MIN_GAP:
            return False
    return True

def check_course_component_conflict(timetable, day, course_code, comp_type, TIME_SLOTS, is_auditorium=False):
    for slot_idx in range(len(TIME_SLOTS)):
        if timetable[day][slot_idx]['type'] is not None:
            existing_code = timetable[day][slot_idx]['code']
            existing_type = timetable[day][slot_idx]['type']
            
            if existing_code == course_code and 'Courses' not in existing_code:
                if is_auditorium:
                    if comp_type == 'LEC' and existing_type == 'LEC':
                        return True
                    if (comp_type == 'LEC' and existing_type == 'TUT') or \
                       (comp_type == 'TUT' and existing_type == 'LEC'):
                        return True
                else:
                    if comp_type == 'LEC' and existing_type == 'LEC':
                        return True
                    if (comp_type == 'LEC' and existing_type == 'TUT') or \
                       (comp_type == 'TUT' and existing_type == 'LEC'):
                        return True
    return False

def add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, comp_type, section, reason):
    existing = next((u for u in unscheduled_components if u.code == code), None)
    if existing:
        if comp_type not in existing.component_type:
            existing.component_type += f", {comp_type}"
        if reason not in existing.reason:
            existing.reason += f"; {reason}"
    else:
        unscheduled_components.append(UnscheduledComponent(department, semester, code, name, faculty, comp_type, 1, section, reason))

# ---------------------------
# Basket scheduling
# ---------------------------
def check_basket_slot_conflict(new_day, new_slots, existing_basket_slots):
    for existing in existing_basket_slots:
        if existing['day'] == new_day:
            for new_slot in new_slots:
                for existing_slot in existing['slots']:
                    if new_slot == existing_slot:
                        return True
            
            new_min = min(new_slots)
            new_max = max(new_slots)
            existing_min = min(existing['slots'])
            existing_max = max(existing['slots'])
            
            if abs(new_min - existing_max) == 1 or abs(new_max - existing_min) == 1:
                return True
    return False

def get_basket_config_for_semester(semester, electives_data):
    if electives_data and semester in electives_data:
        baskets = electives_data[semester]
        config = {
            'departments': ['CSE', 'DSAI', 'ECE'],
            'lectures': 2,
            'tutorials': 1,
        }
        
        if 'B1' in baskets and 'B2' in baskets and 'B3' in baskets and 'B4' in baskets:
            config['baskets'] = ['B1', 'B2', 'B3', 'B4']
        elif 'B1' in baskets and 'B2' in baskets:
            config['baskets'] = ['B1', 'B2']
        else:
            config['label'] = 'ELECTIVE'
        
        return config
    else:
        if semester == 1:
            return {
                'departments': ['CSE', 'DSAI', 'ECE'],
                'lectures': 2,
                'tutorials': 1,
                'label': 'ELECTIVE'
            }
        elif semester == 3:
            return {
                'departments': ['CSE', 'DSAI', 'ECE'],
                'lectures': 2,
                'tutorials': 1,
                'label': 'ELECTIVE'
            }
        elif semester == 5:
            return {
                'departments': ['CSE', 'DSAI', 'ECE'],
                'lectures': 2,
                'tutorials': 1,
                'baskets': ['B1', 'B2']
            }
        elif semester == 7:
            return {
                'departments': ['CSE', 'DSAI', 'ECE'],
                'lectures': 2,
                'tutorials': 1,
                'baskets': ['B1', 'B2', 'B3', 'B4']
            }
        else:
            return None

def schedule_basket_slots(semester, all_department_timetables, professor_schedule, room_schedule, course_room_mapping, config, TIME_SLOTS, electives_data, lecture_rooms, large_rooms, small_rooms, medium_rooms):
    BASKET_LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    BASKET_TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    
    basket_config = get_basket_config_for_semester(semester, electives_data)
    if not basket_config:
        return {}
    
    departments = basket_config['departments']
    
    if 'baskets' in basket_config:
        all_basket_slots = {}
        existing_slots = []
        
        for basket_label in basket_config['baskets']:
            lec_count = basket_config['lectures']
            tut_count = basket_config['tutorials']
            
            basket_result = schedule_single_basket_with_constraints(
                semester, basket_label, departments, lec_count, tut_count,
                all_department_timetables, professor_schedule, room_schedule, 
                course_room_mapping, existing_slots, config, TIME_SLOTS, lecture_rooms, large_rooms, small_rooms, medium_rooms, electives_data
            )
            all_basket_slots[basket_label] = basket_result
            
            for lecture in basket_result.get('lectures', []):
                existing_slots.append({
                    'day': lecture['day'],
                    'slots': lecture['slots'],
                    'type': 'lecture',
                    'basket': basket_label,
                    'rooms': lecture['rooms']
                })
            for tutorial in basket_result.get('tutorials', []):
                existing_slots.append({
                    'day': tutorial['day'],
                    'slots': tutorial['slots'],
                    'type': 'tutorial',
                    'basket': basket_label,
                    'rooms': tutorial['rooms']
                })
        
        return all_basket_slots
    else:
        return schedule_single_basket_with_constraints(
        semester, basket_config['label'], departments, 
        basket_config['lectures'], basket_config['tutorials'],
        all_department_timetables, professor_schedule, room_schedule, course_room_mapping, [], config, TIME_SLOTS, lecture_rooms, large_rooms, small_rooms, medium_rooms, electives_data
    )

def schedule_single_basket_with_constraints(semester, label, departments, lec_count, tut_count,
                                         all_department_timetables, professor_schedule, room_schedule, 
                                         course_room_mapping, existing_basket_slots, config, TIME_SLOTS, lecture_rooms, large_rooms, small_rooms, medium_rooms, electives_data):
    BASKET_LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    BASKET_TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    
    scheduled_slots = {'lectures': [], 'tutorials': []}
    scheduled_lecture_days = set()
    
    # Prefer large rooms, but allow using other lecture rooms so we can allocate N unique rooms
    if lecture_rooms is None:
        lecture_rooms = []
    if large_rooms:
        room_pool = list(dict.fromkeys(large_rooms + lecture_rooms))  # preserves order, unique
    else:
        room_pool = list(dict.fromkeys(lecture_rooms))
    
    basket_faculty = []
    n_value = 1
    electives_list = []
    electives_counts = []  # New: store count values for each elective
    
    if electives_data and semester in electives_data and label in electives_data[semester]:
        basket_faculty = electives_data[semester][label].get('faculty', [])
        electives_list = electives_data[semester][label].get('electives', [])
        electives_counts = electives_data[semester][label].get('counts', [60] * len(electives_list))  # Default to 60 if not specified
        n_value = electives_data[semester][label].get('n_value', 1)
    
    print(f"Basket {label} has {len(electives_list)} electives: {electives_list}")
    print(f"Using N={n_value} rooms per slot")
    print(f"Elective counts: {electives_counts}")
    
    # Schedule lectures
    for lec_idx in range(lec_count):
        best_slot = None
        best_rooms = []
        best_electives = []
        
        for attempt in range(2000):
            day = random.randint(0, len(config.get("days", DEFAULT_CONFIG["days"]))-1)
            if day in scheduled_lecture_days:
                continue
                
            starts = get_all_possible_start_indices_for_duration('LEC', TIME_SLOTS)
            
            for start_idx in starts:
                valid_for_all = True
                slot_indices = []
                
                for dept in departments:
                    if dept not in all_department_timetables:
                        continue
                    if semester not in all_department_timetables[dept]:
                        continue
                        
                    for section_key in all_department_timetables[dept][semester]:
                        timetable = all_department_timetables[dept][semester][section_key]
                        
                        accumulated = 0
                        i = start_idx
                        temp_slot_indices = []
                        
                        while i < len(TIME_SLOTS) and accumulated < BASKET_LECTURE_MIN:
                            if is_minor_slot(TIME_SLOTS[i]):
                                valid_for_all = False
                                break
                            if is_break_time_slot(TIME_SLOTS[i], semester, 'LEC', config):
                                valid_for_all = False
                                break
                            if timetable[day][i]['type'] is not None:
                                valid_for_all = False
                                break
                            temp_slot_indices.append(i)
                            accumulated += slot_minutes(TIME_SLOTS[i])
                            i += 1
                        
                        if not valid_for_all or accumulated < BASKET_LECTURE_MIN:
                            valid_for_all = False
                            break
                        
                        if not slot_indices:
                            slot_indices = temp_slot_indices
                
                if valid_for_all and slot_indices:
                    if check_basket_slot_conflict(day, slot_indices, existing_basket_slots):
                        valid_for_all = False
                
                if valid_for_all and slot_indices and electives_list:
                    # Get rooms that are NOT already used in this time slot
                    # Collect all rooms already used in this time slot
                    used_rooms_in_slot = set()
                    for existing_slot in existing_basket_slots:
                        if existing_slot['day'] == day:
                            used_rooms_in_slot.update(existing_slot.get('rooms', []))
                    
                    # Get available rooms (not used in this slot)
                    available_rooms = [r for r in room_pool if r not in used_rooms_in_slot]
                    random.shuffle(available_rooms)
                    
                    # Assign rooms based on elective counts
                    temp_rooms = []
                    temp_electives = []
                    
                    # Sort electives by count (descending) to prioritize larger classes
                    elective_data = list(zip(electives_list, electives_counts))
                    elective_data.sort(key=lambda x: x[1], reverse=True)
                    
                    for elective, count in elective_data:
                        # Find a suitable room based on count
                        suitable_room = None
                        if count > 120:
                            # Try to find a large room
                            for room in available_rooms:
                                if room in large_rooms:
                                    suitable_room = room
                                    break
                        elif count > 60:
                            # Try to find a medium room
                            for room in available_rooms:
                                if room in medium_rooms:
                                    suitable_room = room
                                    break
                        else:
                            # Any room will do
                            suitable_room = available_rooms[0] if available_rooms else None
                        
                        if suitable_room:
                            temp_rooms.append(suitable_room)
                            temp_electives.append(elective)
                            available_rooms.remove(suitable_room)
                    
                    # Check if all rooms are available
                    rooms_available = True
                    for room in temp_rooms:
                        if room not in room_schedule:
                            room_schedule[room] = {d: set() for d in range(len(config.get("days", DEFAULT_CONFIG["days"])))}
                        if any(si in room_schedule[room][day] for si in slot_indices):
                            rooms_available = False
                            break
                    
                    if rooms_available and temp_rooms:
                        best_slot = (day, slot_indices)
                        best_rooms = temp_rooms
                        best_electives = temp_electives
                
                if best_slot:
                    break
            if best_slot:
                break
        
        if best_slot:
            scheduled_slots['lectures'].append({
                'day': best_slot[0], 
                'slots': best_slot[1], 
                'rooms': best_rooms,
                'electives': best_electives
            })
            scheduled_lecture_days.add(best_slot[0])
            
            # Mark all rooms as occupied
            for room in best_rooms:
                for si in best_slot[1]:
                    room_schedule[room][best_slot[0]].add(si)
            
            # Mark faculty as occupied
            for i, faculty in enumerate(basket_faculty[:len(best_electives)]):
                if faculty not in professor_schedule:
                    professor_schedule[faculty] = {d: set() for d in range(len(config.get("days", DEFAULT_CONFIG["days"])))}
                for si in best_slot[1]:
                    professor_schedule[faculty][best_slot[0]].add(si)
            
            print(f"Scheduled {label} Lecture {lec_idx + 1} with {len(best_electives)} electives on day {best_slot[0]}")
            for i, elective in enumerate(best_electives):
                faculty = basket_faculty[i] if i < len(basket_faculty) else "TBD"
                count = electives_counts[electives_list.index(elective)] if i < len(electives_counts) else 60
                print(f"  - {elective} (Count: {count}) taught by {faculty} in room {best_rooms[i]}")
        else:
            print(f"WARNING: Could not schedule {label} Lecture {lec_idx + 1}")
    
    # Schedule tutorials (similar logic)
    for tut_idx in range(tut_count):
        best_slot = None
        best_rooms = []
        best_electives = []
        
        for attempt in range(2000):
            day = random.randint(0, len(config.get("days", DEFAULT_CONFIG["days"]))-1)
            if day in scheduled_lecture_days:
                continue
                
            starts = get_all_possible_start_indices_for_duration('TUT', TIME_SLOTS)
            
            for start_idx in starts:
                valid_for_all = True
                slot_indices = []
                
                for dept in departments:
                    if dept not in all_department_timetables:
                        continue
                    if semester not in all_department_timetables[dept]:
                        continue
                        
                    for section_key in all_department_timetables[dept][semester]:
                        timetable = all_department_timetables[dept][semester][section_key]
                        
                        accumulated = 0
                        i = start_idx
                        temp_slot_indices = []
                        
                        while i < len(TIME_SLOTS) and accumulated < BASKET_TUTORIAL_MIN:
                            if is_minor_slot(TIME_SLOTS[i]):
                                valid_for_all = False
                                break
                            if is_break_time_slot(TIME_SLOTS[i], semester, 'TUT', config):
                                valid_for_all = False
                                break
                            if timetable[day][i]['type'] is not None:
                                valid_for_all = False
                                break
                            temp_slot_indices.append(i)
                            accumulated += slot_minutes(TIME_SLOTS[i])
                            i += 1
                        
                        if not valid_for_all or accumulated < BASKET_TUTORIAL_MIN:
                            valid_for_all = False
                            break
                        
                        if not slot_indices:
                            slot_indices = temp_slot_indices
                
                if valid_for_all and slot_indices:
                    if check_basket_slot_conflict(day, slot_indices, existing_basket_slots):
                        valid_for_all = False
                
                if valid_for_all and slot_indices and electives_list:
                    # Get rooms that are NOT already used in this time slot
                    # Collect all rooms already used in this time slot
                    used_rooms_in_slot = set()
                    for existing_slot in existing_basket_slots:
                        if existing_slot['day'] == day:
                            used_rooms_in_slot.update(existing_slot.get('rooms', []))
                    
                    # Get available rooms (not used in this slot)
                    available_rooms = [r for r in room_pool if r not in used_rooms_in_slot]
                    random.shuffle(available_rooms)
                    
                    # Assign rooms based on elective counts
                    temp_rooms = []
                    temp_electives = []
                    
                    # Sort electives by count (descending) to prioritize larger classes
                    elective_data = list(zip(electives_list, electives_counts))
                    elective_data.sort(key=lambda x: x[1], reverse=True)
                    
                    for elective, count in elective_data:
                        # Find a suitable room based on count
                        suitable_room = None
                        if count > 120:
                            # Try to find a large room
                            for room in available_rooms:
                                if room in large_rooms:
                                    suitable_room = room
                                    break
                        elif count > 60:
                            # Try to find a medium room
                            for room in available_rooms:
                                if room in medium_rooms:
                                    suitable_room = room
                                    break
                        else:
                            # Any room will do
                            suitable_room = available_rooms[0] if available_rooms else None
                        
                        if suitable_room:
                            temp_rooms.append(suitable_room)
                            temp_electives.append(elective)
                            available_rooms.remove(suitable_room)
                    
                    # Check if all rooms are available
                    rooms_available = True
                    for room in temp_rooms:
                        if room not in room_schedule:
                            room_schedule[room] = {d: set() for d in range(len(config.get("days", DEFAULT_CONFIG["days"])))}
                        if any(si in room_schedule[room][day] for si in slot_indices):
                            rooms_available = False
                            break
                    
                    if rooms_available and temp_rooms:
                        best_slot = (day, slot_indices)
                        best_rooms = temp_rooms
                        best_electives = temp_electives
                
                if best_slot:
                    break
            if best_slot:
                    break
        
        if best_slot:
            scheduled_slots['tutorials'].append({
                'day': best_slot[0], 
                'slots': best_slot[1], 
                'rooms': best_rooms,
                'electives': best_electives
            })
            
            # Mark all rooms as occupied
            for room in best_rooms:
                for si in best_slot[1]:
                    room_schedule[room][best_slot[0]].add(si)
            
            # Mark faculty as occupied
            for i, faculty in enumerate(basket_faculty[:len(best_electives)]):
                if faculty not in professor_schedule:
                    professor_schedule[faculty] = {d: set() for d in range(len(config.get("days", DEFAULT_CONFIG["days"])))}
                for si in best_slot[1]:
                    professor_schedule[faculty][best_slot[0]].add(si)
            
            print(f"Scheduled {label} Tutorial {tut_idx + 1} with {len(best_electives)} electives on day {best_slot[0]}")
            for i, elective in enumerate(best_electives):
                faculty = basket_faculty[i] if i < len(basket_faculty) else "TBD"
                count = electives_counts[electives_list.index(elective)] if i < len(electives_counts) else 60
                print(f"  - {elective} (Count: {count}) taught by {faculty} in room {best_rooms[i]}")
        else:
            print(f"WARNING: Could not schedule {label} Tutorial {tut_idx + 1}")
    
    return scheduled_slots

# ---------------------------
# Write electives to output
# ---------------------------
def write_electives_to_output(all_department_timetables, output_filename, electives_data):
    if not electives_data:
        return
    
    electives_wb = Workbook()
    if "Sheet" in electives_wb.sheetnames:
        electives_wb.remove(electives_wb["Sheet"])
    
    departments = set()
    for semester_data in all_department_timetables.values():
        departments.update(semester_data.keys())
    
    if electives_data:
        for semester in electives_data:
            for basket in electives_data[semester]:
                config = get_basket_config_for_semester(semester, electives_data)
                if config and 'departments' in config:
                    departments.update(config['departments'])
    
    for department in departments:
        ws = electives_wb.create_sheet(title=str(department))
        
        headers = ["Semester", "Basket", "Electives", "Faculty", "N (Rooms per slot)", "Count"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_idx = 2
        
        for semester in sorted(electives_data.keys()):
            for basket_label, basket_data in electives_data[semester].items():
                ws.cell(row=row_idx, column=1, value=str(semester))
                ws.cell(row=row_idx, column=2, value=basket_label)
                
                electives_list = basket_data.get('electives', [])
                electives_str = ", ".join(electives_list)
                ws.cell(row=row_idx, column=3, value=electives_str)
                
                faculty_list = basket_data.get('faculty', [])
                faculty_str = ", ".join(faculty_list) if faculty_list else ""
                ws.cell(row=row_idx, column=4, value=faculty_str)
                
                n_value = basket_data.get('n_value', 1)
                ws.cell(row=row_idx, column=5, value=n_value)
                
                counts_list = basket_data.get('counts', [])
                counts_str = ", ".join(str(c) for c in counts_list) if counts_list else ""
                ws.cell(row=row_idx, column=6, value=counts_str)
                
                for col in range(1, 7):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                row_idx += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
    
    electives_output_path = os.path.join(OUTPUT_DIR, "electives_output.xlsx")
    electives_wb.save(electives_output_path)
    print(f"Electives information saved to {electives_output_path}")

# ---------------------------
# --- NEW FUNCTION ---
# Write a sheet containing ONLY the basket/elective schedule
# ---------------------------
def write_basket_only_sheet(ws, timetable, semester, electives_data, config, TIME_SLOTS):
    DAYS = config.get("days", DEFAULT_CONFIG["days"])
    BASKET_LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    BASKET_TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    
    # --- Header and Formatting Setup ---
    header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
    ws.append(header)
    
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    basket_fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
    break_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    minor_fill = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid")
    
    # --- Build a map of only the basket slots ---
    basket_slot_mapping = {}
    for day_idx, day_name in enumerate(DAYS):
        for slot_idx in range(len(TIME_SLOTS)):
            # We only care about slots that have a basket course (identified by the name 'Course')
            if timetable[day_idx][slot_idx]['type'] is not None and 'Course' in timetable[day_idx][slot_idx]['name']:
                key = (day_idx, slot_idx)
                if key not in basket_slot_mapping:
                    basket_slot_mapping[key] = {
                        'type': timetable[day_idx][slot_idx]['type'],
                        'electives': [],
                        'faculty': timetable[day_idx][slot_idx]['faculty']
                    }
                
                # FIX: Remove duplicates in electives for display
                # Only add each elective once per time slot, regardless of how many departments it appears in
                if 'electives' in timetable[day_idx][slot_idx] and timetable[day_idx][slot_idx]['electives']:
                    for elective in timetable[day_idx][slot_idx]['electives']:
                        # Create a unique identifier for each elective (code + room)
                        elective_id = f"{elective['code']}_{elective['room']}"
                        # Check if this elective is already in our list for this time slot
                        if not any(e['code'] == elective['code'] and e['room'] == elective['room'] 
                                 for e in basket_slot_mapping[key]['electives']):
                            basket_slot_mapping[key]['electives'].append({
                                'code': elective['code'],
                                'room': elective['room'],
                                'faculty': elective['faculty'],
                                'count': elective.get('count', 60)  # Default to 60 if not specified
                            })

    # --- Write the main grid, but only for basket slots ---
    for day_idx, day_name in enumerate(DAYS):
        ws.append([day_name] + [''] * len(TIME_SLOTS))
        row_num = ws.max_row
        merges = []
        
        for slot_idx in range(len(TIME_SLOTS)):
            cell_obj = ws.cell(row=row_num, column=slot_idx + 2)
            
            if is_minor_slot(TIME_SLOTS[slot_idx]):
                cell_obj.value = "Minor Slot"
                cell_obj.fill = minor_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue

            if is_break_time_slot(TIME_SLOTS[slot_idx], semester, config=config):
                cell_obj.value = "BREAK"
                cell_obj.fill = break_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue

            # Check if this slot is part of the basket schedule
            key = (day_idx, slot_idx)
            if key not in basket_slot_mapping:
                # It's not a basket slot, leave it empty but bordered
                cell_obj.border = border
                continue

            # It IS a basket slot, now write it out.
            basket_info = basket_slot_mapping[key]
            display_parts = []
            
            # Remove duplicates in electives for display
            seen_electives = set()
            unique_electives = []
            for elective in basket_info['electives']:
                elective_id = f"{elective['code']}_{elective['room']}"
                if elective_id not in seen_electives:
                    seen_electives.add(elective_id)
                    unique_electives.append(elective)
            
            for elective in unique_electives:
                count = elective.get('count', 60)
                display_parts.append(f"{elective['code']} (Count: {count})\nRoom: {elective['room']}")
            
            fac = basket_info['faculty']
            display = '\n'.join(display_parts)
            display += f"\n{basket_info['type']}\n{fac}"
            
            # Find the full span of this block
            span = [slot_idx]
            j = slot_idx + 1
            while (j < len(TIME_SLOTS) and 
                   (day_idx, j) in basket_slot_mapping and 
                   basket_slot_mapping[(day_idx, j)]['type'] == basket_info['type']):
                span.append(j)
                j += 1
            
            cell_obj.value = display
            cell_obj.fill = basket_fill
            merges.append((slot_idx + 2, slot_idx + 2 + len(span) - 1, display, basket_fill))
            
            cell_obj.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell_obj.border = border

        # Merge cells for continuous blocks
        for start_col, end_col, val, fill in merges:
            if end_col > start_col:
                rng = f"{get_column_letter(start_col)}{row_num}:{get_column_letter(end_col)}{row_num}"
                try:
                    ws.merge_cells(rng)
                    mc = ws[f"{get_column_letter(start_col)}{row_num}"]
                    mc.value = val
                    mc.fill = fill
                    mc.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    mc.border = border
                except:
                    pass

    # --- Column and Row Formatting ---
    for col_idx in range(1, len(TIME_SLOTS)+2):
        try:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        except:
            pass
    for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
        ws.row_dimensions[row[0].row].height = 40

    # --- Add Basket Information Section ---
    current_row = len(DAYS) + 4
    ws.cell(row=current_row, column=1, value="Cross-Department Elective Information")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 2

    if electives_data and semester in electives_data:
        if 'B1' in electives_data[semester] and 'B2' in electives_data[semester] and 'B3' in electives_data[semester] and 'B4' in electives_data[semester]:
            for basket_label in ['B1', 'B2', 'B3', 'B4']:
                ws.cell(row=current_row, column=1, value=f"{basket_label} Electives:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1
                electives_list = electives_data[semester][basket_label].get('electives', [])
                counts_list = electives_data[semester][basket_label].get('counts', [60] * len(electives_list))
                
                # Display electives with their counts
                electives_with_counts = []
                for i, elective in enumerate(electives_list):
                    count = counts_list[i] if i < len(counts_list) else 60
                    electives_with_counts.append(f"{elective} (Count: {count})")
                
                electives_str = ", ".join(electives_with_counts)
                ws.cell(row=current_row, column=1, value=electives_str)
                current_row += 1
                
                faculty_list = electives_data[semester][basket_label].get('faculty', [])
                faculty_str = ", ".join(faculty_list) if faculty_list else "TBD"
                ws.cell(row=current_row, column=1, value=f"Faculty: {faculty_str}")
                current_row += 1
                n_value = electives_data[semester][basket_label].get('n_value', 1)
                ws.cell(row=current_row, column=1, value=f"Rooms per slot: {n_value}")
                current_row += 1
                ws.cell(row=current_row, column=1, value=f"Lectures: 2 ({BASKET_LECTURE_MIN} min each)")
                ws.cell(row=current_row, column=2, value=f"Tutorials: 1 ({BASKET_TUTORIAL_MIN} min each)")
                current_row += 1
                ws.cell(row=current_row, column=1, value="Shared across: CSE, DSAI, ECE")
                current_row += 2
        elif 'B1' in electives_data[semester] and 'B2' in electives_data[semester]:
            for basket_label in ['B1', 'B2']:
                ws.cell(row=current_row, column=1, value=f"{basket_label} Electives:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1
                electives_list = electives_data[semester][basket_label].get('electives', [])
                counts_list = electives_data[semester][basket_label].get('counts', [60] * len(electives_list))
                
                # Display electives with their counts
                electives_with_counts = []
                for i, elective in enumerate(electives_list):
                    count = counts_list[i] if i < len(counts_list) else 60
                    electives_with_counts.append(f"{elective} (Count: {count})")
                
                electives_str = ", ".join(electives_with_counts)
                ws.cell(row=current_row, column=1, value=electives_str)
                current_row += 1
                
                faculty_list = electives_data[semester][basket_label].get('faculty', [])
                faculty_str = ", ".join(faculty_list) if faculty_list else "TBD"
                ws.cell(row=current_row, column=1, value=f"Faculty: {faculty_str}")
                current_row += 1
                n_value = electives_data[semester][basket_label].get('n_value', 1)
                ws.cell(row=current_row, column=1, value=f"Rooms per slot: {n_value}")
                current_row += 1
                ws.cell(row=current_row, column=1, value=f"Lectures: 2 ({BASKET_LECTURE_MIN} min each)")
                ws.cell(row=current_row, column=2, value=f"Tutorials: 1 ({BASKET_TUTORIAL_MIN} min each)")
                current_row += 1
                ws.cell(row=current_row, column=1, value="Shared across: CSE, DSAI, ECE")
                current_row += 2
        else:
            ws.cell(row=current_row, column=1, value="ELECTIVE slots shared across CSE, DSAI, ECE")
            current_row += 1
            for basket_label, basket_data in electives_data[semester].items():
                electives_list = basket_data.get('electives', [])
                counts_list = basket_data.get('counts', [60] * len(electives_list))
                
                # Display electives with their counts
                electives_with_counts = []
                for i, elective in enumerate(electives_list):
                    count = counts_list[i] if i < len(counts_list) else 60
                    electives_with_counts.append(f"{elective} (Count: {count})")
                
                electives_str = ", ".join(electives_with_counts)
                ws.cell(row=current_row, column=1, value=electives_str)
                current_row += 1
                
                faculty_list = basket_data.get('faculty', [])
                faculty_str = ", ".join(faculty_list) if faculty_list else "TBD"
                ws.cell(row=current_row, column=1, value=f"Faculty: {faculty_str}")
                current_row += 1
                n_value = basket_data.get('n_value', 1)
                ws.cell(row=current_row, column=1, value=f"Rooms per slot: {n_value}")
                current_row += 1
            ws.cell(row=current_row, column=1, value=f"Lectures: 2 ({BASKET_LECTURE_MIN} min each)")
            ws.cell(row=current_row, column=2, value=f"Tutorials: 1 ({BASKET_TUTORIAL_MIN} min each)")
            current_row += 2

        # --- Scheduled Elective Details ---
        ws.cell(row=current_row, column=1, value="Scheduled Elective Details")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1

        headers = ["Elective", "Count", "Type", "Day", "Time", "Room"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = border
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        current_row += 1

        scheduled_electives_list = []
        for (day_idx, slot_idx), basket_info in basket_slot_mapping.items():
            day_name = DAYS[day_idx]
            start_time = TIME_SLOTS[slot_idx][0].strftime('%H:%M')
            end_slot_idx = slot_idx
            while (end_slot_idx + 1 < len(TIME_SLOTS) and 
                   (day_idx, end_slot_idx + 1) in basket_slot_mapping and 
                   basket_slot_mapping[(day_idx, end_slot_idx + 1)]['type'] == basket_info['type']):
                end_slot_idx += 1
            end_time = TIME_SLOTS[end_slot_idx][1].strftime('%H:%M')
            time_str = f"{start_time} - {end_time}"

            seen_electives = set()
            unique_electives = []
            for elective in basket_info['electives']:
                elective_id = f"{elective['code']}_{elective['room']}"
                if elective_id not in seen_electives:
                    seen_electives.add(elective_id)
                    unique_electives.append(elective)

            for elective in unique_electives:
                scheduled_electives_list.append({
                    'Elective': elective['code'],
                    'Count': elective.get('count', 60),
                    'Type': basket_info['type'],
                    'Day': day_name,
                    'Time': time_str,
                    'Room': elective['room']
                })

        scheduled_electives_list.sort(key=lambda x: (DAYS.index(x['Day']), x['Time']))

        for elective_info in scheduled_electives_list:
            ws.cell(row=current_row, column=1, value=elective_info['Elective'])
            ws.cell(row=current_row, column=2, value=elective_info['Count'])
            ws.cell(row=current_row, column=3, value=elective_info['Type'])
            ws.cell(row=current_row, column=4, value=elective_info['Day'])
            ws.cell(row=current_row, column=5, value=elective_info['Time'])
            ws.cell(row=current_row, column=6, value=elective_info['Room'])
            
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            current_row += 1
        
        if not scheduled_electives_list:
            ws.cell(row=current_row, column=1, value="No electives were scheduled.")
            current_row += 1

# ---------------------------
# --- NEW FUNCTION ---
# Write a sheet containing free room information for each time slot
# ---------------------------
def write_free_room_sheet(wb, room_schedule, config, TIME_SLOTS, lecture_rooms, computer_lab_rooms, auditorium_rooms):
    DAYS = config.get("days", DEFAULT_CONFIG["days"])
    
    # Create a new sheet for free rooms
    ws = wb.create_sheet(title="Free Rooms")
    
    # Set up the header
    header = ['Day/Time'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
    ws.append(header)
    
    # Style the header
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Define cell styles
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    free_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    busy_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
    break_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")  # Gray
    
    # Get all rooms
    all_rooms = lecture_rooms + computer_lab_rooms + auditorium_rooms
    
    # For each day, create a row for each room
    for day_idx, day_name in enumerate(DAYS):
        # Add a row header for the day
        ws.append([day_name] + [''] * len(TIME_SLOTS))
        day_row = ws.max_row
        ws.cell(row=day_row, column=1).font = Font(bold=True)
        
        # Add a row for each room
        for room in all_rooms:
            ws.append([room] + [''] * len(TIME_SLOTS))
            room_row = ws.max_row
            
            for slot_idx in range(len(TIME_SLOTS)):
                cell = ws.cell(row=room_row, column=slot_idx + 2)
                
                # Check if this is a break time
                if is_break_time_slot(TIME_SLOTS[slot_idx], config=config):
                    cell.value = "BREAK"
                    cell.fill = break_fill
                # Check if this room is occupied at this time slot
                elif room in room_schedule and slot_idx in room_schedule[room][day_idx]:
                    cell.value = "OCCUPIED"
                    cell.fill = busy_fill
                else:
                    cell.value = "FREE"
                    cell.fill = free_fill
                
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add an empty row for spacing
        ws.append([''] * (len(TIME_SLOTS) + 1))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    for col_idx in range(2, len(TIME_SLOTS) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Add a legend
    current_row = ws.max_row + 2
    ws.cell(row=current_row, column=1, value="Legend:")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    legend_items = [
        ("FREE", free_fill, "Room is available"),
        ("OCCUPIED", busy_fill, "Room is in use"),
        ("BREAK", break_fill, "Break time")
    ]
    
    for status, fill, description in legend_items:
        ws.cell(row=current_row, column=1, value=status)
        ws.cell(row=current_row, column=1).fill = fill
        ws.cell(row=current_row, column=2, value=description)
        current_row += 1

# ---------------------------
# Main generation function
# ---------------------------
def generate_all_timetables():
    config = load_configuration()
    
    DAYS = config.get("days", DEFAULT_CONFIG["days"])
    LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    LAB_MIN = config.get("LAB_MIN", DEFAULT_CONFIG["LAB_MIN"])
    TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    SELF_STUDY_MIN = config.get("SELF_STUDY_MIN", DEFAULT_CONFIG["SELF_STUDY_MIN"])
    
    BASKET_LECTURE_MIN = LECTURE_MIN
    BASKET_TUTORIAL_MIN = TUTORIAL_MIN
    
    MORNING_BREAK_START = parse_time_string(config.get("MORNING_BREAK_START", "10:30"))
    MORNING_BREAK_END = parse_time_string(config.get("MORNING_BREAK_END", "10:45"))
    LUNCH_BREAK_START = parse_time_string(config.get("LUNCH_BREAK_START", "13:00"))
    LUNCH_BREAK_END = parse_time_string(config.get("LUNCH_BREAK_END", "13:45"))
    LECTURE_TUTORIAL_BREAK_START = parse_time_string(config.get("LECTURE_TUTORIAL_BREAK_START", "15:30"))
    LECTURE_TUTORIAL_BREAK_END = parse_time_string(config.get("LECTURE_TUTORIAL_BREAK_END", "15:40"))
    
    print("\n" + "="*60)
    print("LOADED CONFIGURATION")
    print("="*60)
    print(f"Lecture Duration: {LECTURE_MIN} minutes")
    print(f"Lab Duration: {LAB_MIN} minutes")
    print(f"Tutorial Duration: {TUTORIAL_MIN} minutes")
    print(f"Self-Study Duration: {SELF_STUDY_MIN} minutes")
    print(f"Morning Break: {MORNING_BREAK_START.strftime('%H:%M')} - {MORNING_BREAK_END.strftime('%H:%M')}")
    print(f"Lunch Break: {LUNCH_BREAK_START.strftime('%H:%M')} - {LUNCH_BREAK_END.strftime('%H:%M')}")
    print(f"Evening Break: {LECTURE_TUTORIAL_BREAK_START.strftime('%H:%M')} - {LECTURE_TUTORIAL_BREAK_END.strftime('%H:%M')}")
    print("="*60 + "\n")
    
    TIME_SLOTS = generate_time_slots(config)
    
    df, lecture_rooms, computer_lab_rooms, large_rooms, auditorium_rooms, small_rooms, medium_rooms = load_data()
    electives_data = load_electives()
    
    room_schedule = {}
    professor_schedule = {}
    course_room_mapping = {}
    unscheduled_components = []

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    overview = wb.create_sheet("Overview")
    overview.append(["Combined Timetable for All Departments and Semesters"])
    overview.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    overview.append([])
    overview.append(["Department", "Semester", "Sheet Name"])
    row_index = 5

    SUBJECT_COLORS = [
        "FF6B6B", "4ECDC4", "FF9F1C", "5D5FEF", "45B7D1",
        "F72585", "7209B7", "3A0CA3", "4361EE", "4CC9F0",
        "06D6A0", "FFD166", "EF476F", "118AB2", "073B4C"
    ]

    all_department_timetables = {}
    
    # Initialize all timetables
    for department in df['Department'].unique():
        all_department_timetables[department] = {}
        sems = sorted(df[df['Department'] == department]['Semester'].unique())
        
        for semester in sems:
            dept_upper = str(department).strip().upper()
            num_sections = 2 if (dept_upper == "CSE" and int(semester) in [1, 3, 5]) else 1
            
            all_department_timetables[department][semester] = {}
            
            for section in range(num_sections):
                section_key = chr(65 + section) if num_sections > 1 else 'A'
                all_department_timetables[department][semester][section_key] = {
                    d: {s: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': '', 'electives': [], 'lab_rooms': []} 
                        for s in range(len(TIME_SLOTS))} 
                    for d in range(len(DAYS))
                }
    
    # Schedule basket slots
    for semester in [1, 3, 5, 7]:
        basket_config = get_basket_config_for_semester(semester, electives_data)
        if basket_config:
            print(f"\n=== Scheduling basket slots for Semester {semester} ===")
            basket_slots = schedule_basket_slots(
                semester, all_department_timetables, professor_schedule, 
                room_schedule, course_room_mapping, config, TIME_SLOTS, electives_data, lecture_rooms, large_rooms, small_rooms, medium_rooms
            )
            
            departments = basket_config['departments']
            for dept in departments:
                if dept not in all_department_timetables or semester not in all_department_timetables[dept]:
                    continue
                    
                for section_key in all_department_timetables[dept][semester]:
                    timetable = all_department_timetables[dept][semester][section_key]
                    
                    if isinstance(basket_slots, dict) and 'B1' in basket_slots:
                        if semester == 5:
                            basket_labels = ['B1', 'B2']
                        elif semester == 7:
                            basket_labels = ['B1', 'B2', 'B3', 'B4']
                        else:
                            basket_labels = ['B1', 'B2']
                            
                        for basket_label in basket_labels:
                            basket_data = basket_slots.get(basket_label, {})
                            print(f"Applying {basket_label} slots to {dept} Section {section_key}")
                            
                            basket_faculty = []
                            electives_counts = []
                            if electives_data and semester in electives_data and basket_label in electives_data[semester]:
                                basket_faculty = electives_data[semester][basket_label].get('faculty', [])
                                electives_counts = electives_data[semester][basket_label].get('counts', [60] * len(basket_faculty))
                            faculty_str = ", ".join(basket_faculty) if basket_faculty else "Multiple Faculty"
                            
                            for lec_info in basket_data.get('lectures', []):
                                day = lec_info['day']
                                slots = lec_info['slots']
                                rooms = lec_info.get('rooms', [])
                                electives = lec_info.get('electives', [])
                                
                                for si in slots:
                                    if timetable[day][si]['type'] is None:
                                        timetable[day][si]['type'] = 'LEC'
                                        timetable[day][si]['name'] = f"{basket_label} Course"
                                        timetable[day][si]['faculty'] = faculty_str
                                        timetable[day][si]['electives'] = []
                                        
                                        for i, elective in enumerate(electives):
                                            room = rooms[i] if i < len(rooms) else ''
                                            faculty = basket_faculty[i] if i < len(basket_faculty) else "TBD"
                                            count = electives_counts[i] if i < len(electives_counts) else 60
                                            
                                            timetable[day][si]['electives'].append({
                                                'code': elective,
                                                'room': room,
                                                'faculty': faculty,
                                                'count': count
                                            })
                            
                            for tut_info in basket_data.get('tutorials', []):
                                day = tut_info['day']
                                slots = tut_info['slots']
                                rooms = tut_info.get('rooms', [])
                                electives = tut_info.get('electives', [])
                                
                                for si in slots:
                                    if timetable[day][si]['type'] is None:
                                        timetable[day][si]['type'] = 'TUT'
                                        timetable[day][si]['name'] = f"{basket_label} Course"
                                        timetable[day][si]['faculty'] = faculty_str
                                        timetable[day][si]['electives'] = []
                                        
                                        for i, elective in enumerate(electives):
                                            room = rooms[i] if i < len(rooms) else ''
                                            faculty = basket_faculty[i] if i < len(basket_faculty) else "TBD"
                                            count = electives_counts[i] if i < len(electives_counts) else 60
                                            
                                            timetable[day][si]['electives'].append({
                                                'code': elective,
                                                'room': room,
                                                'faculty': faculty,
                                                'count': count
                                            })
                    else:
                        basket_label = 'ELECTIVE'
                        basket_data = basket_slots
                        
                        basket_faculty = []
                        electives_counts = []
                        if electives_data and semester in electives_data and basket_label in electives_data[semester]:
                            basket_faculty = electives_data[semester][basket_label].get('faculty', [])
                            electives_counts = electives_data[semester][basket_label].get('counts', [60] * len(basket_faculty))
                            faculty_str = ", ".join(basket_faculty) if basket_faculty else "Multiple Faculty"
                            
                        for lec_info in basket_data.get('lectures', []):
                            day = lec_info['day']
                            slots = lec_info['slots']
                            rooms = lec_info.get('rooms', [])
                            electives = lec_info.get('electives', [])
                            
                            for si in slots:
                                if timetable[day][si]['type'] is None:
                                    timetable[day][si]['type'] = 'LEC'
                                    timetable[day][si]['name'] = f"{basket_label} Course"
                                    timetable[day][si]['faculty'] = faculty_str
                                    timetable[day][si]['electives'] = []
                                    
                                    for i, elective in enumerate(electives):
                                        room = rooms[i] if i < len(rooms) else ''
                                        faculty = basket_faculty[i] if i < len(basket_faculty) else "TBD"
                                        count = electives_counts[i] if i < len(electives_counts) else 60
                                        
                                        timetable[day][si]['electives'].append({
                                            'code': elective,
                                            'room': room,
                                            'faculty': faculty,
                                            'count': count
                                        })
                        
                        for tut_info in basket_data.get('tutorials', []):
                            day = tut_info['day']
                            slots = tut_info['slots']
                            rooms = tut_info.get('rooms', [])
                            electives = tut_info.get('electives', [])
                            
                            for si in slots:
                                if timetable[day][si]['type'] is None:
                                    timetable[day][si]['type'] = 'TUT'
                                    timetable[day][si]['name'] = f"{basket_label} Course"
                                    timetable[day][si]['faculty'] = faculty_str
                                    timetable[day][si]['electives'] = []
                                    
                                    for i, elective in enumerate(electives):
                                        room = rooms[i] if i < len(rooms) else ''
                                        faculty = basket_faculty[i] if i < len(basket_faculty) else "TBD"
                                        count = electives_counts[i] if i < len(electives_counts) else 60
                                        
                                        timetable[day][si]['electives'].append({
                                            'code': elective,
                                            'room': room,
                                            'faculty': faculty,
                                            'count': count
                                        })
    
    # Schedule auditorium courses
    auditorium_courses_map = {}
    for _, course in df.iterrows():
        if is_auditorium_course(course):
            code = str(course.get('Course Code', '')).strip()
            name = str(course.get('Course Name', '')).strip()
            dept = str(course.get('Department', '')).strip()
            sem = int(course.get('Semester', 0))
            
            if code not in auditorium_courses_map:
                auditorium_courses_map[code] = {
                    'name': name,
                    'departments': set(),
                    'semester': sem,
                    'course_data': course
                }
            auditorium_courses_map[code]['departments'].add(dept)
    
    for code, course_info in auditorium_courses_map.items():
        name = course_info['name']
        semester = course_info['semester']
        departments = list(course_info['departments'])
        course_data = course_info['course_data']
        
        required_departments = []
        if 'CSE' in departments:
            if 'DSAI' in departments or 'ECE' in departments:
                required_departments = ['CSE', 'DSAI', 'ECE']
            else:
                required_departments = ['CSE']
        elif 'DSAI' in departments and 'ECE' in departments:
            required_departments = ['DSAI', 'ECE']
        else:
            required_departments = departments
        
        print(f"Scheduling auditorium course {code} for departments: {required_departments}")
        
        lec_sessions, tut_sessions, lab_sessions, ss_sessions, lab_duration = calculate_required_sessions(course_data, config)
        print(f"Auditorium course {code}: L={lec_sessions}, T={tut_sessions}, P={lab_sessions}, S={ss_sessions}")
        
        lecture_days = set()
        tutorial_days = set()
        lab_days = set()
        
        # Schedule lectures
        for lec_idx in range(lec_sessions):
            for attempt in range(2000):
                day = random.randint(0, len(DAYS)-1)
                if day in lecture_days:
                    continue
                if day in tutorial_days:
                    continue
                if day in lab_days:
                    continue
                    
                starts = get_all_possible_start_indices_for_duration('LEC', TIME_SLOTS)
                
                for start_idx in starts:
                    valid_for_all = True
                    slot_indices = []
                    
                    for dept in required_departments:
                        if dept not in all_department_timetables:
                            continue
                        if semester not in all_department_timetables[dept]:
                            continue
                            
                        for section_key in all_department_timetables[dept][semester]:
                            timetable = all_department_timetables[dept][semester][section_key]
                            
                            accumulated = 0
                            i = start_idx
                            temp_slot_indices = []
                            
                            while i < len(TIME_SLOTS) and accumulated < LECTURE_MIN:
                                if is_minor_slot(TIME_SLOTS[i]):
                                    valid_for_all = False
                                    break
                                if is_break_time_slot(TIME_SLOTS[i], semester, 'LEC', config):
                                    valid_for_all = False
                                    break
                                if timetable[day][i]['type'] is not None:
                                    valid_for_all = False
                                    break
                                temp_slot_indices.append(i)
                                accumulated += slot_minutes(TIME_SLOTS[i])
                                i += 1
                            
                            if not valid_for_all or accumulated < LECTURE_MIN:
                                valid_for_all = False
                                break
                            
                            if not slot_indices:
                                slot_indices = temp_slot_indices
                    
                    if not valid_for_all or not slot_indices:
                        continue
                    
                    if not auditorium_rooms:
                        print(f"No auditorium rooms available for {code}")
                        valid_for_all = False
                        break
                    
                    room = None
                    random.shuffle(auditorium_rooms)
                    for auditorium in auditorium_rooms:
                        if auditorium not in room_schedule:
                            room_schedule[auditorium] = {d: set() for d in range(len(DAYS))}
                        if all(si not in room_schedule[auditorium][day] for si in slot_indices):
                            room = auditorium
                            break
                    
                    if room is None:
                        valid_for_all = False
                        break
                    
                    for dept in required_departments:
                        if dept not in all_department_timetables:
                            continue
                        if semester not in all_department_timetables[dept]:
                            continue
                            
                        for section_key in all_department_timetables[dept][semester]:
                            timetable = all_department_timetables[dept][semester][section_key]
                            
                            faculty = select_faculty_for_section(course_data.get('Faculty', 'TBD'), section_key)
                            
                            for idx, si in enumerate(slot_indices):
                                timetable[day][si]['type'] = 'LEC'
                                timetable[day][si]['code'] = code if idx == 0 else ''
                                timetable[day][si]['name'] = name if idx == 0 else ''
                                timetable[day][si]['faculty'] = faculty if idx == 0 else ''
                                timetable[day][si]['classroom'] = room if idx == 0 else ''
                    
                    for si in slot_indices:
                        room_schedule[room][day].add(si)
                    
                    lecture_days.add(day)
                    
                    print(f"Scheduled auditorium course {code} lecture {lec_idx + 1} on day {day} in room {room}")
                    break
                
                if valid_for_all:
                    break
        
        # Schedule tutorials
        for tut_idx in range(tut_sessions):
            for attempt in range(2000):
                day = random.randint(0, len(DAYS)-1)
                if day in lecture_days:
                    continue
                if day in tutorial_days:
                    continue
                if day in lab_days:
                    continue
                    
                starts = get_all_possible_start_indices_for_duration('TUT', TIME_SLOTS)
                
                for start_idx in starts:
                    valid_for_all = True
                    slot_indices = []
                    
                    for dept in required_departments:
                        if dept not in all_department_timetables:
                            continue
                        if semester not in all_department_timetables[dept]:
                            continue
                            
                        for section_key in all_department_timetables[dept][semester]:
                            timetable = all_department_timetables[dept][semester][section_key]
                            
                            accumulated = 0
                            i = start_idx
                            temp_slot_indices = []
                            
                            while i < len(TIME_SLOTS) and accumulated < TUTORIAL_MIN:
                                if is_minor_slot(TIME_SLOTS[i]):
                                    valid_for_all = False
                                    break
                                if is_break_time_slot(TIME_SLOTS[i], semester, 'TUT', config):
                                    valid_for_all = False
                                    break
                                if timetable[day][i]['type'] is not None:
                                    valid_for_all = False
                                    break
                                temp_slot_indices.append(i)
                                accumulated += slot_minutes(TIME_SLOTS[i])
                                i += 1
                            
                            if not valid_for_all or accumulated < TUTORIAL_MIN:
                                valid_for_all = False
                                break
                            
                            if not slot_indices:
                                slot_indices = temp_slot_indices
                    
                    if not valid_for_all or not slot_indices:
                        continue
                    
                    if not auditorium_rooms:
                        print(f"No auditorium rooms available for {code}")
                        valid_for_all = False
                        break
                    
                    room = None
                    random.shuffle(auditorium_rooms)
                    for auditorium in auditorium_rooms:
                        if auditorium not in room_schedule:
                            room_schedule[auditorium] = {d: set() for d in range(len(DAYS))}
                        if all(si not in room_schedule[auditorium][day] for si in slot_indices):
                            room = auditorium
                            break
                    
                    if room is None:
                        valid_for_all = False
                        break
                    
                    for dept in required_departments:
                        if dept not in all_department_timetables:
                            continue
                        if semester not in all_department_timetables[dept]:
                            continue
                            
                        for section_key in all_department_timetables[dept][semester]:
                            timetable = all_department_timetables[dept][semester][section_key]
                            
                            faculty = select_faculty_for_section(course_data.get('Faculty', 'TBD'), section_key)
                            
                            for idx, si in enumerate(slot_indices):
                                timetable[day][si]['type'] = 'TUT'
                                timetable[day][si]['code'] = code if idx == 0 else ''
                                timetable[day][si]['name'] = name if idx == 0 else ''
                                timetable[day][si]['faculty'] = faculty if idx == 0 else ''
                                timetable[day][si]['classroom'] = room if idx == 0 else ''
                    
                    for si in slot_indices:
                        room_schedule[room][day].add(si)
                    
                    tutorial_days.add(day)
                    
                    print(f"Scheduled auditorium course {code} tutorial {tut_idx + 1} on day {day} in room {room}")
                    break
                
                if valid_for_all:
                    break
        
        # Schedule labs with different times for different sections
        for lab_idx in range(lab_sessions):
            for dept_idx, dept in enumerate(required_departments):
                for attempt in range(2000):
                    day = random.randint(0, len(DAYS)-1)
                    if day in lecture_days:
                        continue
                    if day in tutorial_days:
                        continue
                    # Each department gets its own lab day to avoid conflicts
                    dept_lab_day_key = f"{dept}_{day}"
                    if dept_lab_day_key in lab_days:
                        continue
                        
                    starts = get_all_possible_start_indices_for_duration('LAB', TIME_SLOTS)
                    
                    for start_idx in starts:
                        valid_for_all = True
                        slot_indices = []
                        
                        # Check all sections for this department
                        if dept not in all_department_timetables:
                            continue
                        if semester not in all_department_timetables[dept]:
                            continue
                            
                        for section_key in all_department_timetables[dept][semester]:
                            timetable = all_department_timetables[dept][semester][section_key]
                            
                            accumulated = 0
                            i = start_idx
                            temp_slot_indices = []
                            
                            while i < len(TIME_SLOTS) and accumulated < lab_duration:
                                if is_minor_slot(TIME_SLOTS[i]):
                                    valid_for_all = False
                                    break
                                if is_break_time_slot(TIME_SLOTS[i], semester, 'LAB', config):
                                    valid_for_all = False
                                    break
                                if timetable[day][i]['type'] is not None:
                                    valid_for_all = False
                                    break
                                temp_slot_indices.append(i)
                                accumulated += slot_minutes(TIME_SLOTS[i])
                                i += 1
                            
                            if not valid_for_all or accumulated < lab_duration:
                                valid_for_all = False
                                break
                            
                            if not slot_indices:
                                slot_indices = temp_slot_indices
                        
                        if not valid_for_all or not slot_indices:
                            continue
                        
                        # Need two lab rooms for this department
                        lab_rooms_needed = 2
                        available_lab_rooms = []
                        
                        random.shuffle(computer_lab_rooms)
                        for lab_room in computer_lab_rooms:
                            if lab_room not in room_schedule:
                                room_schedule[lab_room] = {d: set() for d in range(len(DAYS))}
                            if all(si not in room_schedule[lab_room][day] for si in slot_indices):
                                available_lab_rooms.append(lab_room)
                                if len(available_lab_rooms) >= lab_rooms_needed:
                                    break
                        
                        if len(available_lab_rooms) < lab_rooms_needed:
                            valid_for_all = False
                            continue
                        
                        # Assign rooms to sections
                        sections = list(all_department_timetables[dept][semester].keys())
                        for section_key in all_department_timetables[dept][semester]:
                            timetable = all_department_timetables[dept][semester][section_key]
                            faculty = select_faculty_for_section(course_data.get('Faculty', 'TBD'), section_key)
                            
                            # Store both rooms in lab_rooms field
                            lab_rooms_list = available_lab_rooms[:lab_rooms_needed]
                            
                            for idx, si in enumerate(slot_indices):
                                timetable[day][si]['type'] = 'LAB'
                                timetable[day][si]['code'] = code if idx == 0 else ''
                                timetable[day][si]['name'] = name if idx == 0 else ''
                                timetable[day][si]['faculty'] = faculty if idx == 0 else ''
                                timetable[day][si]['classroom'] = lab_rooms_list[0] if idx == 0 else ''
                                timetable[day][si]['lab_rooms'] = lab_rooms_list
                        
                        # Mark rooms as occupied
                        for room in available_lab_rooms:
                            for si in slot_indices:
                                room_schedule[room][day].add(si)
                        
                        lab_days.add(dept_lab_day_key)
                        
                        print(f"Scheduled auditorium course {code} lab for {dept} on day {day} in rooms {available_lab_rooms[:lab_rooms_needed]}")
                        break
                    
                    if valid_for_all:
                        break
    
    # Schedule regular courses
    for department in df['Department'].unique():
        sems = sorted(df[df['Department'] == department]['Semester'].unique())
        
        for semester in sems:
            dept_upper = str(department).strip().upper()
            num_sections = 2 if (dept_upper == "CSE" and int(semester) in [1, 3, 5]) else 1

            courses = df[(df['Department'] == department) & (df['Semester'] == semester)]
            if 'Schedule' in courses.columns:
                courses = courses[(courses['Schedule'].fillna('Yes').str.upper() == 'YES') | (courses['Schedule'].isna())]
            if courses.empty:
                continue

            basket_config = get_basket_config_for_semester(int(semester), electives_data)
            is_basket_semester = basket_config is not None and dept_upper in basket_config['departments']
            
            if 'P' in courses.columns:
                lab_courses = courses[courses['P'] > 0].copy()
                non_lab_courses = courses[courses['P'] == 0].copy()
            else:
                lab_courses = courses.head(0)
                non_lab_courses = courses.copy()

            if not lab_courses.empty:
                lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
                lab_courses = lab_courses.sort_values('priority', ascending=False)
            non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
            non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)

            combined = pd.concat([lab_courses, non_lab_courses])
            combined['is_elective'] = combined.apply(is_elective, axis=1)
            
            if is_basket_semester:
                combined = combined[~combined['is_elective']]
                
            courses_combined = combined.sort_values(by=['is_elective', 'priority'], ascending=[False, False]).drop_duplicates()

            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65 + section)}"
                ws = wb.create_sheet(title=section_title)

                overview.cell(row=row_index, column=1, value=department)
                overview.cell(row=row_index, column=2, value=str(semester))
                overview.cell(row=row_index, column=3, value=section_title)
                row_index += 1

                section_key = chr(65 + section) if num_sections > 1 else 'A'
                timetable = all_department_timetables[department][semester][section_key]

                section_subject_color = {}
                color_iter = iter(SUBJECT_COLORS)
                course_faculty_map = {}
                auditorium_course_codes = []

                for _, c in courses_combined.iterrows():
                    code = str(c.get('Course Code', '')).strip()
                    if code and code not in section_subject_color:
                        try:
                            section_subject_color[code] = next(color_iter)
                        except StopIteration:
                            section_subject_color[code] = random.choice(SUBJECT_COLORS)
                        course_faculty_map[code] = select_faculty_for_section(c.get('Faculty', 'TBD'), section_key)
                    
                    if is_auditorium_course(c):
                        auditorium_course_codes.append(code)

                for _, course in courses_combined.iterrows():
                    code = str(course.get('Course Code', '')).strip()
                    name = str(course.get('Course Name', '')).strip()
                    faculty = select_faculty_for_section(course.get('Faculty', 'TBD'), section_key)

                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}

                    is_aud = is_auditorium_course(course)
                    if is_aud:
                        continue
                        
                    lec_sessions, tut_sessions, lab_sessions, ss_sessions, lab_duration = calculate_required_sessions(course, config)
                    
                    print(f"Course {code}: L={lec_sessions}, T={tut_sessions}, P={lab_sessions}, S={ss_sessions}")
                    
                    room_type = get_required_room_type(course)

                    def schedule_component(required_minutes, comp_type, attempts_limit=10000):
                        for attempt in range(attempts_limit):
                            day = random.randint(0, len(DAYS)-1)
                            starts = get_all_possible_start_indices_for_duration(comp_type, TIME_SLOTS)
                            for start_idx in starts:
                                if check_course_component_conflict(timetable, day, code, comp_type, TIME_SLOTS, is_aud):
                                    continue
                                
                                # FIX: Add the missing arguments to the function call
                                slot_indices, candidate_room = find_consecutive_slots_for_minutes(
                                    timetable, day, start_idx, required_minutes, semester,
                                    professor_schedule, faculty, room_schedule, room_type,
                                    code, course_room_mapping, comp_type, config, TIME_SLOTS, lecture_rooms, computer_lab_rooms, auditorium_rooms, small_rooms, medium_rooms, large_rooms
                                )

                                if slot_indices is None:
                                    continue
                                if not check_professor_availability(professor_schedule, faculty, day, slot_indices[0], len(slot_indices), TIME_SLOTS):
                                    continue
                                if candidate_room is None:
                                    continue
                                    
                                for si_idx, si in enumerate(slot_indices):
                                    timetable[day][si]['type'] = 'LEC' if comp_type == 'LEC' else ('LAB' if comp_type == 'LAB' else ('TUT' if comp_type == 'TUT' else 'SS'))
                                    timetable[day][si]['code'] = code if si_idx == 0 else ''
                                    timetable[day][si]['name'] = name if si_idx == 0 else ''
                                    timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                                    timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                                    # Store lab rooms for display
                                    if comp_type == 'LAB':
                                        timetable[day][si]['lab_rooms'] = [candidate_room]
                                    professor_schedule[faculty][day].add(si)
                                    if candidate_room not in room_schedule:
                                        room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                                    room_schedule[candidate_room][day].add(si)
                                return True
                        return False

                    for _ in range(lec_sessions):
                        ok = schedule_component(LECTURE_MIN, 'LEC', attempts_limit=800)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LEC', section, "Number of collisions exceeded limit")

                    for _ in range(tut_sessions):
                        ok = schedule_component(TUTORIAL_MIN, 'TUT', attempts_limit=600)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'TUT', section, "No slot available")

                    for _ in range(lab_sessions):
                        ok = schedule_component(lab_duration, 'LAB', attempts_limit=800)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LAB', section, "Lab not scheduled")

                    for _ in range(ss_sessions):
                        ok = schedule_component(SELF_STUDY_MIN, 'SS', attempts_limit=400)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'SS', section, "Self-study not scheduled")

                # --- CHANGE: Call with include_basket_details=False ---
                write_timetable_to_sheet(ws, timetable, courses_combined, section_subject_color, 
                                        course_faculty_map, course_room_mapping, semester, 
                                        is_basket_semester, {}, config, TIME_SLOTS, auditorium_course_codes, computer_lab_rooms, include_basket_details=False, electives_data=electives_data)

    # Format overview sheet
    for col in range(1, 4):
        overview.column_dimensions[get_column_letter(col)].width = 20
    for row_ in overview.iter_rows(min_row=1, max_row=4):
        for cell in row_:
            cell.font = Font(bold=True)
    for cell in overview[4]:
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    for row_ in overview.iter_rows(min_row=5, max_row=row_index-1):
        for cell in row_:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))

    # --- CHANGE: Add new section to create dedicated elective sheets ---
    electives_data = load_electives()
    basket_semesters = [1, 3, 5, 7]
    all_departments = df['Department'].unique()

    for semester in basket_semesters:
        basket_config = get_basket_config_for_semester(semester, electives_data)
        if not basket_config:
            continue

        template_department = None
        for dept in all_departments:
            if semester in all_department_timetables.get(dept, {}):
                template_department = dept
                break
        
        if template_department:
            template_section_key = list(all_department_timetables[template_department][semester].keys())[0]
            template_timetable = all_department_timetables[template_department][semester][template_section_key]
            
            sheet_name = f"Semester_{semester}_Electives"
            ws = wb.create_sheet(title=sheet_name)
            
            write_basket_only_sheet(
                ws, 
                template_timetable, 
                semester, 
                electives_data, 
                config, 
                TIME_SLOTS
            )
            
            overview.cell(row=row_index, column=1, value="Electives")
            overview.cell(row=row_index, column=2, value=str(semester))
            overview.cell(row=row_index, column=3, value=sheet_name)
            row_index += 1

    # --- NEW: Add free room sheet ---
    write_free_room_sheet(wb, room_schedule, config, TIME_SLOTS, lecture_rooms, computer_lab_rooms, auditorium_rooms)
    overview.cell(row=row_index, column=1, value="Free Rooms")
    overview.cell(row=row_index, column=2, value="All")
    overview.cell(row=row_index, column=3, value="Free Rooms")
    row_index += 1

    out_filename = os.path.join(OUTPUT_DIR, "timetable_all_departments.xlsx")
    try:
        wb.save(out_filename)
        print(f"\nCombined timetable saved as {out_filename}")
    except Exception as e:
        print(f"Failed to save: {e}")
        traceback.print_exc()

    try:
        create_teacher_and_unscheduled_from_combined(out_filename, unscheduled_components, config)
    except Exception as e:
        print("Failed to generate teacher/unscheduled workbooks:", e)
        traceback.print_exc()

    try:
        write_electives_to_output(all_department_timetables, out_filename, electives_data)
    except Exception as e:
        print("Failed to write electives output:", e)
        traceback.print_exc()

    return out_filename

def write_timetable_to_sheet(ws, timetable, courses_combined, section_subject_color, 
                             course_faculty_map, course_room_mapping, semester, 
                             is_basket_semester, basket_slots, config, TIME_SLOTS, auditorium_course_codes, computer_lab_rooms, include_basket_details=True, electives_data=None):
    # --- FIX: Define DAYS from config object passed to the function ---
    DAYS = config.get("days", DEFAULT_CONFIG["days"])
    
    BASKET_LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    BASKET_TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    
    header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
    ws.append(header)
    
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    lec_fill = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")
    lab_fill = PatternFill(start_color="7CFC00", end_color="7CFC00", fill_type="solid")
    tut_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    ss_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    break_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    minor_fill = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid")
    basket_fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
    auditorium_fill = PatternFill(start_color="9370DB", end_color="9370DB", fill_type="solid")

    # Create a mapping of basket slots to collect all electives for each time slot
    basket_slot_mapping = {}
    for day_idx, day_name in enumerate(DAYS):
        for slot_idx in range(len(TIME_SLOTS)):
            if timetable[day_idx][slot_idx]['type'] is not None and 'Course' in timetable[day_idx][slot_idx]['name']:
                key = (day_idx, slot_idx)
                if key not in basket_slot_mapping:
                    basket_slot_mapping[key] = {
                        'type': timetable[day_idx][slot_idx]['type'],
                        'electives': []
                    }
                
                # FIX: Remove duplicates in electives for display
                # Only add each elective once per time slot, regardless of how many departments it appears in
                if 'electives' in timetable[day_idx][slot_idx] and timetable[day_idx][slot_idx]['electives']:
                    for elective in timetable[day_idx][slot_idx]['electives']:
                        # Create a unique identifier for each elective (code + room)
                        elective_id = f"{elective['code']}_{elective['room']}"
                        # Check if this elective is already in our list for this time slot
                        if not any(e['code'] == elective['code'] and e['room'] == elective['room'] 
                                 for e in basket_slot_mapping[key]['electives']):
                            basket_slot_mapping[key]['electives'].append({
                                'code': elective['code'],
                                'room': elective['room'],
                                'faculty': elective['faculty']
                            })

    for day_idx, day_name in enumerate(DAYS):
        ws.append([day_name] + [''] * len(TIME_SLOTS))
        row_num = ws.max_row
        merges = []
        
        for slot_idx in range(len(TIME_SLOTS)):
            cell_obj = ws.cell(row=row_num, column=slot_idx + 2)
            
            if is_minor_slot(TIME_SLOTS[slot_idx]):
                cell_obj.value = "Minor Slot"
                cell_obj.fill = minor_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue

            if is_break_time_slot(TIME_SLOTS[slot_idx], semester, config=config):
                cell_obj.value = "BREAK"
                cell_obj.fill = break_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue

            if timetable[day_idx][slot_idx]['type'] is None:
                cell_obj.border = border
                continue

            typ = timetable[day_idx][slot_idx]['type']
            code = timetable[day_idx][slot_idx]['code']
            cls = timetable[day_idx][slot_idx]['classroom']
            fac = timetable[day_idx][slot_idx]['faculty']
            name = timetable[day_idx][slot_idx]['name']

            # Check if this is a basket slot
            key = (day_idx, slot_idx)
            if key in basket_slot_mapping:
                # This is a basket slot, display all electives
                basket_info = basket_slot_mapping[key]
                display_parts = []
                
                # FIX: Remove duplicates in electives for display
                seen_electives = set()
                unique_electives = []
                for elective in basket_info['electives']:
                    elective_id = f"{elective['code']}_{elective['room']}"
                    if elective_id not in seen_electives:
                        seen_electives.add(elective_id)
                        unique_electives.append(elective)
                
                for elective in unique_electives:
                    display_parts.append(f"{elective['code']}\nRoom: {elective['room']}")
                
                display = '\n'.join(display_parts)
                display += f"\n{basket_info['type']}\n{fac}"
                
                span = [slot_idx]
                j = slot_idx + 1
                while (j < len(TIME_SLOTS) and 
                       (day_idx, j) in basket_slot_mapping and 
                       basket_slot_mapping[(day_idx, j)]['type'] == basket_info['type']):
                    span.append(j)
                    j += 1
                
                cell_obj.value = display
                cell_obj.fill = basket_fill
                merges.append((slot_idx + 2, slot_idx + 2 + len(span) - 1, display, basket_fill))
            elif cls and any(auditorium in cls for auditorium in ["Auditorium", "240"]):
                display = f"{code}\n{typ}\nRoom: {cls}\n{fac}"
                fill = auditorium_fill
                
                span = [slot_idx]
                j = slot_idx + 1
                while (j < len(TIME_SLOTS) and 
                       timetable[day_idx][j]['type'] is not None and 
                       timetable[day_idx][j]['code'] == ''):
                    span.append(j)
                    j += 1
                
                cell_obj.value = display
                cell_obj.fill = fill
                merges.append((slot_idx + 2, slot_idx + 2 + len(span) - 1, display, fill))
            else:
                # Regular course
                span = [slot_idx]
                j = slot_idx + 1
                while (j < len(TIME_SLOTS) and 
                       timetable[day_idx][j]['type'] is not None and 
                       timetable[day_idx][j]['code'] == ''):
                    span.append(j)
                    j += 1
                
                # Special handling for lab slots to show two rooms
                if typ == 'LAB':
                    if 'lab_rooms' in timetable[day_idx][slot_idx] and timetable[day_idx][slot_idx]['lab_rooms']:
                        lab_rooms_list = timetable[day_idx][slot_idx]['lab_rooms']
                        if len(lab_rooms_list) >= 2:
                            rooms_display = f"Rooms: {lab_rooms_list[0]}, {lab_rooms_list[1]}"
                        else:
                            rooms_display = f"Room: {lab_rooms_list[0]}" if lab_rooms_list else f"Room: {cls}"
                    else:
                        rooms_display = f"Room: {cls}"
                        if cls and computer_lab_rooms:
                            for room in computer_lab_rooms:
                                if room != cls:
                                    rooms_display = f"Rooms: {cls}, {room}"
                                    break
                    
                    display = f"{typ}\n{rooms_display}\n{fac}"
                else:
                    display = f"{typ}\nRoom: {cls}\n{fac}"
                
                if code in section_subject_color:
                    subj_color = section_subject_color[code]
                    fill = PatternFill(start_color=subj_color, end_color=subj_color, fill_type="solid")
                else:
                    fill = {'LEC': lec_fill, 'LAB': lab_fill, 'TUT': tut_fill, 'SS': ss_fill}.get(typ, lec_fill)
                
                cell_obj.value = display
                cell_obj.fill = fill
                merges.append((slot_idx + 2, slot_idx + 2 + len(span) - 1, display, fill))
            
            cell_obj.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell_obj.border = border

        for start_col, end_col, val, fill in merges:
            if end_col > start_col:
                rng = f"{get_column_letter(start_col)}{row_num}:{get_column_letter(end_col)}{row_num}"
                try:
                    ws.merge_cells(rng)
                    mc = ws[f"{get_column_letter(start_col)}{row_num}"]
                    mc.value = val
                    mc.fill = fill
                    mc.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    mc.border = border
                except:
                    pass

    for col_idx in range(1, len(TIME_SLOTS)+2):
        try:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        except:
            pass

    for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
        ws.row_dimensions[row[0].row].height = 40

    current_row = len(DAYS) + 4

    # Self-Study Only Courses section
    ss_courses = []
    for _, course in courses_combined.iterrows():
        l = int(course['L']) if pd.notna(course['L']) else 0
        t = int(course['T']) if pd.notna(course['T']) else 0
        p = int(course['P']) if pd.notna(course['P']) else 0
        s = int(course['S']) if pd.notna(course['S']) else 0
        if s > 0 and l == 0 and t == 0 and p == 0:
            ss_courses.append({
                'code': str(course['Course Code']),
                'name': str(course['Course Name']),
                'faculty': str(course['Faculty'])
            })

    if ss_courses:
        ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        headers = ['Course Code', 'Course Name', 'Faculty']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True)
        current_row += 1

        for course in ss_courses:
            ws.cell(row=current_row, column=1, value=course['code'])
            ws.cell(row=current_row, column=2, value=course['name'])
            ws.cell(row=current_row, column=3, value=course['faculty'])
            current_row += 1

        current_row += 2

    # --- CHANGE: Wrap entire basket information section in a conditional ---
    # Add Basket Information
    if include_basket_details and is_basket_semester and electives_data and semester in electives_data:
        ws.cell(row=current_row, column=1, value="Cross-Department Elective Information")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 2

        if 'B1' in electives_data[semester] and 'B2' in electives_data[semester] and 'B3' in electives_data[semester] and 'B4' in electives_data[semester]:
            for basket_label in ['B1', 'B2', 'B3', 'B4']:
                ws.cell(row=current_row, column=1, value=f"{basket_label} Electives:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1
                
                electives_list = electives_data[semester][basket_label].get('electives', [])
                electives_str = ", ".join(electives_list)
                ws.cell(row=current_row, column=1, value=electives_str)
                current_row += 1
                
                faculty_list = electives_data[semester][basket_label].get('faculty', [])
                faculty_str = ", ".join(faculty_list) if faculty_list else "TBD"
                ws.cell(row=current_row, column=1, value=f"Faculty: {faculty_str}")
                current_row += 1
                
                n_value = electives_data[semester][basket_label].get('n_value', 1)
                ws.cell(row=current_row, column=1, value=f"Rooms per slot: {n_value}")
                current_row += 1
                
                ws.cell(row=current_row, column=1, value=f"Lectures: 2 ({BASKET_LECTURE_MIN} min each)")
                ws.cell(row=current_row, column=2, value=f"Tutorials: 1 ({BASKET_TUTORIAL_MIN} min each)")
                current_row += 1
                
                ws.cell(row=current_row, column=1, value="Shared across: CSE, DSAI, ECE")
                current_row += 2
        elif 'B1' in electives_data[semester] and 'B2' in electives_data[semester]:
            for basket_label in ['B1', 'B2']:
                ws.cell(row=current_row, column=1, value=f"{basket_label} Electives:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1
                
                electives_list = electives_data[semester][basket_label].get('electives', [])
                electives_str = ", ".join(electives_list)
                ws.cell(row=current_row, column=1, value=electives_str)
                current_row += 1
                
                faculty_list = electives_data[semester][basket_label].get('faculty', [])
                faculty_str = ", ".join(faculty_list) if faculty_list else "TBD"
                ws.cell(row=current_row, column=1, value=f"Faculty: {faculty_str}")
                current_row += 1
                
                n_value = electives_data[semester][basket_label].get('n_value', 1)
                ws.cell(row=current_row, column=1, value=f"Rooms per slot: {n_value}")
                current_row += 1
                
                ws.cell(row=current_row, column=1, value=f"Lectures: 2 ({BASKET_LECTURE_MIN} min each)")
                ws.cell(row=current_row, column=2, value=f"Tutorials: 1 ({BASKET_TUTORIAL_MIN} min each)")
                current_row += 1
                
                ws.cell(row=current_row, column=1, value="Shared across: CSE, DSAI, ECE")
                current_row += 2
        else:
            ws.cell(row=current_row, column=1, value="ELECTIVE slots shared across CSE, DSAI, ECE")
            current_row += 1
            
            for basket_label, basket_data in electives_data[semester].items():
                electives_list = basket_data.get('electives', [])
                electives_str = ", ".join(electives_list)
                ws.cell(row=current_row, column=1, value=electives_str)
                current_row += 1
                
                faculty_list = basket_data.get('faculty', [])
                faculty_str = ", ".join(faculty_list) if faculty_list else "TBD"
                ws.cell(row=current_row, column=1, value=f"Faculty: {faculty_str}")
                current_row += 1
                
                n_value = basket_data.get('n_value', 1)
                ws.cell(row=current_row, column=1, value=f"Rooms per slot: {n_value}")
                current_row += 1
            
            ws.cell(row=current_row, column=1, value=f"Lectures: 2 ({BASKET_LECTURE_MIN} min each)")
            ws.cell(row=current_row, column=2, value=f"Tutorials: 1 ({BASKET_TUTORIAL_MIN} min each)")
            current_row += 2

        # --- CORRECTED SECTION: Scheduled Elective Details ---
        ws.cell(row=current_row, column=1, value="Scheduled Elective Details")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1

        headers = ["Elective", "Type", "Day", "Time", "Room"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = border
            cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        current_row += 1

        # Collect scheduled electives from the basket_slot_mapping
        scheduled_electives_list = []
        for (day_idx, slot_idx), basket_info in basket_slot_mapping.items():
            day_name = DAYS[day_idx]
            start_time = TIME_SLOTS[slot_idx][0].strftime('%H:%M')
            
            # Find the end of the block to get the full duration
            end_slot_idx = slot_idx
            while (end_slot_idx + 1 < len(TIME_SLOTS) and 
                   (day_idx, end_slot_idx + 1) in basket_slot_mapping and 
                   basket_slot_mapping[(day_idx, end_slot_idx + 1)]['type'] == basket_info['type']):
                end_slot_idx += 1
            end_time = TIME_SLOTS[end_slot_idx][1].strftime('%H:%M')
            time_str = f"{start_time} - {end_time}"

            # FIX: Remove duplicates in scheduled electives
            seen_electives = set()
            unique_electives = []
            for elective in basket_info['electives']:
                elective_id = f"{elective['code']}_{elective['room']}"
                if elective_id not in seen_electives:
                    seen_electives.add(elective_id)
                    unique_electives.append(elective)

            for elective in unique_electives:
                scheduled_electives_list.append({
                    'Elective': elective['code'],
                    'Type': basket_info['type'],
                    'Day': day_name,
                    'Time': time_str,
                    'Room': elective['room']
                })

        # Sort the list for consistency (by day, then time)
        scheduled_electives_list.sort(key=lambda x: (DAYS.index(x['Day']), x['Time']))

        # Write the collected electives to the sheet
        for elective_info in scheduled_electives_list:
            ws.cell(row=current_row, column=1, value=elective_info['Elective'])
            ws.cell(row=current_row, column=2, value=elective_info['Type'])
            ws.cell(row=current_row, column=3, value=elective_info['Day'])
            ws.cell(row=current_row, column=4, value=elective_info['Time'])
            ws.cell(row=current_row, column=5, value=elective_info['Room'])
            
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            current_row += 1
        
        if not scheduled_electives_list:
            ws.cell(row=current_row, column=1, value="No electives were scheduled.")
            current_row += 1

        current_row += 2 # Add space before the next section (Legend)

    # Legend
    legend_title = ws.cell(row=current_row, column=1, value="Legend")
    legend_title.font = Font(bold=True, size=12)
    current_row += 2

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    legend_headers = ['Subject Code', 'Color', 'Subject Name', 'Faculty', 'LTPS', 'Room']
    for col, header in enumerate(legend_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = border
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row += 1

    all_course_codes = set()
    for _, course in courses_combined.iterrows():
        code = str(course.get('Course Code', '')).strip()
        if code:
            all_course_codes.add(code)
    
    for code in all_course_codes:
        if code not in section_subject_color:
            continue
            
        color = section_subject_color[code]
        assigned_room = course_room_mapping.get(code, "—")
        
        ws.row_dimensions[current_row].height = 30

        ltps_value = ""
        course_name = ''
        fac_name = ''
        
        for _, course_row in courses_combined.iterrows():
            if str(course_row['Course Code']) == code:
                l = str(int(course_row['L'])) if pd.notna(course_row['L']) else "0"
                t = str(int(course_row['T'])) if pd.notna(course_row['T']) else "0"
                p = str(int(course_row['P'])) if pd.notna(course_row['P']) else "0"
                s = str(int(course_row['S'])) if pd.notna(course_row['S']) and 'S' in course_row else "0"
                ltps_value = f"{l}-{t}-{p}-{s}"
                course_name = str(course_row['Course Name'])
                break

        if code in course_faculty_map:
            fac_name = course_faculty_map[code]

        is_auditorium = code in auditorium_course_codes
        
        cells = [
            (code, None),
            ('', PatternFill(start_color=color, end_color=color, fill_type="solid")),
            (course_name, None),
            (fac_name, None),
            (ltps_value, None),
            (assigned_room, None)
        ]

        if is_auditorium:
            cells[0] = (f"{code} (240)", None)
            cells[1] = ('', auditorium_fill)

        for col, (value, fill) in enumerate(cells, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)

        current_row += 1

# ---------------------------
# Teacher and Unscheduled workbooks
# ---------------------------
def split_faculty_names(fac_str):
    if fac_str is None:
        return []
    s = str(fac_str).strip()
    if s == '' or s.lower() in ['nan', 'none']:
        return []
    parts = [s]
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            break
    return parts if parts else [s]

def parse_cell_for_course(cell_value):
    if cell_value is None:
        return ('', '', '', '')
    text = str(cell_value).strip()
    if text == '':
        return ('', '', '')

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    faculty = ''
    room = ''
    code = ''
    typ = ''

    for ln in lines:
        if 'room' in ln.lower() or 'Room:' in ln:
            parts = ln.split(':')
            if len(parts) >= 2:
                room = parts[-1].strip()

    if len(lines) >= 1:
        last = lines[-1]
        if 'room' not in last.lower() and 'courses' not in last.lower() and ':' not in last:
            faculty = last

    first = lines[0] if lines else ''
    if first:
        tokens = first.split()
        if len(tokens) >= 2 and tokens[1].upper() in ['LEC', 'LAB', 'TUT', 'SS']:
            code = tokens[0].strip()
            typ = tokens[1].strip().upper()
        else:
            code = tokens[0].strip() if tokens else ''
            for t in ['LEC', 'LAB', 'TUT', 'SS']:
                if t in text.upper():
                    typ = t
                    break

    if not faculty and len(lines) >= 2:
        for cand in lines[1:]:
            if any(ch.isalpha() for ch in cand) and 'room' not in cand.lower() and 'courses' not in cand.lower() and ':' not in cand:
                faculty = cand
                break

    return (code, typ, room, faculty)

def create_teacher_and_unscheduled_from_combined(timetable_filename, unscheduled_components, config):
    try:
        wb = load_workbook(timetable_filename, data_only=True)
    except Exception as e:
        print(f"Failed to open {timetable_filename}: {e}")
        return

    teacher_slots = {}
    slot_headers = []
    
    electives_data = load_electives()
    elective_faculty_map = {}
    elective_room_map = {}
    
    if electives_data:
        for semester in electives_data:
            for basket_label, basket_data in electives_data[semester].items():
                electives_list = basket_data.get('electives', [])
                faculty_list = basket_data.get('faculty', [])
                
                for i, faculty in enumerate(faculty_list):
                    if faculty not in elective_faculty_map:
                        elective_faculty_map[faculty] = []
                        elective_room_map[faculty] = []
                    
                    if i < len(electives_list):
                        elective_faculty_map[faculty].append(electives_list[i])

    for sheetname in wb.sheetnames:
        if sheetname.lower() == 'overview':
            continue
        ws = wb[sheetname]
        header = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else '' for c in range(2, ws.max_column + 1)]
        if len(header) > len(slot_headers):
            slot_headers = header
        
        for r in range(2, ws.max_row + 1):
            day = ws.cell(r, 1).value
            if not day or str(day) not in config.get("days", DEFAULT_CONFIG["days"]):
                break
            day_idx = config.get("days", DEFAULT_CONFIG["days"]).index(day)
            
            for c in range(2, ws.max_column + 1):
                code, typ, room, faculty = parse_cell_for_course(ws.cell(r, c).value)
                
                if 'Course' in str(ws.cell(r, c).value) and faculty:
                    faculty_list = split_faculty_names(faculty)
                    
                    for f in faculty_list:
                        if not f or str(f).strip().upper() in ["BREAK", "MINOR SLOT", "NAN", "NONE", "", "MULTIPLE FACULTY"]:
                            continue
                            
                        if f in elective_faculty_map:
                            for elective in elective_faculty_map[f]:
                                teacher_slots.setdefault(f, {d: {i: '' for i in range(len(slot_headers))} for d in range(len(config.get("days", DEFAULT_CONFIG["days"])))})
                                teacher_slots[f][day_idx][c - 2] = f"{elective} {typ}\n({sheetname})\nRoom: {room}"
                                
                                if f not in elective_room_map:
                                    elective_room_map[f] = []
                                elective_room_map[f].append(room)
                else:
                    for f in split_faculty_names(faculty):
                        if not f or str(f).strip().upper() in ["BREAK", "MINOR SLOT", "NAN", "NONE", "", "MULTIPLE FACULTY"]:
                            continue

                        teacher_slots.setdefault(f, {d: {i: '' for i in range(len(slot_headers))} for d in range(len(config.get("days", DEFAULT_CONFIG["days"])))})
                        teacher_slots[f][day_idx][c - 2] = f"{code} {typ}\n({sheetname})\nRoom: {room}" if code else ''

    twb = Workbook()
    if "Sheet" in twb.sheetnames:
        twb.remove(twb["Sheet"])

    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    alt_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for teacher in sorted(teacher_slots.keys()):
        safe_name = teacher[:31] or "Unknown"
        ws = twb.create_sheet(title=safe_name)

        ws.merge_cells("A1:{}1".format(get_column_letter(len(slot_headers) + 1)))
        title_cell = ws.cell(row=1, column=1, value=f"{teacher} — Weekly Timetable")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.append(["Day"] + slot_headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for d, day in enumerate(config.get("days", DEFAULT_CONFIG["days"])):
            row = [day] + [teacher_slots[teacher][d][i] for i in range(len(slot_headers))]
            ws.append(row)
            row_idx = ws.max_row
            if d % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill
            for cell in ws[row_idx]:
                cell.alignment = cell_align
                cell.border = border
            ws.row_dimensions[row_idx].height = 35

        ws.column_dimensions["A"].width = 15
        for col in range(2, len(slot_headers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 20

    twb.save(os.path.join(OUTPUT_DIR, "teacher_timetables.xlsx"))
    print("Saved teacher_timetables.xlsx")

    uwb = Workbook()
    ws = uwb.active
    ws.title = "Unscheduled Courses"

    headers = ["Course Code", "Department", "Semester", "Reason"]
    ws.append(headers)

    unscheduled_unique = {}
    for u in unscheduled_components:
        if u.code not in unscheduled_unique:
            reason_text = str(u.reason).strip() if hasattr(u, "reason") and u.reason else "Scheduling conflict"
            unscheduled_unique[u.code] = {
                "Course Code": u.code,
                "Department": u.department,
                "Semester": u.semester,
                "Reason": reason_text
            }

    for entry in unscheduled_unique.values():
        ws.append([entry[h] for h in headers])

    uwb.save(os.path.join(OUTPUT_DIR, "unscheduled_courses.xlsx"))
    print(f"Saved unscheduled_courses.xlsx with {len(unscheduled_unique)} unique courses")

if __name__ == "__main__":
    try:
        generate_all_timetables()
    except Exception as e:
        print("Error running TT_gen:", e)
        traceback.print_exc()