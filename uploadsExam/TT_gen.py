# TT_gen.py -- Timetable generator with basket system for cross-department electives
# Run: python TT_gen.py
# Requires: pandas, openpyxl

import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
import traceback
import os
import json
import math

# ---------------------------
# Default Configuration
# ---------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")

DEFAULT_CONFIG = {
    "days": ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'],
    "LECTURE_MIN": 90,
    "LAB_MIN": 120,
    "TUTORIAL_MIN": 60,
    "SELF_STUDY_MIN": 60,
    "MORNING_BREAK_START": "10:30",
    "MORNING_BREAK_END": "10:45",
    "LUNCH_BREAK_START": "13:00",
    "LUNCH_BREAK_END": "13:45",
    "LECTURE_TUTORIAL_BREAK_START": "15:30",
    "LECTURE_TUTORIAL_BREAK_END": "15:40",
    "TIME_SLOTS": [
        ["07:30", "09:00"],
        ["09:00", "09:30"],
        ["09:30", "10:00"],
        ["10:00", "10:30"],
        ["10:30", "10:45"],
        ["10:45", "11:00"],
        ["11:00", "11:30"],
        ["11:30", "12:00"],
        ["12:00", "12:15"],
        ["12:15", "12:30"],
        ["12:30", "13:15"],
        ["13:15", "13:30"],
        ["13:30", "14:00"],
        ["14:00", "14:30"],
        ["14:30", "15:00"],
        ["15:00", "15:30"],
        ["15:30", "15:40"],
        ["15:40", "16:00"],
        ["16:00", "16:30"],
        ["16:30", "17:00"],
        ["17:00", "17:30"],
        ["17:30", "18:00"],
        ["18:00", "18:30"],
        ["18:30", "23:59"]
    ],
    "USE_CUSTOM_SLOTS": False
}

# ---------------------------
# Load Configuration Function
# ---------------------------
def load_configuration():
    """Load the latest configuration from config.json or use defaults"""
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            config = json.load(f)
            print("✓ Loaded configuration from config.json")
    except Exception as e:
        print(f"⚠ Could not load config.json: {e}")
        print("Using default configuration")
        config = DEFAULT_CONFIG.copy()
    
    return config

# ---------------------------
# Parse time strings
# ---------------------------
def parse_time_string(time_str):
    """Convert time string like '10:30' to time object"""
    try:
        hour, minute = map(int, time_str.split(':'))
        return time(hour, minute)
    except Exception:
        print(f"⚠ Error parsing time: {time_str}, using default")
        return time(10, 30)

# ---------------------------
# Time slot generation from config
# ---------------------------
def generate_time_slots(config):
    """Generate time slots from config.json or use defaults"""
    use_custom = config.get("USE_CUSTOM_SLOTS", False)
    
    if use_custom:
        time_slots_config = config.get("TIME_SLOTS", DEFAULT_CONFIG["TIME_SLOTS"])
        print(f"✓ Using custom time slots ({len(time_slots_config)} slots)")
    else:
        time_slots_config = DEFAULT_CONFIG["TIME_SLOTS"]
        print(f"✓ Using default time slots ({len(time_slots_config)} slots)")
    
    slots = []
    for slot_config in time_slots_config:
        try:
            start_str, end_str = slot_config
            start = parse_time_string(start_str)
            end = parse_time_string(end_str)
            slots.append((start, end))
        except Exception as e:
            print(f"⚠ Error parsing time slot {slot_config}: {e}")
            continue
    
    if not slots:
        print("⚠ No valid time slots found, using fallback")
        # Fallback to a basic set of slots
        slots = [
            (time(9, 0), time(10, 30)),
            (time(11, 0), time(12, 30)),
            (time(14, 0), time(15, 30)),
            (time(16, 0), time(17, 30))
        ]
    
    return slots

# ---------------------------
# Rest of the TT_gen.py code (with modifications to use config inside functions)
# ---------------------------

@dataclass
class UnscheduledComponent:
    department: str
    semester: int
    code: str
    name: str
    faculty: str
    component_type: str
    sessions: int
    section: int
    reason: str

INPUT_DIR = os.path.join(BASE_DIR, "inputs")
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")

# ---------------------------
# Load CSVs
# ---------------------------
def load_data():
    """Load CSV data files"""
    try:
        df = pd.read_csv(os.path.join(INPUT_DIR, 'combined.csv'))
    except FileNotFoundError:
        raise SystemExit("Error: 'combined.csv' not found in working directory.")

    try:
        rooms_df = pd.read_csv(os.path.join(INPUT_DIR, 'rooms.csv'))
    except FileNotFoundError:
        rooms_df = pd.DataFrame(columns=['roomNumber', 'type'])

    lecture_rooms = rooms_df[rooms_df.get('type', '').str.upper() == 'LECTURE_ROOM']['roomNumber'].tolist()
    computer_lab_rooms = rooms_df[rooms_df.get('type', '').str.upper() == 'COMPUTER_LAB']['roomNumber'].tolist()
    large_rooms = rooms_df[rooms_df.get('type', '').str.upper() == 'SEATER_120']['roomNumber'].tolist()
    
    return df, lecture_rooms, computer_lab_rooms, large_rooms

# ---------------------------
# Load Electives
# ---------------------------
def load_electives():
    """Load electives from elective.csv or elective.xlsx"""
    electives_path_csv = os.path.join(INPUT_DIR, "elective.csv")
    electives_path_xlsx = os.path.join(INPUT_DIR, "elective.xlsx")
    
    electives_df = None
    
    # Try to load from CSV first
    if os.path.exists(electives_path_csv):
        try:
            electives_df = pd.read_csv(electives_path_csv)
            print(f"Loaded electives from {electives_path_csv}")
        except Exception as e:
            print(f"Error loading electives from CSV: {e}")
    
    # If CSV not found or failed, try Excel
    if electives_df is None and os.path.exists(electives_path_xlsx):
        try:
            electives_df = pd.read_excel(electives_path_xlsx)
            print(f"Loaded electives from {electives_path_xlsx}")
        except Exception as e:
            print(f"Error loading electives from Excel: {e}")
    
    if electives_df is None:
        print("Warning: No electives file found. Using default basket configuration.")
        return None
    
    # Process the electives data
    electives_dict = {}
    
    for _, row in electives_df.iterrows():
        sem_str = str(row.iloc[0]).strip().lower()
        electives_str = str(row.iloc[1]).strip()
        
        # Parse semester
        if "1st" in sem_str:
            sem = 1
            basket_label = "ELECTIVE"
        elif "3rd" in sem_str:
            sem = 3
            basket_label = "ELECTIVE"
        elif "5th(b1)" in sem_str:
            sem = 5
            basket_label = "B1"
        elif "5th(b2)" in sem_str:
            sem = 5
            basket_label = "B2"
        else:
            continue
        
        # Parse electives list
        electives_list = [e.strip() for e in electives_str.split(',') if e.strip()]
        
        if sem not in electives_dict:
            electives_dict[sem] = {}
        
        electives_dict[sem][basket_label] = electives_list
    
    return electives_dict

# ---------------------------
# Helper functions (modified to use config)
# ---------------------------
def slot_minutes(slot):
    s, e = slot
    s_m = s.hour*60 + s.minute
    e_m = e.hour*60 + e.minute
    if e_m < s_m:
        e_m += 24*60
    return e_m - s_m

def overlaps(a_start, a_end, b_start, b_end):
    a_s_min = a_start.hour*60 + a_start.minute
    a_e_min = a_end.hour*60 + a_end.minute
    b_s_min = b_start.hour*60 + b_start.minute
    b_e_min = b_end.hour*60 + b_end.minute
    return (a_s_min < b_e_min) and (b_s_min < a_e_min)

def is_break_time_slot(slot, semester=None, comp_type=None, config=None):
    """
    Check if a time slot falls within any break period.
    The config parameter contains the break times.
    """
    if config is None:
        config = load_configuration()
    
    start, end = slot
    # Parse break times from config
    MORNING_BREAK_START = parse_time_string(config.get("MORNING_BREAK_START", "10:30"))
    MORNING_BREAK_END = parse_time_string(config.get("MORNING_BREAK_END", "10:45"))
    LUNCH_BREAK_START = parse_time_string(config.get("LUNCH_BREAK_START", "13:00"))
    LUNCH_BREAK_END = parse_time_string(config.get("LUNCH_BREAK_END", "13:45"))
    LECTURE_TUTORIAL_BREAK_START = parse_time_string(config.get("LECTURE_TUTORIAL_BREAK_START", "15:30"))
    LECTURE_TUTORIAL_BREAK_END = parse_time_string(config.get("LECTURE_TUTORIAL_BREAK_END", "15:40"))
    
    # Debug output
    # print(f"Checking slot {start.strftime('%H:%M')}-{end.strftime('%H:%M')} against breaks:")
    # print(f"  Morning: {MORNING_BREAK_START.strftime('%H:%M')}-{MORNING_BREAK_END.strftime('%H:%M')}")
    # print(f"  Lunch: {LUNCH_BREAK_START.strftime('%H:%M')}-{LUNCH_BREAK_END.strftime('%H:%M')}")
    # print(f"  Evening: {LECTURE_TUTORIAL_BREAK_START.strftime('%H:%M')}-{LECTURE_TUTORIAL_BREAK_END.strftime('%H:%M')}")
    
    # Regular breaks
    if overlaps(start, end, MORNING_BREAK_START, MORNING_BREAK_END):
        # print("  -> Overlaps with morning break")
        return True
    if overlaps(start, end, LUNCH_BREAK_START, LUNCH_BREAK_END):
        # print("  -> Overlaps with lunch break")
        return True
    # Special break for lectures and tutorials (not labs)
    if comp_type in ['LEC', 'TUT'] and overlaps(start, end, LECTURE_TUTORIAL_BREAK_START, LECTURE_TUTORIAL_BREAK_END):
        # print("  -> Overlaps with lecture/tutorial break")
        return True
    return False

def is_minor_slot(slot):
    start, end = slot
    if start == time(7, 30) and end == time(9, 0):
        return True
    if start == time(18, 30):
        return True
    return False

def is_lecture_unfriendly_slot(slot):
    """
    Check if this slot is not suitable for 90-minute lectures.
    17:30-18:30 is only 60 minutes, not enough for a lecture.
    """
    start, end = slot
    # Check if it's the evening slot (17:30-18:30 or later)
    if start >= time(17, 30):
        return True
    return False

def select_faculty_for_section(faculty_field, section_char='A'):
    """
    Select faculty for a specific section.
    If faculty_field contains '&' separator, assign different faculty to different sections.
    Section A gets first faculty, Section B gets second faculty.
    """
    if pd.isna(faculty_field) or str(faculty_field).strip().lower() in ['nan', 'none', '']:
        return "TBD"
    
    s = str(faculty_field).strip()
    
    # Check for '&' separator - this indicates different faculty for different sections
    if '&' in s:
        faculties = [f.strip() for f in s.split('&')]
        if len(faculties) >= 2:
            # Section A gets first faculty, Section B gets second faculty
            if section_char.upper() == 'A':
                return faculties[0]
            elif section_char.upper() == 'B':
                return faculties[1]
            else:
                # For any other section, use the first faculty
                return faculties[0]
        else:
            # Fallback if split doesn't work properly
            return faculties[0] if faculties else "TBD"
    else:
        # No '&' separator, use the single faculty for all sections
        for sep in ['/', ',', ';']:
            if sep in s:
                return s.split(sep)[0].strip()
        return s

def is_elective(course_row):
    """Check if a course is an elective based on name or code"""
    name = str(course_row.get('Course Name', '')).lower()
    code = str(course_row.get('Course Code', '')).lower()
    keywords = ["elective", "oe", "open elective", "pe", "program elective"]
    return any(k in name for k in keywords) or any(k in code for k in keywords)

def get_course_priority(row):
    try:
        l = int(row.get('L', 0)) if pd.notna(row.get('L', 0)) else 0
        t = int(row.get('T', 0)) if pd.notna(row.get('T', 0)) else 0
        p = int(row.get('P', 0)) if pd.notna(row.get('P', 0)) else 0
        return -(l + t + p)
    except Exception:
        return 0

def calculate_required_sessions(course_row, config):
    """
    Calculate the number of sessions and their durations for each component type.
    L, T, P, S values represent credit/contact hours.
    """
    l = int(course_row['L']) if ('L' in course_row and pd.notna(course_row['L'])) else 0
    t = int(course_row['T']) if ('T' in course_row and pd.notna(course_row['T'])) else 0
    p = int(course_row['P']) if ('P' in course_row and pd.notna(course_row['P'])) else 0
    s = int(course_row['S']) if ('S' in course_row and pd.notna(course_row['S'])) else 0
    
    # Get durations from config
    LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    LAB_MIN = config.get("LAB_MIN", DEFAULT_CONFIG["LAB_MIN"])  # This is the key line
    TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    SELF_STUDY_MIN = config.get("SELF_STUDY_MIN", DEFAULT_CONFIG["SELF_STUDY_MIN"])
    
    # Calculate number of sessions based on credits and standard session durations
    lec_sessions = max(1, math.ceil((l * 60) / LECTURE_MIN)) if l > 0 else 0
    tut_sessions = t if t > 0 else 0  # T value typically equals number of tutorials
    lab_sessions = 1 if p > 0 else 0  # P > 0 means there's one lab session
    
    # FIX: Use LAB_MIN from config instead of calculating based on P value
    lab_duration = LAB_MIN if p > 0 else 0
    
    ss_sessions = s if s > 0 else 0  # S value typically equals number of self-study sessions
    
    return (lec_sessions, tut_sessions, lab_sessions, ss_sessions, lab_duration)

def get_required_room_type(course_row):
    try:
        p = int(course_row['P']) if ('P' in course_row and pd.notna(course_row['P'])) else 0
        return 'COMPUTER_LAB' if p > 0 else 'LECTURE_ROOM'
    except Exception:
        return 'LECTURE_ROOM'

# ---------------------------
# Room allocation
# ---------------------------
def find_suitable_room_for_slot(course_code, room_type, day, slot_indices, room_schedule, course_room_mapping, config, lecture_rooms, computer_lab_rooms):
    if course_code in course_room_mapping:
        fixed_room = course_room_mapping[course_code]
        for si in slot_indices:
            if si in room_schedule[fixed_room][day]:
                return None
        return fixed_room

    pool = computer_lab_rooms if room_type == 'COMPUTER_LAB' else lecture_rooms
    if not pool:
        return None
    random.shuffle(pool)
    for room in pool:
        if room not in room_schedule:
            room_schedule[room] = {d: set() for d in range(len(config.get("days", DEFAULT_CONFIG["days"])))}
        if all(si not in room_schedule[room][day] for si in slot_indices):
            course_room_mapping[course_code] = room
            return room
    return None

def find_consecutive_slots_for_minutes(timetable, day, start_idx, required_minutes,
                                       semester, professor_schedule, faculty,
                                       room_schedule, room_type, course_code, course_room_mapping, comp_type, config, TIME_SLOTS, lecture_rooms, computer_lab_rooms):
    LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    BASKET_LECTURE_MIN = LECTURE_MIN
    BASKET_TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    
    n = len(TIME_SLOTS)
    slot_indices = []
    i = start_idx
    accumulated = 0

    while i < n and accumulated < required_minutes:
        if is_minor_slot(TIME_SLOTS[i]):
            return None, None
        if is_break_time_slot(TIME_SLOTS[i], semester, comp_type, config):
            return None, None
        # Check if slot is already occupied (including by basket electives)
        if timetable[day][i]['type'] is not None:
            return None, None
        if faculty in professor_schedule and i in professor_schedule[faculty][day]:
            return None, None
        
        if room_type == 'COMPUTER_LAB' and not computer_lab_rooms:
            return None, None
        if room_type == 'LECTURE_ROOM' and not lecture_rooms:
            return None, None

        slot_indices.append(i)
        accumulated += slot_minutes(TIME_SLOTS[i])
        i += 1

    if accumulated >= required_minutes:
        room = find_suitable_room_for_slot(course_code, room_type, day, slot_indices, room_schedule, course_room_mapping, config, lecture_rooms, computer_lab_rooms)
        if room is not None:
            return slot_indices, room

    return None, None

def get_all_possible_start_indices_for_duration(comp_type=None, TIME_SLOTS=None):
    """
    Get all possible start indices, but filter out lecture-unfriendly slots for lectures.
    """
    idxs = list(range(len(TIME_SLOTS)))
    random.shuffle(idxs)
    
    # If this is a lecture, filter out the 17:30-18:30 slot
    if comp_type == 'LEC':
        idxs = [i for i in idxs if not is_lecture_unfriendly_slot(TIME_SLOTS[i])]
    
    return idxs

def check_professor_availability(professor_schedule, faculty, day, start_idx, duration_slots, TIME_SLOTS):
    if faculty not in professor_schedule:
        return True
    if not professor_schedule[faculty][day]:
        return True
    new_start = TIME_SLOTS[start_idx][0]
    new_start_m = new_start.hour*60 + new_start.minute
    MIN_GAP = 180
    for s in professor_schedule[faculty][day]:
        exist_start = TIME_SLOTS[s][0]
        exist_m = exist_start.hour*60 + exist_start.minute
        if abs(exist_m - new_start_m) < MIN_GAP:
            return False
    return True

def check_course_component_conflict(timetable, day, course_code, comp_type, TIME_SLOTS):
    """
    Check if scheduling a component of this course on this day would conflict
    with existing components of the same course.
    """
    for slot_idx in range(len(TIME_SLOTS)):
        if timetable[day][slot_idx]['type'] is not None:
            existing_code = timetable[day][slot_idx]['code']
            existing_type = timetable[day][slot_idx]['type']
            
            # Check if it's the same course (ignoring basket courses)
            if existing_code == course_code and 'Courses' not in existing_code:
                # Two lectures shouldn't be on the same day
                if comp_type == 'LEC' and existing_type == 'LEC':
                    return True
                # One lecture and one tutorial shouldn't be on the same day
                if (comp_type == 'LEC' and existing_type == 'TUT') or \
                   (comp_type == 'TUT' and existing_type == 'LEC'):
                    return True
    return False

def add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, comp_type, section, reason):
    existing = next((u for u in unscheduled_components if u.code == code), None)
    if existing:
        if comp_type not in existing.component_type:
            existing.component_type += f", {comp_type}"
        if reason not in existing.reason:
            existing.reason += f"; {reason}"
    else:
        unscheduled_components.append(UnscheduledComponent(department, semester, code, name, faculty, comp_type, 1, section, reason))

# ---------------------------
# Basket scheduling for cross-department electives
# ---------------------------
def check_basket_slot_conflict(new_day, new_slots, existing_basket_slots):
    """Check if new basket slot conflicts with existing ones (overlap or adjacency)"""
    for existing in existing_basket_slots:
        if existing['day'] == new_day:
            # Check for overlap
            for new_slot in new_slots:
                for existing_slot in existing['slots']:
                    if new_slot == existing_slot:
                        return True  # Direct overlap
            
            # Check for adjacency (slots next to each other)
            new_min = min(new_slots)
            new_max = max(new_slots)
            existing_min = min(existing['slots'])
            existing_max = max(existing['slots'])
            
            # If slots are adjacent (no gap)
            if abs(new_min - existing_max) == 1 or abs(new_max - existing_min) == 1:
                return True
    
    return False

def get_basket_config_for_semester(semester, electives_data):
    """Get basket configuration for a semester from ELECTIVES_DATA or default config"""
    if electives_data and semester in electives_data:
        # Use electives data if available
        baskets = electives_data[semester]
        
        # Create a config similar to the original BASKET_CONFIG
        config = {
            'departments': ['CSE', 'DSAI', 'ECE'],
            'lectures': 2,
            'tutorials': 1,
        }
        
        if 'B1' in baskets and 'B2' in baskets:
            # Semester 5 case
            config['baskets'] = ['B1', 'B2']
        else:
            # Semesters 1 and 3 case
            config['label'] = 'ELECTIVE'
        
        return config
    else:
        # Fall back to default config if no electives data
        if semester == 1:
            return {
                'departments': ['CSE', 'DSAI', 'ECE'],
                'lectures': 2,
                'tutorials': 1,
                'label': 'ELECTIVE'
            }
        elif semester == 3:
            return {
                'departments': ['CSE', 'DSAI', 'ECE'],
                'lectures': 2,
                'tutorials': 1,
                'label': 'ELECTIVE'
            }
        elif semester == 5:
            return {
                'departments': ['CSE', 'DSAI', 'ECE'],
                'lectures': 2,
                'tutorials': 1,
                'baskets': ['B1', 'B2']
            }
        else:
            return None

def schedule_basket_slots(semester, all_department_timetables, professor_schedule, room_schedule, course_room_mapping, config, TIME_SLOTS, electives_data, large_rooms):
    """
    Schedule basket slots that are shared across CSE, DSAI, ECE departments.
    Returns dict mapping department to scheduled slot info.
    """
    BASKET_LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    BASKET_TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    
    basket_config = get_basket_config_for_semester(semester, electives_data)
    if not basket_config:
        return {}
    
    departments = basket_config['departments']
    
    # For semester 5, we need to split into B1 and B2
    if 'baskets' in basket_config:
        # Schedule all baskets together to ensure no conflicts
        all_basket_slots = {}
        existing_slots = []  # Track all scheduled slots to prevent conflicts
        
        # Schedule B1 first, then B2 to ensure both get scheduled
        for basket_label in basket_config['baskets']:
            # Each basket gets specified lectures and tutorials
            lec_count = basket_config['lectures']
            tut_count = basket_config['tutorials']
            
            basket_result = schedule_single_basket_with_constraints(
                semester, basket_label, departments, lec_count, tut_count,
                all_department_timetables, professor_schedule, room_schedule, 
                course_room_mapping, existing_slots, config, TIME_SLOTS, large_rooms
            )
            all_basket_slots[basket_label] = basket_result
            
            # Add scheduled slots to existing_slots for conflict checking
            for lecture in basket_result.get('lectures', []):
                existing_slots.append({
                    'day': lecture['day'],
                    'slots': lecture['slots'],
                    'type': 'lecture',
                    'basket': basket_label
                })
            for tutorial in basket_result.get('tutorials', []):
                existing_slots.append({
                    'day': tutorial['day'],
                    'slots': tutorial['slots'],
                    'type': 'tutorial',
                    'basket': basket_label
                })
        
        return all_basket_slots
    else:
        # For semesters 1 and 3, schedule as a single ELECTIVE group
        return schedule_single_basket_with_constraints(
            semester, basket_config['label'], departments, 
            basket_config['lectures'], basket_config['tutorials'],
            all_department_timetables, professor_schedule, room_schedule, course_room_mapping, [], config, TIME_SLOTS, large_rooms
        )

def schedule_single_basket_with_constraints(semester, label, departments, lec_count, tut_count,
                                         all_department_timetables, professor_schedule, room_schedule, 
                                         course_room_mapping, existing_basket_slots, config, TIME_SLOTS, large_rooms):
    """
    Schedule a single basket/elective group across multiple departments at the same time
    with constraints to avoid adjacency and overlap with other baskets.
    """
    BASKET_LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    BASKET_TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    
    scheduled_slots = {'lectures': [], 'tutorials': []}
    scheduled_lecture_days = set()  # Track days with lectures to avoid multiple lectures on the same day
    
    # Use large rooms for cross-department sessions
    room_pool = large_rooms if large_rooms else []
    
    # Schedule lectures with constraints
    for lec_idx in range(lec_count):
        best_slot = None
        best_room = None
        
        for attempt in range(2000):  # Increased attempts for constraints
            day = random.randint(0, len(config.get("days", DEFAULT_CONFIG["days"]))-1)
            # Ensure two lectures are not on the same day
            if day in scheduled_lecture_days:
                continue
                
            starts = get_all_possible_start_indices_for_duration('LEC', TIME_SLOTS)  # Use lecture-friendly slots
            
            for start_idx in starts:
                # Check if this slot works for ALL departments and sections
                valid_for_all = True
                slot_indices = []
                
                for dept in departments:
                    # Check if this department exists in our timetables
                    if dept not in all_department_timetables:
                        continue
                        
                    if semester not in all_department_timetables[dept]:
                        continue
                        
                    # For each section in this department
                    for section_key in all_department_timetables[dept][semester]:
                        timetable = all_department_timetables[dept][semester][section_key]
                        
                        # Check consecutive slots using BASKET_LECTURE_MIN for basket lectures
                        accumulated = 0
                        i = start_idx
                        temp_slot_indices = []
                        
                        while i < len(TIME_SLOTS) and accumulated < BASKET_LECTURE_MIN:
                            if is_minor_slot(TIME_SLOTS[i]):
                                valid_for_all = False
                                break
                            if is_break_time_slot(TIME_SLOTS[i], semester, 'LEC', config):
                                valid_for_all = False
                                break
                            if timetable[day][i]['type'] is not None:
                                valid_for_all = False
                                break
                            temp_slot_indices.append(i)
                            accumulated += slot_minutes(TIME_SLOTS[i])
                            i += 1
                        
                        if not valid_for_all or accumulated < BASKET_LECTURE_MIN:
                            valid_for_all = False
                            break
                        
                        # Store slot indices from the first valid check
                        if not slot_indices:
                            slot_indices = temp_slot_indices
                
                # Check for conflicts with existing basket slots
                if valid_for_all and slot_indices:
                    if check_basket_slot_conflict(day, slot_indices, existing_basket_slots):
                        valid_for_all = False
                
                if valid_for_all and room_pool and slot_indices:
                    # Find a room that's free for all these slots
                    random.shuffle(room_pool)
                    for room in room_pool:
                        if room not in room_schedule:
                            room_schedule[room] = {d: set() for d in range(len(config.get("days", DEFAULT_CONFIG["days"])))}
                        if all(si not in room_schedule[room][day] for si in slot_indices):
                            best_slot = (day, slot_indices)
                            best_room = room
                            break
                
                if best_slot:
                    break
            if best_slot:
                break
        
        if best_slot:
            scheduled_slots['lectures'].append({'day': best_slot[0], 'slots': best_slot[1], 'room': best_room})
            scheduled_lecture_days.add(best_slot[0])  # Mark this day as having a lecture
            # Mark room as occupied
            for si in best_slot[1]:
                room_schedule[best_room][best_slot[0]].add(si)
            print(f"Scheduled {label} Lecture {lec_idx + 1} on day {best_slot[0]} in room {best_room}")
        else:
            print(f"WARNING: Could not schedule {label} Lecture {lec_idx + 1}")
    
    # Schedule tutorials with constraints using BASKET_TUTORIAL_MIN
    for tut_idx in range(tut_count):
        best_slot = None
        best_room = None
        
        for attempt in range(2000):  # Increased attempts for constraints
            day = random.randint(0, len(config.get("days", DEFAULT_CONFIG["days"]))-1)
            # Ensure tutorial is not on the same day as any lecture of this basket
            if day in scheduled_lecture_days:
                continue
                
            starts = get_all_possible_start_indices_for_duration('TUT', TIME_SLOTS)  # Use tutorial-friendly slots
            
            for start_idx in starts:
                valid_for_all = True
                slot_indices = []
                
                for dept in departments:
                    # Check if this department exists in our timetables
                    if dept not in all_department_timetables:
                        continue
                        
                    if semester not in all_department_timetables[dept]:
                        continue
                        
                    # For each section in this department
                    for section_key in all_department_timetables[dept][semester]:
                        timetable = all_department_timetables[dept][semester][section_key]
                        
                        # Check consecutive slots using BASKET_TUTORIAL_MIN for basket tutorials
                        accumulated = 0
                        i = start_idx
                        temp_slot_indices = []
                        
                        while i < len(TIME_SLOTS) and accumulated < BASKET_TUTORIAL_MIN:
                            if is_minor_slot(TIME_SLOTS[i]):
                                valid_for_all = False
                                break
                            if is_break_time_slot(TIME_SLOTS[i], semester, 'TUT', config):
                                valid_for_all = False
                                break
                            if timetable[day][i]['type'] is not None:
                                valid_for_all = False
                                break
                            temp_slot_indices.append(i)
                            accumulated += slot_minutes(TIME_SLOTS[i])
                            i += 1
                        
                        if not valid_for_all or accumulated < BASKET_TUTORIAL_MIN:
                            valid_for_all = False
                            break
                        
                        # Store slot indices from the first valid check
                        if not slot_indices:
                            slot_indices = temp_slot_indices
                
                # Check for conflicts with existing basket slots
                if valid_for_all and slot_indices:
                    if check_basket_slot_conflict(day, slot_indices, existing_basket_slots):
                        valid_for_all = False
                
                if valid_for_all and room_pool and slot_indices:
                    random.shuffle(room_pool)
                    for room in room_pool:
                        if room not in room_schedule:
                            room_schedule[room] = {d: set() for d in range(len(config.get("days", DEFAULT_CONFIG["days"])))}
                        if all(si not in room_schedule[room][day] for si in slot_indices):
                            best_slot = (day, slot_indices)
                            best_room = room
                            break
                
                if best_slot:
                    break
            if best_slot:
                break
        
        if best_slot:
            scheduled_slots['tutorials'].append({'day': best_slot[0], 'slots': best_slot[1], 'room': best_room})
            for si in best_slot[1]:
                room_schedule[best_room][best_slot[0]].add(si)
            print(f"Scheduled {label} Tutorial {tut_idx + 1} on day {best_slot[0]} in room {best_room}")
        else:
            print(f"WARNING: Could not schedule {label} Tutorial {tut_idx + 1}")
    
    return scheduled_slots

def schedule_single_basket(semester, label, departments, lec_count, tut_count,
                          all_department_timetables, professor_schedule, room_schedule, course_room_mapping, config, TIME_SLOTS, large_rooms):
    """
    Schedule a single basket/elective group across multiple departments at the same time.
    (Legacy function kept for compatibility)
    """
    return schedule_single_basket_with_constraints(
        semester, label, departments, lec_count, tut_count,
        all_department_timetables, professor_schedule, room_schedule, course_room_mapping, [], config, TIME_SLOTS, large_rooms
    )

# ---------------------------
# Write electives to output
# ---------------------------
def write_electives_to_output(all_department_timetables, output_filename, electives_data):
    if not electives_data:
        print("No electives data to write.")
        return
    
    electives_wb = Workbook()
    if "Sheet" in electives_wb.sheetnames:
        electives_wb.remove(electives_wb["Sheet"])
    
    # Create a sheet for each department
    departments = set()
    for semester_data in all_department_timetables.values():
        departments.update(semester_data.keys())
    
    # Also include departments from electives data that might not have timetables
    if electives_data:
        for semester in electives_data:
            for basket in electives_data[semester]:
                # Add departments from basket config
                config = get_basket_config_for_semester(semester, electives_data)
                if config and 'departments' in config:
                    departments.update(config['departments'])
    
    for department in departments:
        ws = electives_wb.create_sheet(title=str(department))
        
        # Add headers
        headers = ["Semester", "Basket", "Electives"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_idx = 2
        
        # Add electives for each semester from ELECTIVES_DATA
        for semester in sorted(electives_data.keys()):
            for basket_label, electives_list in electives_data[semester].items():
                # Add semester
                ws.cell(row=row_idx, column=1, value=str(semester))
                
                # Add basket label
                ws.cell(row=row_idx, column=2, value=basket_label)
                
                # Add electives list
                electives_str = ", ".join(electives_list)
                ws.cell(row=row_idx, column=3, value=electives_str)
                
                # Add borders
                for col in range(1, 4):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                row_idx += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 50
    
    # Save the electives file
    electives_output_path = os.path.join(OUTPUT_DIR, "electives_output.xlsx")
    electives_wb.save(electives_output_path)
    print(f"Electives information saved to {electives_output_path}")

# ---------------------------
# Main generation function
# ---------------------------
def generate_all_timetables():
    # Load the latest configuration
    config = load_configuration()
    
    # Extract configuration values
    DAYS = config.get("days", DEFAULT_CONFIG["days"])
    LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    LAB_MIN = config.get("LAB_MIN", DEFAULT_CONFIG["LAB_MIN"])
    TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    SELF_STUDY_MIN = config.get("SELF_STUDY_MIN", DEFAULT_CONFIG["SELF_STUDY_MIN"])
    
    # Basket-specific durations - using same as regular
    BASKET_LECTURE_MIN = LECTURE_MIN
    BASKET_TUTORIAL_MIN = TUTORIAL_MIN
    
    # Parse break times
    MORNING_BREAK_START = parse_time_string(config.get("MORNING_BREAK_START", "10:30"))
    MORNING_BREAK_END = parse_time_string(config.get("MORNING_BREAK_END", "10:45"))
    LUNCH_BREAK_START = parse_time_string(config.get("LUNCH_BREAK_START", "13:00"))
    LUNCH_BREAK_END = parse_time_string(config.get("LUNCH_BREAK_END", "13:45"))
    LECTURE_TUTORIAL_BREAK_START = parse_time_string(config.get("LECTURE_TUTORIAL_BREAK_START", "15:30"))
    LECTURE_TUTORIAL_BREAK_END = parse_time_string(config.get("LECTURE_TUTORIAL_BREAK_END", "15:40"))
    
    # Print loaded configuration
    print("\n" + "="*60)
    print("LOADED CONFIGURATION")
    print("="*60)
    print(f"Lecture Duration: {LECTURE_MIN} minutes")
    print(f"Lab Duration: {LAB_MIN} minutes")
    print(f"Tutorial Duration: {TUTORIAL_MIN} minutes")
    print(f"Self-Study Duration: {SELF_STUDY_MIN} minutes")
    print(f"Morning Break: {MORNING_BREAK_START.strftime('%H:%M')} - {MORNING_BREAK_END.strftime('%H:%M')}")
    print(f"Lunch Break: {LUNCH_BREAK_START.strftime('%H:%M')} - {LUNCH_BREAK_END.strftime('%H:%M')}")
    print(f"Evening Break: {LECTURE_TUTORIAL_BREAK_START.strftime('%H:%M')} - {LECTURE_TUTORIAL_BREAK_END.strftime('%H:%M')}")
    print("="*60 + "\n")
    
    # Generate time slots from the latest configuration
    TIME_SLOTS = generate_time_slots(config)
    
    # Load data
    df, lecture_rooms, computer_lab_rooms, large_rooms = load_data()
    
    # Load electives data once
    electives_data = load_electives()
    
    room_schedule = {}
    professor_schedule = {}
    course_room_mapping = {}
    unscheduled_components = []

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    overview = wb.create_sheet("Overview")
    overview.append(["Combined Timetable for All Departments and Semesters"])
    overview.append(["Generated on:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    overview.append([])
    overview.append(["Department", "Semester", "Sheet Name"])
    row_index = 5

    SUBJECT_COLORS = [
        "FF6B6B", "4ECDC4", "FF9F1C", "5D5FEF", "45B7D1",
        "F72585", "7209B7", "3A0CA3", "4361EE", "4CC9F0",
        "06D6A0", "FFD166", "EF476F", "118AB2", "073B4C"
    ]

    # Create a structure to store all department timetables for basket scheduling
    all_department_timetables = {}
    
    # First pass: Initialize all timetables without scheduling anything
    for department in df['Department'].unique():
        all_department_timetables[department] = {}
        sems = sorted(df[df['Department'] == department]['Semester'].unique())
        
        for semester in sems:
            dept_upper = str(department).strip().upper()
            num_sections = 2 if (dept_upper == "CSE" and int(semester) in [1, 3, 5]) else 1
            
            all_department_timetables[department][semester] = {}
            
            for section in range(num_sections):
                section_key = chr(65 + section) if num_sections > 1 else 'A'
                all_department_timetables[department][semester][section_key] = {
                    d: {s: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} 
                        for s in range(len(TIME_SLOTS))} 
                    for d in range(len(DAYS))
                }
    
    # Second pass: Schedule basket slots for all semesters that need them
    for semester in [1, 3, 5]:  # Only these semesters have basket electives
        basket_config = get_basket_config_for_semester(semester, electives_data)
        if basket_config:
            print(f"\n=== Scheduling basket slots for Semester {semester} ===")
            print(f"Basket lecture duration: {BASKET_LECTURE_MIN} minutes")
            print(f"Basket tutorial duration: {BASKET_TUTORIAL_MIN} minutes")
            basket_slots = schedule_basket_slots(
                semester, all_department_timetables, professor_schedule, 
                room_schedule, course_room_mapping, config, TIME_SLOTS, electives_data, large_rooms
            )
            
            # Apply basket slots to all relevant departments
            departments = basket_config['departments']
            for dept in departments:
                if dept not in all_department_timetables or semester not in all_department_timetables[dept]:
                    continue
                    
                # For each section in this department
                for section_key in all_department_timetables[dept][semester]:
                    timetable = all_department_timetables[dept][semester][section_key]
                    
                    # Apply ALL basket slots (for Semester 5, this includes both B1 and B2)
                    if isinstance(basket_slots, dict) and 'B1' in basket_slots:
                        # Semester 5 case - apply both B1 and B2 to all sections
                        for basket_label in ['B1', 'B2']:
                            basket_data = basket_slots.get(basket_label, {})
                            print(f"Applying {basket_label} slots to {dept} Section {section_key}")
                            
                            # Fill in basket lectures
                            for lec_info in basket_data.get('lectures', []):
                                day = lec_info['day']
                                slots = lec_info['slots']
                                room = lec_info['room']
                                for idx, si in enumerate(slots):
                                    timetable[day][si]['type'] = 'LEC'
                                    timetable[day][si]['code'] = f"{basket_label} Courses" if idx == 0 else ''
                                    timetable[day][si]['name'] = '' if idx == 0 else ''
                                    timetable[day][si]['faculty'] = 'Multiple Faculty' if idx == 0 else ''
                                    timetable[day][si]['classroom'] = room if idx == 0 else ''
                            
                            # Fill in basket tutorials
                            for tut_info in basket_data.get('tutorials', []):
                                day = tut_info['day']
                                slots = tut_info['slots']
                                room = tut_info['room']
                                for idx, si in enumerate(slots):
                                    timetable[day][si]['type'] = 'TUT'
                                    timetable[day][si]['code'] = f"{basket_label} Courses" if idx == 0 else ''
                                    timetable[day][si]['name'] = '' if idx == 0 else ''
                                    timetable[day][si]['faculty'] = 'Multiple Faculty' if idx == 0 else ''
                                    timetable[day][si]['classroom'] = room if idx == 0 else ''
                    else:
                        # Semesters 1 and 3 case
                        basket_label = 'ELECTIVE'
                        basket_data = basket_slots
                        
                        # Fill in basket lectures
                        for lec_info in basket_data.get('lectures', []):
                            day = lec_info['day']
                            slots = lec_info['slots']
                            room = lec_info['room']
                            for idx, si in enumerate(slots):
                                timetable[day][si]['type'] = 'LEC'
                                timetable[day][si]['code'] = f"{basket_label} Courses" if idx == 0 else ''
                                timetable[day][si]['name'] = '' if idx == 0 else ''
                                timetable[day][si]['faculty'] = 'Multiple Faculty' if idx == 0 else ''
                                timetable[day][si]['classroom'] = room if idx == 0 else ''
                        
                        # Fill in basket tutorials
                        for tut_info in basket_data.get('tutorials', []):
                            day = tut_info['day']
                            slots = tut_info['slots']
                            room = tut_info['room']
                            for idx, si in enumerate(slots):
                                timetable[day][si]['type'] = 'TUT'
                                timetable[day][si]['code'] = f"{basket_label} Courses" if idx == 0 else ''
                                timetable[day][si]['name'] = '' if idx == 0 else ''
                                timetable[day][si]['faculty'] = 'Multiple Faculty' if idx == 0 else ''
                                timetable[day][si]['classroom'] = room if idx == 0 else ''
    
    # Third pass: Schedule all other courses
    for department in df['Department'].unique():
        sems = sorted(df[df['Department'] == department]['Semester'].unique())
        
        for semester in sems:
            dept_upper = str(department).strip().upper()
            num_sections = 2 if (dept_upper == "CSE" and int(semester) in [1, 3, 5]) else 1

            courses = df[(df['Department'] == department) & (df['Semester'] == semester)]
            if 'Schedule' in courses.columns:
                courses = courses[(courses['Schedule'].fillna('Yes').str.upper() == 'YES') | (courses['Schedule'].isna())]
            if courses.empty:
                continue

            # Check if this is a basket semester
            basket_config = get_basket_config_for_semester(int(semester), electives_data)
            is_basket_semester = basket_config is not None and dept_upper in basket_config['departments']
            
            # Split courses
            if 'P' in courses.columns:
                lab_courses = courses[courses['P'] > 0].copy()
                non_lab_courses = courses[courses['P'] == 0].copy()
            else:
                lab_courses = courses.head(0)
                non_lab_courses = courses.copy()

            if not lab_courses.empty:
                lab_courses['priority'] = lab_courses.apply(get_course_priority, axis=1)
                lab_courses = lab_courses.sort_values('priority', ascending=False)
            non_lab_courses['priority'] = non_lab_courses.apply(get_course_priority, axis=1)
            non_lab_courses = non_lab_courses.sort_values('priority', ascending=False)

            combined = pd.concat([lab_courses, non_lab_courses])
            combined['is_elective'] = combined.apply(is_elective, axis=1)
            
            # Filter out basket electives since they're already scheduled
            if is_basket_semester:
                combined = combined[~combined['is_elective']]
                
            courses_combined = combined.sort_values(by=['is_elective', 'priority'], ascending=[False, False]).drop_duplicates()

            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65 + section)}"
                ws = wb.create_sheet(title=section_title)

                overview.cell(row=row_index, column=1, value=department)
                overview.cell(row=row_index, column=2, value=str(semester))
                overview.cell(row=row_index, column=3, value=section_title)
                row_index += 1

                section_key = chr(65 + section) if num_sections > 1 else 'A'
                timetable = all_department_timetables[department][semester][section_key]

                section_subject_color = {}
                color_iter = iter(SUBJECT_COLORS)
                course_faculty_map = {}

                for _, c in courses_combined.iterrows():
                    code = str(c.get('Course Code', '')).strip()
                    if code and code not in section_subject_color:
                        try:
                            section_subject_color[code] = next(color_iter)
                        except StopIteration:
                            section_subject_color[code] = random.choice(SUBJECT_COLORS)
                        # Use section-specific faculty selection
                        course_faculty_map[code] = select_faculty_for_section(c.get('Faculty', 'TBD'), section_key)

                # Schedule non-basket courses
                for _, course in courses_combined.iterrows():
                    code = str(course.get('Course Code', '')).strip()
                    name = str(course.get('Course Name', '')).strip()
                    # Select faculty for this specific section
                    faculty = select_faculty_for_section(course.get('Faculty', 'TBD'), section_key)

                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {d: set() for d in range(len(DAYS))}

                    # Use the new calculate_required_sessions function
                    lec_sessions, tut_sessions, lab_sessions, ss_sessions, lab_duration = calculate_required_sessions(course, config)
                    room_type = get_required_room_type(course)

                    def schedule_component(required_minutes, comp_type, attempts_limit=10000):
                        for attempt in range(attempts_limit):
                            day = random.randint(0, len(DAYS)-1)
                            starts = get_all_possible_start_indices_for_duration(comp_type, TIME_SLOTS)  # Use component-specific slot filtering
                            for start_idx in starts:
                                # Check for LEC + TUT conflict on the same day
                                if check_course_component_conflict(timetable, day, code, comp_type, TIME_SLOTS):
                                    continue
                                
                                slot_indices, candidate_room = find_consecutive_slots_for_minutes(
                                    timetable, day, start_idx, required_minutes, semester,
                                    professor_schedule, faculty, room_schedule, room_type,
                                    code, course_room_mapping, comp_type, config, TIME_SLOTS, lecture_rooms, computer_lab_rooms)

                                if slot_indices is None:
                                    continue
                                if not check_professor_availability(professor_schedule, faculty, day, slot_indices[0], len(slot_indices), TIME_SLOTS):
                                    continue
                                if candidate_room is None:
                                    continue
                                    
                                for si_idx, si in enumerate(slot_indices):
                                    timetable[day][si]['type'] = 'LEC' if comp_type == 'LEC' else ('LAB' if comp_type == 'LAB' else ('TUT' if comp_type == 'TUT' else 'SS'))
                                    timetable[day][si]['code'] = code if si_idx == 0 else ''
                                    timetable[day][si]['name'] = name if si_idx == 0 else ''
                                    timetable[day][si]['faculty'] = faculty if si_idx == 0 else ''
                                    timetable[day][si]['classroom'] = candidate_room if si_idx == 0 else ''
                                    professor_schedule[faculty][day].add(si)
                                    if candidate_room not in room_schedule:
                                        room_schedule[candidate_room] = {d: set() for d in range(len(DAYS))}
                                    room_schedule[candidate_room][day].add(si)
                                return True
                        return False

                    # Schedule the calculated number of sessions
                    for _ in range(lec_sessions):
                        ok = schedule_component(LECTURE_MIN, 'LEC', attempts_limit=800)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LEC', section, "Number of collisions exceeded limit")

                    for _ in range(tut_sessions):
                        ok = schedule_component(TUTORIAL_MIN, 'TUT', attempts_limit=600)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'TUT', section, "No slot available")

                    for _ in range(lab_sessions):
                        # Use the calculated lab duration based on P value
                        ok = schedule_component(lab_duration, 'LAB', attempts_limit=800)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'LAB', section, "Lab not scheduled")

                    for _ in range(ss_sessions):
                        ok = schedule_component(SELF_STUDY_MIN, 'SS', attempts_limit=400)
                        if not ok:
                            add_unscheduled_course(unscheduled_components, department, semester, code, name, faculty, 'SS', section, "Self-study not scheduled")

                # Get basket slots for display
                basket_slots = {}
                if is_basket_semester:
                    if isinstance(basket_slots, dict) and 'B1' in basket_slots:
                        # Semester 5 case - include both B1 and B2
                        basket_slots = {}
                        for basket_label in ['B1', 'B2']:
                            basket_slots[basket_label] = {'lectures': [], 'tutorials': []}
                    else:
                        # Semesters 1 and 3 case
                        basket_slots = {'lectures': [], 'tutorials': []}

                # Write to Excel sheet
                write_timetable_to_sheet(ws, timetable, courses_combined, section_subject_color, 
                                        course_faculty_map, course_room_mapping, semester, is_basket_semester, basket_slots, config, TIME_SLOTS)

    # Format overview sheet
    for col in range(1, 4):
        overview.column_dimensions[get_column_letter(col)].width = 20
    for row_ in overview.iter_rows(min_row=1, max_row=4):
        for cell in row_:
            cell.font = Font(bold=True)
    for cell in overview[4]:
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    for row_ in overview.iter_rows(min_row=5, max_row=row_index-1):
        for cell in row_:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))

    # Save workbook
    out_filename = os.path.join(OUTPUT_DIR, "timetable_all_departments.xlsx")
    try:
        wb.save(out_filename)
        print(f"\nCombined timetable saved as {out_filename}")
    except Exception as e:
        print(f"Failed to save: {e}")
        traceback.print_exc()

    try:
        create_teacher_and_unscheduled_from_combined(out_filename, unscheduled_components, config)
    except Exception as e:
        print("Failed to generate teacher/unscheduled workbooks:", e)
        traceback.print_exc()

    # Write electives information to a separate file
    try:
        write_electives_to_output(all_department_timetables, out_filename, electives_data)
    except Exception as e:
        print("Failed to write electives output:", e)
        traceback.print_exc()

    return out_filename

def write_timetable_to_sheet(ws, timetable, courses_combined, section_subject_color, 
                             course_faculty_map, course_room_mapping, semester, 
                             is_basket_semester, basket_slots, config, TIME_SLOTS):
    """Write timetable data to Excel sheet with proper formatting"""
    
    DAYS = config.get("days", DEFAULT_CONFIG["days"])
    BASKET_LECTURE_MIN = config.get("LECTURE_MIN", DEFAULT_CONFIG["LECTURE_MIN"])
    BASKET_TUTORIAL_MIN = config.get("TUTORIAL_MIN", DEFAULT_CONFIG["TUTORIAL_MIN"])
    
    header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
    ws.append(header)
    
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    lec_fill = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")
    lab_fill = PatternFill(start_color="7CFC00", end_color="7CFC00", fill_type="solid")
    tut_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    ss_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    break_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    minor_fill = PatternFill(start_color="9ACD32", end_color="9ACD32", fill_type="solid")
    basket_fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")

    for day_idx, day_name in enumerate(DAYS):
        ws.append([day_name] + [''] * len(TIME_SLOTS))
        row_num = ws.max_row
        merges = []
        
        for slot_idx in range(len(TIME_SLOTS)):
            cell_obj = ws.cell(row=row_num, column=slot_idx + 2)
            
            if is_minor_slot(TIME_SLOTS[slot_idx]):
                cell_obj.value = "Minor Slot"
                cell_obj.fill = minor_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue

            if is_break_time_slot(TIME_SLOTS[slot_idx], semester, config=config):
                cell_obj.value = "BREAK"
                cell_obj.fill = break_fill
                cell_obj.font = Font(bold=True)
                cell_obj.alignment = Alignment(horizontal='center', vertical='center')
                cell_obj.border = border
                continue

            if timetable[day_idx][slot_idx]['type'] is None:
                cell_obj.border = border
                continue

            typ = timetable[day_idx][slot_idx]['type']
            code = timetable[day_idx][slot_idx]['code']
            cls = timetable[day_idx][slot_idx]['classroom']
            fac = timetable[day_idx][slot_idx]['faculty']

            if code:
                span = [slot_idx]
                j = slot_idx + 1
                while j < len(TIME_SLOTS) and timetable[day_idx][j]['type'] is not None and timetable[day_idx][j]['code'] == '':
                    span.append(j)
                    j += 1
                
                # Check if this is a basket slot
                if 'Courses' in code:  # B1 Courses, B2 Courses, or ELECTIVE Courses
                    display = f"{code}\n{typ}\nRoom: {cls}\n{fac}"
                    fill = basket_fill
                else:
                    display = f"{typ}\nRoom: {cls}\n{fac}"
                    if code in section_subject_color:
                        subj_color = section_subject_color[code]
                        fill = PatternFill(start_color=subj_color, end_color=subj_color, fill_type="solid")
                    else:
                        fill = {'LEC': lec_fill, 'LAB': lab_fill, 'TUT': tut_fill, 'SS': ss_fill}.get(typ, lec_fill)

                cell_obj.value = display
                cell_obj.fill = fill
                merges.append((slot_idx + 2, slot_idx + 2 + len(span) - 1, display, fill))
            
            cell_obj.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            cell_obj.border = border

        for start_col, end_col, val, fill in merges:
            if end_col > start_col:
                rng = f"{get_column_letter(start_col)}{row_num}:{get_column_letter(end_col)}{row_num}"
                try:
                    ws.merge_cells(rng)
                    mc = ws[f"{get_column_letter(start_col)}{row_num}"]
                    mc.value = val
                    mc.fill = fill
                    mc.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    mc.border = border
                except Exception:
                    pass

    for col_idx in range(1, len(TIME_SLOTS)+2):
        try:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        except Exception:
            pass

    for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
        ws.row_dimensions[row[0].row].height = 40

    current_row = len(DAYS) + 4

    # Self-Study Only Courses section
    ss_courses = []
    for _, course in courses_combined.iterrows():
        l = int(course['L']) if pd.notna(course['L']) else 0
        t = int(course['T']) if pd.notna(course['T']) else 0
        p = int(course['P']) if pd.notna(course['P']) else 0
        s = int(course['S']) if pd.notna(course['S']) else 0
        if s > 0 and l == 0 and t == 0 and p == 0:
            ss_courses.append({
                'code': str(course['Course Code']),
                'name': str(course['Course Name']),
                'faculty': str(course['Faculty'])
            })

    if ss_courses:
        ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        headers = ['Course Code', 'Course Name', 'Faculty']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = Font(bold=True)
        current_row += 1

        for course in ss_courses:
            ws.cell(row=current_row, column=1, value=course['code'])
            ws.cell(row=current_row, column=2, value=course['name'])
            ws.cell(row=current_row, column=3, value=course['faculty'])
            current_row += 1

        current_row += 2

    # Add Basket Information if applicable
    electives_data = load_electives()
    if is_basket_semester and electives_data and semester in electives_data:
        ws.cell(row=current_row, column=1, value="Cross-Department Elective Information")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 2

        if 'B1' in electives_data[semester] and 'B2' in electives_data[semester]:
            # Semester 5 - show both baskets
            for basket_label in ['B1', 'B2']:
                ws.cell(row=current_row, column=1, value=f"{basket_label} Electives:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1
                
                # List the electives for this basket
                electives_list = electives_data[semester][basket_label]
                electives_str = ", ".join(electives_list)
                ws.cell(row=current_row, column=1, value=electives_str)
                current_row += 1
                
                ws.cell(row=current_row, column=1, value=f"Lectures: 2 ({BASKET_LECTURE_MIN} min each)")
                ws.cell(row=current_row, column=2, value=f"Tutorials: 1 ({BASKET_TUTORIAL_MIN} min each)")
                current_row += 1
                
                ws.cell(row=current_row, column=1, value="Shared across: CSE, DSAI, ECE")
                current_row += 2
        else:
            # Semesters 1 and 3
            ws.cell(row=current_row, column=1, value="ELECTIVE slots shared across CSE, DSAI, ECE")
            current_row += 1
            
            # List the electives
            for basket_label, electives_list in electives_data[semester].items():
                electives_str = ", ".join(electives_list)
                ws.cell(row=current_row, column=1, value=electives_str)
                current_row += 1
            
            ws.cell(row=current_row, column=1, value=f"Lectures: 2 ({BASKET_LECTURE_MIN} min each)")
            ws.cell(row=current_row, column=2, value=f"Tutorials: 1 ({BASKET_TUTORIAL_MIN} min each)")
            current_row += 2

    # Legend
    legend_title = ws.cell(row=current_row, column=1, value="Legend")
    legend_title.font = Font(bold=True, size=12)
    current_row += 2

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    legend_headers = ['Subject Code', 'Color', 'Subject Name', 'Faculty', 'LTPS', 'Room']
    for col, header in enumerate(legend_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.border = border
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    current_row += 1

    for code, color in section_subject_color.items():
        assigned_room = course_room_mapping.get(code, "—")
        if not assigned_room or assigned_room == "—":
            continue

        ws.row_dimensions[current_row].height = 30

        ltps_value = ""
        course_name = ''
        fac_name = ''
        
        for _, course_row in courses_combined.iterrows():
            if str(course_row['Course Code']) == code:
                l = str(int(course_row['L'])) if pd.notna(course_row['L']) else "0"
                t = str(int(course_row['T'])) if pd.notna(course_row['T']) else "0"
                p = str(int(course_row['P'])) if pd.notna(course_row['P']) else "0"
                s = str(int(course_row['S'])) if pd.notna(course_row['S']) and 'S' in course_row else "0"
                ltps_value = f"{l}-{t}-{p}-{s}"
                course_name = str(course_row['Course Name'])
                break

        if code in course_faculty_map:
            fac_name = course_faculty_map[code]

        cells = [
            (code, None),
            ('', PatternFill(start_color=color, end_color=color, fill_type="solid")),
            (course_name, None),
            (fac_name, None),
            (ltps_value, None),
            (assigned_room, None)
        ]

        for col, (value, fill) in enumerate(cells, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = border
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=2)

        current_row += 1

# ---------------------------
# Teacher and Unscheduled workbooks
# ---------------------------
def split_faculty_names(fac_str):
    if fac_str is None:
        return []
    s = str(fac_str).strip()
    if s == '' or s.lower() in ['nan', 'none']:
        return []
    parts = [s]
    for sep in ['/', ',', '&', ';']:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            break
    return parts if parts else [s]

def parse_cell_for_course(cell_value):
    if cell_value is None:
        return ('', '', '', '')
    text = str(cell_value).strip()
    if text == '':
        return ('', '', '', '')

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    faculty = ''
    room = ''
    code = ''
    typ = ''

    for ln in lines:
        if 'room' in ln.lower() or 'Room:' in ln:
            parts = ln.split(':')
            if len(parts) >= 2:
                room = parts[-1].strip()

    if len(lines) >= 1:
        last = lines[-1]
        if 'room' not in last.lower() and 'courses' not in last.lower() and ':' not in last:
            faculty = last

    first = lines[0] if lines else ''
    if first:
        tokens = first.split()
        if len(tokens) >= 2 and tokens[1].upper() in ['LEC', 'LAB', 'TUT', 'SS']:
            code = tokens[0].strip()
            typ = tokens[1].strip().upper()
        else:
            code = tokens[0].strip() if tokens else ''
            for t in ['LEC', 'LAB', 'TUT', 'SS']:
                if t in text.upper():
                    typ = t
                    break

    if not faculty and len(lines) >= 2:
        for cand in lines[1:]:
            if any(ch.isalpha() for ch in cand) and 'room' not in cand.lower() and 'courses' not in cand.lower() and ':' not in cand:
                faculty = cand
                break

    return (code, typ, room, faculty)

def create_teacher_and_unscheduled_from_combined(timetable_filename, unscheduled_components, config):
    try:
        wb = load_workbook(timetable_filename, data_only=True)
    except Exception as e:
        print(f"Failed to open {timetable_filename}: {e}")
        return

    teacher_slots = {}
    slot_headers = []

    for sheetname in wb.sheetnames:
        if sheetname.lower() == 'overview':
            continue
        ws = wb[sheetname]
        header = [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value else '' for c in range(2, ws.max_column + 1)]
        if len(header) > len(slot_headers):
            slot_headers = header
        
        for r in range(2, ws.max_row + 1):
            day = ws.cell(r, 1).value
            if not day or str(day) not in config.get("days", DEFAULT_CONFIG["days"]):
                break
            day_idx = config.get("days", DEFAULT_CONFIG["days"]).index(day)
            
            for c in range(2, ws.max_column + 1):
                code, typ, room, faculty = parse_cell_for_course(ws.cell(r, c).value)
                for f in split_faculty_names(faculty):
                    if not f:
                        continue
                    if str(f).strip().upper() in ["BREAK", "MINOR SLOT", "NAN", "NONE", "", "MULTIPLE FACULTY"]:
                        continue

                    teacher_slots.setdefault(f, {d: {i: '' for i in range(len(slot_headers))} for d in range(len(config.get("days", DEFAULT_CONFIG["days"])))})
                    teacher_slots[f][day_idx][c - 2] = f"{code} {typ}\n({sheetname})\nRoom: {room}" if code else ''

    # Create teacher workbook
    twb = Workbook()
    if "Sheet" in twb.sheetnames:
        twb.remove(twb["Sheet"])

    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    alt_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for teacher in sorted(teacher_slots.keys()):
        safe_name = teacher[:31] or "Unknown"
        ws = twb.create_sheet(title=safe_name)

        ws.merge_cells("A1:{}1".format(get_column_letter(len(slot_headers) + 1)))
        title_cell = ws.cell(row=1, column=1, value=f"{teacher} — Weekly Timetable")
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.append(["Day"] + slot_headers)
        for cell in ws[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for d, day in enumerate(config.get("days", DEFAULT_CONFIG["days"])):
            row = [day] + [teacher_slots[teacher][d][i] for i in range(len(slot_headers))]
            ws.append(row)
            row_idx = ws.max_row
            if d % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill
            for cell in ws[row_idx]:
                cell.alignment = cell_align
                cell.border = border
            ws.row_dimensions[row_idx].height = 35

        ws.column_dimensions["A"].width = 15
        for col in range(2, len(slot_headers) + 2):
            ws.column_dimensions[get_column_letter(col)].width = 20

    twb.save(os.path.join(OUTPUT_DIR, "teacher_timetables.xlsx"))
    print("Saved teacher_timetables.xlsx")

    # Unscheduled workbook
    uwb = Workbook()
    ws = uwb.active
    ws.title = "Unscheduled Courses"

    headers = ["Course Code", "Department", "Semester", "Reason"]
    ws.append(headers)

    unscheduled_unique = {}
    for u in unscheduled_components:
        if u.code not in unscheduled_unique:
            reason_text = str(u.reason).strip() if hasattr(u, "reason") and u.reason else "Scheduling conflict"
            unscheduled_unique[u.code] = {
                "Course Code": u.code,
                "Department": u.department,
                "Semester": u.semester,
                "Reason": reason_text
            }

    for entry in unscheduled_unique.values():
        ws.append([entry[h] for h in headers])

    uwb.save(os.path.join(OUTPUT_DIR, "unscheduled_courses.xlsx"))
    print(f"Saved unscheduled_courses.xlsx with {len(unscheduled_unique)} unique courses")

if __name__ == "__main__":
    try:
        generate_all_timetables()
    except Exception as e:
        print("Error running TT_gen:", e)
        traceback.print_exc()