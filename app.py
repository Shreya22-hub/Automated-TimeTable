from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
import os
import json
import pandas as pd
import zipfile
from io import BytesIO
from datetime import time
import traceback
import importlib
import sys

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "inputs")
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
CONFIG_PATH = os.path.join(BASE_DIR, "config.json")

# Ensure directories exist
os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Default time slots
DEFAULT_TIME_SLOTS = [
    ["07:30", "09:00"],
    ["09:00", "09:30"],
    ["09:30", "10:00"],
    ["10:00", "10:30"],
    ["10:30", "10:45"],
    ["10:45", "11:00"],
    ["11:00", "11:30"],
    ["11:30", "12:00"],
    ["12:00", "12:15"],
    ["12:15", "12:30"],
    ["12:30", "13:15"],
    ["13:15", "13:30"],
    ["13:30", "14:00"],
    ["14:00", "14:30"],
    ["14:30", "15:00"],
    ["15:00", "15:30"],
    ["15:30", "15:40"],
    ["15:40", "16:00"],
    ["16:00", "16:30"],
    ["16:30", "17:00"],
    ["17:00", "17:30"],
    ["17:30", "18:00"],
    ["18:00", "18:30"],
    ["18:30", "23:59"]
]

# Load or create config
def load_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, 'r') as f:
            return json.load(f)
    return {
        "days": ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday'],
        "LECTURE_MIN": 90,
        "LAB_MIN": 120,
        "TUTORIAL_MIN": 60,
        "SELF_STUDY_MIN": 60,
        "MORNING_BREAK_START": "10:30",
        "MORNING_BREAK_END": "10:45",
        "LUNCH_BREAK_START": "13:00",
        "LUNCH_BREAK_END": "13:45",
        "LECTURE_TUTORIAL_BREAK_START": "15:30",
        "LECTURE_TUTORIAL_BREAK_END": "15:40",
        "TIME_SLOTS": DEFAULT_TIME_SLOTS,
        "USE_CUSTOM_SLOTS": False
    }

def save_config(config):
    with open(CONFIG_PATH, 'w') as f:
        json.dump(config, f, indent=4)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/courses')
def get_courses():
    """Get all courses from combined.csv"""
    try:
        csv_path = os.path.join(INPUT_DIR, 'combined.csv')
        if not os.path.exists(csv_path):
            return jsonify({'error': 'combined.csv not found'}), 404
        
        df = pd.read_csv(csv_path)
        # Replace NaN with None for proper JSON serialization
        df = df.where(pd.notna(df), None)
        courses = df.to_dict('records')
        
        # Group by department and semester
        grouped = {}
        for course in courses:
            dept = course.get('Department', 'Unknown')
            sem = course.get('Semester', 'Unknown')
            key = f"{dept} - Semester {sem}"
            if key not in grouped:
                grouped[key] = []
            # Convert None values to empty strings for display
            clean_course = {k: (v if v is not None else '') for k, v in course.items()}
            grouped[key].append(clean_course)
        
        return jsonify({'courses': grouped})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/electives')
def get_electives():
    """Get electives from elective.csv or elective.xlsx"""
    try:
        electives_csv = os.path.join(INPUT_DIR, "elective.csv")
        electives_xlsx = os.path.join(INPUT_DIR, "elective.xlsx")
        
        df = None
        if os.path.exists(electives_csv):
            df = pd.read_csv(electives_csv)
        elif os.path.exists(electives_xlsx):
            df = pd.read_excel(electives_xlsx)
        else:
            return jsonify({'error': 'No electives file found'}), 404
        
        # Replace NaN with None for proper JSON serialization
        df = df.where(pd.notna(df), None)
        electives = df.to_dict('records')
        # Convert None values to empty strings for display
        clean_electives = [{k: (v if v is not None else '') for k, v in elective.items()} for elective in electives]
        
        return jsonify({'electives': clean_electives})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/config', methods=['GET', 'POST'])
def handle_config():
    """Get or update configuration"""
    if request.method == 'GET':
        config = load_config()
        return jsonify(config)
    else:
        try:
            data = request.json
            config = load_config()
            
            # Update durations
            if 'LECTURE_MIN' in data:
                config['LECTURE_MIN'] = int(data['LECTURE_MIN'])
            if 'LAB_MIN' in data:
                config['LAB_MIN'] = int(data['LAB_MIN'])
            if 'TUTORIAL_MIN' in data:
                config['TUTORIAL_MIN'] = int(data['TUTORIAL_MIN'])
            if 'SELF_STUDY_MIN' in data:
                config['SELF_STUDY_MIN'] = int(data['SELF_STUDY_MIN'])
            
            # Update break times
            if 'MORNING_BREAK_START' in data:
                config['MORNING_BREAK_START'] = data['MORNING_BREAK_START']
            if 'MORNING_BREAK_END' in data:
                config['MORNING_BREAK_END'] = data['MORNING_BREAK_END']
            if 'LUNCH_BREAK_START' in data:
                config['LUNCH_BREAK_START'] = data['LUNCH_BREAK_START']
            if 'LUNCH_BREAK_END' in data:
                config['LUNCH_BREAK_END'] = data['LUNCH_BREAK_END']
            if 'LECTURE_TUTORIAL_BREAK_START' in data:
                config['LECTURE_TUTORIAL_BREAK_START'] = data['LECTURE_TUTORIAL_BREAK_START']
            if 'LECTURE_TUTORIAL_BREAK_END' in data:
                config['LECTURE_TUTORIAL_BREAK_END'] = data['LECTURE_TUTORIAL_BREAK_END']
            
            # Update time slots
            if 'TIME_SLOTS' in data:
                config['TIME_SLOTS'] = data['TIME_SLOTS']
            if 'USE_CUSTOM_SLOTS' in data:
                config['USE_CUSTOM_SLOTS'] = bool(data['USE_CUSTOM_SLOTS'])
            
            save_config(config)
            return jsonify({'success': True, 'config': config})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

@app.route('/api/generate', methods=['POST'])
def generate_timetable():
    """Generate timetables"""
    try:
        # Reload the TT_gen module to pick up the latest configuration
        if 'TT_gen' in sys.modules:
            importlib.reload(sys.modules['TT_gen'])
        
        # Import and run the generator
        import TT_gen
        output_file = TT_gen.generate_all_timetables()
        return jsonify({'success': True, 'message': 'Timetables generated successfully'})
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/outputs')
def list_outputs():
    """List all output files"""
    try:
        files = []
        if os.path.exists(OUTPUT_DIR):
            for filename in os.listdir(OUTPUT_DIR):
                filepath = os.path.join(OUTPUT_DIR, filename)
                if os.path.isfile(filepath):
                    size = os.path.getsize(filepath)
                    files.append({
                        'name': filename,
                        'size': f"{size / 1024:.2f} KB"
                    })
        return jsonify({'files': files})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download/<filename>')
def download_file(filename):
    """Download a specific output file"""
    try:
        return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/api/download/all')
def download_all():
    """Download all outputs as a zip file"""
    try:
        memory_file = BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            for filename in os.listdir(OUTPUT_DIR):
                filepath = os.path.join(OUTPUT_DIR, filename)
                if os.path.isfile(filepath):
                    zf.write(filepath, filename)
        
        memory_file.seek(0)
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name='timetable_outputs.zip'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/teachers')
def get_teachers():
    """Get list of teachers from teacher_timetables.xlsx"""
    try:
        teacher_file = os.path.join(OUTPUT_DIR, 'teacher_timetables.xlsx')
        if not os.path.exists(teacher_file):
            return jsonify({'error': 'Teacher timetables not found. Generate timetables first.'}), 404
        
        from openpyxl import load_workbook
        wb = load_workbook(teacher_file)
        teachers = [sheet for sheet in wb.sheetnames]
        wb.close()
        
        return jsonify({'teachers': teachers})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download/teacher/<teacher_name>')
def download_teacher(teacher_name):
    """Download a specific teacher's timetable"""
    try:
        from openpyxl import load_workbook, Workbook
        
        teacher_file = os.path.join(OUTPUT_DIR, 'teacher_timetables.xlsx')
        if not os.path.exists(teacher_file):
            return jsonify({'error': 'Teacher timetables not found'}), 404
        
        wb = load_workbook(teacher_file)
        
        if teacher_name not in wb.sheetnames:
            return jsonify({'error': 'Teacher not found'}), 404
        
        # Create a new workbook with just this teacher's sheet
        new_wb = Workbook()
        new_wb.remove(new_wb.active)
        
        source_sheet = wb[teacher_name]
        target_sheet = new_wb.create_sheet(teacher_name)
        
        # Copy all cells
        from copy import copy
        for row in source_sheet.iter_rows():
            for cell in row:
                target_cell = target_sheet[cell.coordinate]
                target_cell.value = cell.value
                if cell.has_style:
                    target_cell.font = copy(cell.font)
                    target_cell.border = copy(cell.border)
                    target_cell.fill = copy(cell.fill)
                    target_cell.number_format = cell.number_format
                    target_cell.protection = copy(cell.protection)
                    target_cell.alignment = copy(cell.alignment)
        
        # Copy column dimensions
        for col in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
        
        # Copy row dimensions
        for row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
        
        # Copy merged cells
        for merged_cell_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_cell_range))
        
        wb.close()
        
        # Save to BytesIO
        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{teacher_name}_timetable.xlsx'
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)