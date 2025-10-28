import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import math

# -----------------------------
# Step 1: Read Inputs
# -----------------------------
df_strength = pd.read_excel("BranchStrength.xlsx")
df_courses = pd.read_excel("CoursesPerYear.xlsx")
df_common = pd.read_excel("CommonCourse.xlsx")
df_settings = pd.read_excel("Settings.xlsx")
df_faculty = pd.read_csv("FACULTY.csv")
df_room = pd.read_excel("rooms.xlsx")
course_book = pd.read_excel("courselist.xlsx", sheet_name=None)
# -----------------------------
# Step 2: Clean Inputs
# -----------------------------
df_strength["Year"] = df_strength["Year"].astype(str).str.strip().str.title()
df_strength["Branch"] = df_strength["Branch"].astype(str).str.strip().str.upper()
df_courses["Year"] = df_courses["Year"].astype(str).str.strip().str.title()
faculty_list = df_faculty["Name"].astype(str).tolist()
# -----------------------------
# Step 2: Map Sheets to Years Automatically
# -----------------------------
sheet_year_map = {}
for i, sheet_name in enumerate(course_book.keys()):
    if i == 0:
        sheet_year_map[sheet_name] = "1st"
    elif i == 1:
        sheet_year_map[sheet_name] = "2nd"
    elif i == 2:
        sheet_year_map[sheet_name] = "3rd"
    else:
        sheet_year_map[sheet_name] = f"{i+1}th"

# -----------------------------
# Step 3: Prepare Branch Courses
# -----------------------------
branch_courses = {}
for sheet_name, df in course_book.items():
    year = sheet_year_map[sheet_name]
    branch_courses[year] = {}
    for branch in ["CSE", "DSAI", "ECE"]:
        branch_courses[year][branch] = []
        for val in df[branch].dropna().astype(str):
            course_code, credits = val.split(",")
            branch_courses[year][branch].append({
                "course_code": course_code.strip(),
                "credits": int(credits.strip())
            })

# -----------------------------
# Step 3: Extract Settings
# -----------------------------
settings = dict(zip(df_settings["SettingName"], df_settings["Value"]))
credits_per_course = int(settings["CreditsPerCourse"])
max_students_per_slot = int(settings["MaxStudentsPerSlot"])
max_courses_per_slot = int(settings["MaxCoursesPerSlot"])
total_rooms = int(settings["TotalRooms"])
room_capacity_per_course = int(settings["RoomCapacityPerCourse"])

# -----------------------------
# Step 4: Python Structures
# -----------------------------
branches = df_strength["Branch"].unique().tolist()
years = df_strength["Year"].unique().tolist()

branch_strength = {
    year: dict(zip(
        df_strength[df_strength["Year"]==year]["Branch"],
        df_strength[df_strength["Year"]==year]["Strength"]
    ))
    for year in years
}

courses_per_year = dict(zip(df_courses["Year"], df_courses["CoursesPerYear"]))

common_course = {
    "course_code": df_common.loc[0, "CourseCode"],
    "credits": int(df_common.loc[0, "Credits"])
}

# -----------------------------
# Step 5: Initialize Schedule
# -----------------------------
days = ["Day " + str(i+1) for i in range(9)]
slots = ["Morning", "Evening"]

schedule = {day: {slot: {year: [] for year in years} for slot in slots} for day in days}
day_year_credits = {day: {year: {branch: 0 for branch in branches} for year in years} for day in days}
day_slot_total_students = {day: {slot:0 for slot in slots} for day in days}
day_slot_total_courses = {day: {slot:0 for slot in slots} for day in days}
remaining_courses = {year: {branch: courses_per_year[year] for branch in branches} for year in years}
common_assigned = {year: False for year in years}

# -----------------------------
# Step 6: Assign Common Course
# -----------------------------
for day in days:
    for year in years:
        if not common_assigned[year]:
            slot = min(slots, key=lambda s: day_slot_total_students[day][s])
            total_year_students = sum(branch_strength[year].values())
            if day_slot_total_students[day][slot] + total_year_students <= max_students_per_slot and day_slot_total_courses[day][slot] < max_courses_per_slot:
                schedule[day][slot][year].append({
                    "course_code": common_course["course_code"],
                    "credits": common_course["credits"],
                    "branch": "All",
                    "students": total_year_students,
                    "type": "Common"
                })
                for branch in branches:
                    day_year_credits[day][year][branch] += common_course["credits"]
                day_slot_total_students[day][slot] += total_year_students
                day_slot_total_courses[day][slot] += 1
                common_assigned[year] = True

# -----------------------------
# Step 7: Assign Main Courses
# -----------------------------
for day in days:
    for branch in branches:
        for year in years:
            if remaining_courses[year][branch] <= 0:
                continue
            slot = min(slots, key=lambda s: day_slot_total_students[day][s])
            if (day_slot_total_courses[day][slot] >= max_courses_per_slot or 
                day_year_credits[day][year][branch] + credits_per_course > 4 or
                day_slot_total_students[day][slot] + branch_strength[year][branch] > max_students_per_slot):
                continue
            course_number = courses_per_year[year] - remaining_courses[year][branch] + 1
            schedule[day][slot][year].append({
                "course_code": f"{branch}{year[0]}{course_number}",
                "credits": credits_per_course,
                "branch": branch,
                "students": branch_strength[year][branch],
                "type": "Main"
            })
            remaining_courses[year][branch] -= 1
            day_year_credits[day][year][branch] += credits_per_course
            day_slot_total_students[day][slot] += branch_strength[year][branch]
            day_slot_total_courses[day][slot] += 1

# -----------------------------
# Step 8: Prepare Exam Schedule DataFrame
# -----------------------------
rows = []
for day in days:
    row = {"Day": day}
    for slot in slots:
        for year in years:
            course_list = [f"{c['course_code']} ({c['students']} students)" for c in schedule[day][slot][year]]
            row[f"{slot} - {year}"] = ", ".join(course_list) if course_list else "Empty"
    for year in years:
        for branch in branches:
            row[f"Credits {year}-{branch}"] = day_year_credits[day][year][branch]
    for slot in slots:
        row[f"Total Students - {slot}"] = day_slot_total_students[day][slot]
    rows.append(row)

df_schedule = pd.DataFrame(rows)
# -----------------------------
# Step 9 & 10: Room Allocation, Faculty Assignment, and Student Range Display (Fixed)
# -----------------------------

# ---- Load student data ----
student_book = pd.read_excel("students.xlsx", sheet_name=None)
student_ids = {}

for sheet_name, df in student_book.items():
    year = sheet_name.replace(" Year", "")
    for branch in ["CSE", "DSAI", "ECE"]:
        ids = df[branch].dropna().astype(str).tolist()
        student_ids[(year, branch)] = ids

# ---- Prepare room and faculty lists ----
room_names = df_room["Room"].dropna().astype(str).tolist()
faculty_count = len(faculty_list)
faculty_index = 0  # keeps track of next faculty to assign globally
df_rooms_rows = []

for day in days:
    used_faculty_today = set()  # to prevent reuse in same day
    for slot in slots:
        # Faculty assignment for this slot
        assigned_faculty = []
        for _ in range(total_rooms):
            while faculty_list[faculty_index % faculty_count] in used_faculty_today:
                faculty_index += 1
            assigned_faculty.append(faculty_list[faculty_index % faculty_count])
            used_faculty_today.add(faculty_list[faculty_index % faculty_count])
            faculty_index += 1
        room_faculty_mapping = {room_names[i]: assigned_faculty[i] for i in range(total_rooms)}

        # ---- Initialize room states ----
        room_capacity_left = {room: room_capacity_per_course for room in room_names}
        room_courses = {room: [] for room in room_names}  # to allow 2 different courses per room
        room_students = {room: [] for room in room_names}
        courses_in_slot = []
        for year in years:
            courses_in_slot.extend(schedule[day][slot][year])

        # ---- Assign students course-wise ----
        for course in courses_in_slot:
            branch = course["branch"]
            year_digit = course["course_code"][len(branch)]  # extract like '1' from 'CSE11'
            year_map = {"1": "1st", "2": "2nd", "3": "3rd"}
            year_label = year_map.get(year_digit, "1st")

            ids = student_ids.get((year_label, branch), [])
            students_remaining = len(ids) if ids else course["students"]
            batch_size = 24
            idx = 0

            while students_remaining > 0:
                # Try to find a room that has space AND doesn’t already contain this course
                target_room = None
                for r in room_names:
                    if len(room_courses[r]) < 2 and course["course_code"] not in room_courses[r]:
                        target_room = r
                        break
                if not target_room:
                    break  # no free rooms available

                take = min(batch_size, students_remaining)
                start_idx = idx
                end_idx = min(idx + take, len(ids)) - 1
                if ids:
                    first_roll = ids[start_idx]
                    last_roll = ids[end_idx]
                    student_range = f"{first_roll}–{last_roll}"
                else:
                    student_range = f"{take} students"

                # Assign to this room
                room_courses[target_room].append(course["course_code"])
                room_students[target_room].append(student_range)
                room_capacity_left[target_room] -= take
                students_remaining -= take
                idx += take

        # ---- Collect final rows ----
        for r in room_names:
            for i, c in enumerate(room_courses[r]):
                df_rooms_rows.append({
                    "Day": day,
                    "Slot": slot,
                    "Course": c,
                    "Branch": "".join([x for x in c if not x.isdigit()])[:4],
                    "Students": room_students[r][i] if i < len(room_students[r]) else "",
                    "Rooms Assigned": r,
                    "Faculty": room_faculty_mapping[r]
                })

# ---- Convert to DataFrame ----
df_rooms = pd.DataFrame(df_rooms_rows)

# -----------------------------
# Step 10.2: Merge courses in same room with same faculty
# -----------------------------
from collections import defaultdict

merged_rows = []

# key = (Day, Slot, Room, Faculty), value = list of courses + students
room_dict = defaultdict(list)

for row in df_rooms_rows:
    key = (row["Day"], row["Slot"], row["Rooms Assigned"], row["Faculty"])
    course_info = f"{row['Course']} ({row['Students']})"
    room_dict[key].append(course_info)

# create merged rows
for key, course_list in room_dict.items():
    day, slot, room, faculty = key
    merged_rows.append({
        "Day": day,
        "Slot": slot,
        "Rooms Assigned": room,
        "Faculty": faculty,
        "Courses + Students": ", ".join(course_list)
    })

# final DataFrame
df_rooms_merged = pd.DataFrame(merged_rows)


# -----------------------------
# Step 11: Save to Excel
# -----------------------------
with pd.ExcelWriter("exam_schedule_with_rooms_faculty.xlsx", engine="openpyxl") as writer:
    df_schedule.to_excel(writer, sheet_name="Exam Schedule", index=False)
    for day in days:
        df_day = df_rooms[df_rooms["Day"]==day]
        df_day.to_excel(writer, sheet_name=f"Rooms-{day}", index=False)

# -----------------------------
# Step 12: Excel Formatting
# -----------------------------
wb = load_workbook("exam_schedule_with_rooms_faculty.xlsx")
for day in days:
    ws = wb[f"Rooms-{day}"]
    for i, col in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 5
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            ws.row_dimensions[cell.row].height = 30

# -----------------------------
# Step 13: Create Free Slot Sheet
# -----------------------------

free_rows = []

for day in days:
    for slot in slots:
        for year in years:
            for branch in branches:
                # check if any course is assigned to this year & branch in this day+slot
                courses_assigned = schedule[day][slot][year]
                assigned_branches = [c["branch"] for c in courses_assigned]
                
                # if branch not assigned, it is free
                if branch not in assigned_branches and "All" not in assigned_branches:
                    free_rows.append({
                        "Day": day,
                        "Slot": slot,
                        "Year": year,
                        "Branch": branch,
                        "Status": "Free"
                    })

df_free_slots = pd.DataFrame(free_rows)

# -----------------------------
# Step 14: Save Free Slot Sheet
# -----------------------------
with pd.ExcelWriter("exam_schedule_with_rooms_faculty.xlsx", engine="openpyxl", mode="a") as writer:
    df_free_slots.to_excel(writer, sheet_name="Free Slots", index=False)


wb.save("exam_schedule_with_rooms_faculty.xlsx")
print("Exam schedule, room allocation, and faculty assignment saved successfully!")