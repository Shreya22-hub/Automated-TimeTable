import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import math
from datetime import datetime, timedelta
import copy
from collections import defaultdict
import re
import os
import random

# Define the uploads folder
UPLOADS_FOLDER = "uploads"

# Global Helper Functions
def normalize_year(year_text):
    """Normalizes year strings to a standard format (e.g., '1St Year')"""
    text = str(year_text).lower().replace(" ", "").strip()

    if "1st" in text or "first" in text or text == "1":
        return "1St Year"
    elif "2nd" in text or "second" in text or text == "2":
        return "2Nd Year"
    elif "3rd" in text or "third" in text or text == "3":
        return "3Rd Year"
    elif "4th" in text or "fourth" in text or text == "4":
        return "4Th Year"
    elif "5th" in text or "fifth" in text or text == "5":
        return "5Th Year"
    else:
        # Return original text with proper title case
        return str(year_text).title()

def validate_input_files():
    """Validate that all required input files exist and have correct structure"""
    required_files = {
        "BranchStrength.xlsx": ["Year", "Branch", "Strength"],
        "CoursesPerYear.xlsx": ["Year", "CoursesPerYear"],
        "CommonCourse.xlsx": ["CourseCode", "Credits"],  # Will check for similar names
        "Settings.xlsx": ["SettingName", "Value"],
        "FACULTY.csv": ["Name"],
        "rooms.xlsx": ["Room", "Capacity"],
        "courselist.xlsx": None,  # Multiple sheets, will check separately
        "students.xlsx": None   # Multiple sheets, will check separately
    }
    
    print("\n[INFO] Validating Input Files:")
    print("="*50)
    
    for filename, required_cols in required_files.items():
        file_path = os.path.join(UPLOADS_FOLDER, filename)
        try:
            if filename.endswith('.xlsx'):
                df = pd.read_excel(file_path)
                print(f"\n[OK] {filename}: Found {len(df)} rows")
                print(f"  Columns: {df.columns.tolist()}")
                
                if required_cols:
                    missing_cols = [col for col in required_cols if col not in df.columns]
                    if missing_cols:
                        print(f"  [WARNING] Missing columns: {missing_cols}")
                    else:
                        print(f"  [OK] All required columns present")
                        
            elif filename.endswith('.csv'):
                df = pd.read_csv(file_path)
                print(f"\n[OK] {filename}: Found {len(df)} rows")
                print(f"  Columns: {df.columns.tolist()}")
                
                if required_cols:
                    missing_cols = [col for col in required_cols if col not in df.columns]
                    if missing_cols:
                        print(f"  [WARNING] Missing columns: {missing_cols}")
                    else:
                        print(f"  [OK] All required columns present")
                        
        except FileNotFoundError:
            print(f"\n[ERROR] {filename}: FILE NOT FOUND at {file_path}")
        except Exception as e:
            print(f"\n[ERROR] {filename}: ERROR - {e}")
    
    print("="*50)

def generate_timetable(start_date, end_date, branch_slot_allocation=None, max_credits_per_day=5, courses_per_room=2):
    """
    Generate exam timetable with given parameters
    
    Args:
        start_date (str): Start date in 'YYYY-MM-DD' format
        end_date (str): End date in 'YYYY-MM-DD' format
        branch_slot_allocation (dict): Branch allocation per year and slot (optional)
        max_credits_per_day (int): Maximum credits allowed per day per branch (default: 5)
        courses_per_room (int): Number of courses per room (default: 2)
    
    Returns:
        str: Path to generated Excel file
    """
    
    # Validate input files first
    validate_input_files()
    
    # Convert string dates to datetime
    start_date = datetime.strptime(start_date, "%Y-%m-%d")
    end_date = datetime.strptime(end_date, "%Y-%m-%d")
    
    # Public holidays in India (example, update as needed)
    public_holidays = [
        "2025-01-26", "2025-03-14", "2025-03-31", "2025-04-06", "2025-04-18",
        "2025-05-12", "2025-07-06", "2025-08-15", "2025-10-02", "2025-10-14", "2025-10-20"
    ]
    public_holidays = [datetime.strptime(d, "%Y-%m-%d") for d in public_holidays]
    
    # Generate list of valid exam dates (exclude Sundays & holidays)
    exam_dates = []
    current_date = start_date
    while current_date <= end_date:
        if current_date.weekday() != 6 and current_date not in public_holidays:
            exam_dates.append(current_date)
        current_date += timedelta(days=1)
    
    # -----------------------------
    # Step 1: Read Inputs
    # -----------------------------
    df_strength = pd.read_excel(os.path.join(UPLOADS_FOLDER, "BranchStrength.xlsx"))
    df_courses = pd.read_excel(os.path.join(UPLOADS_FOLDER, "CoursesPerYear.xlsx"))
    df_common = pd.read_excel(os.path.join(UPLOADS_FOLDER, "CommonCourse.xlsx"))
    df_settings = pd.read_excel(os.path.join(UPLOADS_FOLDER, "Settings.xlsx"))
    df_faculty = pd.read_csv(os.path.join(UPLOADS_FOLDER, "FACULTY.csv"))
    df_room = pd.read_excel(os.path.join(UPLOADS_FOLDER, "rooms.xlsx"))
    course_book = pd.read_excel(os.path.join(UPLOADS_FOLDER, "courselist.xlsx"), sheet_name=None)
    
    # -----------------------------
    # Step 2: Clean Inputs
    # -----------------------------
    df_strength["Year"] = df_strength["Year"].astype(str).str.strip().str.title()
    df_strength["Branch"] = df_strength["Branch"].astype(str).str.strip().str.upper()
    df_courses["Year"] = df_courses["Year"].astype(str).str.strip().str.title()
    faculty_list = df_faculty["Name"].astype(str).tolist()
    
    # -----------------------------
    # Step 2: Map Sheets to Years Automatically
    # -----------------------------
    sheet_year_map = {}
    year_names = ["1St Year", "2Nd Year", "3Rd Year", "4Th Year", "5Th Year"]
    
    for i, sheet_name in enumerate(course_book.keys()):
        if i < len(year_names):
            sheet_year_map[sheet_name] = year_names[i]
        else:
            sheet_year_map[sheet_name] = f"{i+1}Th Year"
    
    # -----------------------------
    # Step 3: Prepare Branch Courses
    # -----------------------------
    branch_courses = {}
    for sheet_name, df in course_book.items():
        year = sheet_year_map[sheet_name]
        branch_courses[year] = {}
        df.columns = [str(c).strip().upper() for c in df.columns]
        for branch in ["CSE", "DSAI", "ECE"]:
            branch_courses[year][branch] = []
            for val in df.get(branch, pd.Series()).dropna().astype(str):
                try:
                    course_code, credits = val.split(",")
                    branch_courses[year][branch].append({
                        "course_code": course_code.strip(),
                        "credits": int(credits.strip())
                    })
                except:
                    print(f"Skipping invalid course entry: {val}")
    
    # -----------------------------
    # Step 3: Extract Settings with Robust Error Handling
    # -----------------------------
    print("\n[INFO] Settings file information:")
    print(f"Columns in Settings: {df_settings.columns.tolist()}")
    print(f"First few rows:\n{df_settings.head()}")
    print("="*50)
    
    # Default settings
    default_settings = {
        "CreditsPerCourse": 4,
        "MaxStudentsPerSlot": 1000,
        "MaxCoursesPerSlot": 6,
        "TotalRooms": 24
    }
    
    settings = default_settings.copy()
    
    try:
        # Try to find the correct column names
        setting_name_col = None
        value_col = None
        
        for col in df_settings.columns:
            col_lower = str(col).lower()
            if setting_name_col is None and ('name' in col_lower or 'setting' in col_lower):
                setting_name_col = col
            if value_col is None and ('value' in col_lower or 'val' in col_lower):
                value_col = col
        
        if setting_name_col is None or value_col is None:
            print("[WARNING] Could not find required columns in Settings.xlsx")
            print("Expected columns: SettingName, Value (or similar)")
            print(f"Available columns: {df_settings.columns.tolist()}")
            print("Using default settings:")
            for key, value in default_settings.items():
                print(f"  {key}: {value}")
        else:
            print(f"Using columns: {setting_name_col} and {value_col}")
            
            # Read settings from file
            for _, row in df_settings.iterrows():
                setting_name = str(row[setting_name_col]).strip()
                setting_value = row[value_col]
                
                # Check if this setting matches any of our required settings
                for key in default_settings.keys():
                    if key.lower() in setting_name.lower() or setting_name.lower() in key.lower():
                        try:
                            settings[key] = int(setting_value)
                            print(f"  Found {key}: {setting_value}")
                        except (ValueError, TypeError):
                            print(f"  [WARNING] Invalid value for {key}: {setting_value}, using default: {default_settings[key]}")
                            settings[key] = default_settings[key]
                        break
            
            # Check for missing settings
            for key, default_value in default_settings.items():
                if key not in settings or settings[key] == default_value and key not in [str(row[setting_name_col]).strip() for _, row in df_settings.iterrows()]:
                    print(f"  [WARNING] Missing {key}, using default: {default_value}")
                    settings[key] = default_value
    
    except Exception as e:
        print(f"[WARNING] Error reading Settings.xlsx: {e}")
        print("Using default settings:")
        for key, value in default_settings.items():
            print(f"  {key}: {value}")
        settings = default_settings.copy()
    
    # Extract settings with defaults
    credits_per_course = settings.get("CreditsPerCourse", 4)
    max_students_per_slot = settings.get("MaxStudentsPerSlot", 1000)
    max_courses_per_slot = settings.get("MaxCoursesPerSlot", 6)
    total_rooms = settings.get("TotalRooms", 24)
    
    print(f"\nFinal settings:")
    print(f"  CreditsPerCourse: {credits_per_course}")
    print(f"  MaxStudentsPerSlot: {max_students_per_slot}")
    print(f"  MaxCoursesPerSlot: {max_courses_per_slot}")
    print(f"  TotalRooms: {total_rooms}")
    print("="*50)

    # Debug CommonCourse file
    print("\n[INFO] CommonCourse file information:")
    print(f"Columns in CommonCourse: {df_common.columns.tolist()}")
    print(f"First few rows:\n{df_common.head()}")
    print("="*50)

    # FIXED: Handle CommonCourse with proper error handling
    common_course = {}
    try:
        # Try different possible column names
        course_code_col = None
        credits_col = None
        
        for col in df_common.columns:
            col_lower = str(col).lower()
            if 'code' in col_lower or 'course' in col_lower:
                course_code_col = col
            if 'credit' in col_lower:
                credits_col = col
        
        if course_code_col is None or credits_col is None:
            print("[WARNING] Could not find required columns in CommonCourse.xlsx")
            print("Expected columns: CourseCode, Credits (or similar)")
            print(f"Available columns: {df_common.columns.tolist()}")
            # Use default values
            common_course = {
                "course_code": "ENV",
                "credits": 1
            }
        else:
            common_course = {
                "course_code": df_common.loc[0, course_code_col],
                "credits": int(df_common.loc[0, credits_col])
            }
            print(f"Found common course: {common_course}")
            
    except Exception as e:
        print(f"[WARNING] Error reading CommonCourse.xlsx: {e}")
        # Use default values
        common_course = {
            "course_code": "ENV",
            "credits": 1
        }
    
    # -----------------------------
    # Step 3.5: Dynamic Room Capacity Allocation (Fixed Logic)
    # -----------------------------
    # Read actual capacity per room from rooms.xlsx
    df_room["Capacity"] = pd.to_numeric(df_room["Capacity"], errors="coerce").fillna(0).astype(int)

    # --- Expand room ranges like "C403–C408" → C403, C404, ..., C408 ---
    expanded_rooms = []
    for _, row in df_room.iterrows():
        room_name = str(row["Room"]).strip().upper()
        capacity = int(row["Capacity"])

        # handle both hyphen and en-dash
        if "-" in room_name or "–" in room_name:
            parts = room_name.replace("–", "-").split("-")
            try:
                prefix = "".join([c for c in parts[0] if not c.isdigit()])
                start = int("".join([c for c in parts[0] if c.isdigit()]))
                end = int(parts[1])
                for num in range(start, end + 1):
                    expanded_rooms.append({"Room": f"{prefix}{num}", "Capacity": capacity})
            except Exception:
                expanded_rooms.append({"Room": room_name, "Capacity": capacity})
        else:
            expanded_rooms.append({"Room": room_name, "Capacity": capacity})

    df_room = pd.DataFrame(expanded_rooms)
    
    # --- Helper function to check if room is in C403-C408 range ---
    def is_special_room(room_name):
        """Check if room is in C403-C408 range"""
        match = re.match(r'C40([3-8])', room_name.strip().upper())
        return match is not None

    # --- Build per-room capacity map ---
    # Regular rooms: Each course gets FULL capacity
    # Special rooms (C403-C408): Capacity divided by courses_per_room
    room_capacity_map = {}
    for _, row in df_room.iterrows():
        room_name = row["Room"]
        base_capacity = int(row["Capacity"])
        
        if is_special_room(room_name):
            # Special rooms: divide capacity by courses_per_room
            capacity_per_course = base_capacity // courses_per_room
            room_capacity_map[room_name] = capacity_per_course
        else:
            # Regular rooms: full capacity per course
            room_capacity_map[room_name] = base_capacity

    print("\n[INFO] Room capacity mapping (students per course):")
    for r, cap in room_capacity_map.items():
        room_type = "Special (C403-C408)" if is_special_room(r) else "Regular"
        print(f"  {r} [{room_type}]: {cap} students per course")

    # Compute an average just for reference
    if room_capacity_map:
        room_capacity_per_course = int(sum(room_capacity_map.values()) / len(room_capacity_map))
    else:
        room_capacity_per_course = 24  # fallback

    print(f"\nConfiguration: {courses_per_room} courses per room")
    print(f"  Regular rooms: Full capacity per course")
    print(f"  Special rooms (C403-C408): Capacity divided by {courses_per_room} per course")
    
    # -----------------------------
    # Step 4: Python Structures (with debugging)
    # -----------------------------
    branches = df_strength["Branch"].unique().tolist()
    years = df_strength["Year"].unique().tolist()

    # Debug information
    print("\n[INFO] Year Information:")
    print(f"Years in BranchStrength: {df_strength['Year'].unique().tolist()}")
    print(f"Years in CoursesPerYear: {df_courses['Year'].unique().tolist()}")
    print(f"Years in branch_courses: {list(branch_courses.keys())}")
    print(f"Normalized years list: {years}")
    print("="*50)

    # Create branch_strength with both original and normalized year keys
    branch_strength = {}
    branch_strength_normalized = {}

    for year in years:
        norm_year = normalize_year(year)
        year_dict = dict(zip(
            df_strength[df_strength["Year"]==year]["Branch"],
            df_strength[df_strength["Year"]==year]["Strength"]
        ))
        
        # Store with both original and normalized keys
        branch_strength[year] = year_dict
        branch_strength_normalized[norm_year] = year_dict

    courses_per_year = dict(zip(df_courses["Year"], df_courses["CoursesPerYear"]))
    
    slot_max_students = total_rooms * 48

    # FIXED: Handle common_course_map with proper error handling
    common_course_map = {}
    try:
        if df_common.empty:
            print("[WARNING] CommonCourse.xlsx is empty")
        else:
            # Get column names dynamically
            course_code_col = None
            credits_col = None
            year_col = None
            branches_col = None
            
            for col in df_common.columns:
                col_lower = str(col).lower()
                if course_code_col is None and ('code' in col_lower or 'course' in col_lower):
                    course_code_col = col
                if credits_col is None and 'credit' in col_lower:
                    credits_col = col
                if year_col is None and 'year' in col_lower:
                    year_col = col
                if branches_col is None and ('branch' in col_lower or 'branches' in col_lower):
                    branches_col = col
            
            print(f"CommonCourse columns found: Course={course_code_col}, Credits={credits_col}, Year={year_col}, Branches={branches_col}")
            
            if course_code_col and credits_col:
                for _, row in df_common.iterrows():
                    year_norm = normalize_year(str(row[year_col]) if year_col else "1St Year")
                    branches_cell = row[branches_col] if branches_col else ""
                    
                    if pd.isna(branches_cell):
                        branches_for_course = []
                    else:
                        branches_for_course = [b.strip() for b in str(branches_cell).split(",")]
                    
                    common_course_map[row[course_code_col]] = {
                        "credits": int(row[credits_col]),
                        "Year": year_norm,
                        "Branches": branches_for_course
                    }
            else:
                print("[WARNING] Could not find required columns in CommonCourse.xlsx")
                # Use default common course
                common_course_map = {
                    common_course["course_code"]: {
                        "credits": common_course["credits"],
                        "Year": "1St Year",
                        "Branches": ["CSE", "DSAI", "ECE"]
                    }
                }
                
    except Exception as e:
        print(f"[WARNING] Error processing CommonCourse.xlsx: {e}")
        # Use default common course
        common_course_map = {
            common_course["course_code"]: {
                "credits": common_course["credits"],
                "Year": "1St Year",
                "Branches": ["CSE", "DSAI", "ECE"]
            }
        }

    common_assigned = {code: False for code in common_course_map}
    
    # -----------------------------
    # Step 5: Initialize Schedule (FIXED)
    # -----------------------------
    days = [d.strftime("%Y-%m-%d") for d in exam_dates]
    slots = ["Morning", "Evening"]
    schedule = {day: {slot: {year: [] for year in years} for slot in slots} for day in days}
    day_year_credits = {day: {year: {branch: 0 for branch in branches} for year in years} for day in days}
    day_slot_total_students = {day: {slot:0 for slot in slots} for day in days}
    day_slot_total_courses = {day: {slot:0 for slot in slots} for day in days}

    # FIXED: Initialize remaining_courses with proper error handling
    remaining_courses = {}
    for year in years:
        remaining_courses[year] = {}
        for branch in branches:
            # Get courses per year with fallback to 0 if year not found
            courses_count = courses_per_year.get(year, 0)
            remaining_courses[year][branch] = courses_count
    
    # -----------------------------
    # Step 6b: Assign Common Courses
    # -----------------------------
    for day in days:
        for course_code, info in common_course_map.items():
            year_norm = info['Year']
            branches_to_block = info['Branches']
            if common_assigned[course_code]:
                continue
            
            for slot in slots:
                if year_norm not in schedule[day][slot]:
                    schedule[day][slot][year_norm] = []
                if year_norm not in day_year_credits[day]:
                    day_year_credits[day][year_norm] = {b: 0 for b in branches}
                
                conflict = False
                for branch in branches_to_block:
                    if day_year_credits[day][year_norm].get(branch, 0) + info['credits'] > max_credits_per_day:
                        conflict = True
                        break
                    if any(c.get('branch') == branch for c in schedule[day][slot][year_norm]):
                        conflict = True
                        break
                if conflict:
                    continue
                
                total_students = sum(branch_strength_normalized[year_norm].get(b, 0) for b in branches_to_block)
                if day_slot_total_students[day][slot] + total_students > slot_max_students:
                    continue
                
                for branch in branches_to_block:
                    schedule[day][slot][year_norm].append({
                        "course_code": course_code,
                        "credits": info['credits'],
                        "branch": branch,
                        "year": year_norm,
                        "students": branch_strength_normalized[year_norm].get(branch, 0),
                        "type": "Common"
                    })
                    day_year_credits[day][year_norm][branch] += info['credits']
                    day_slot_total_students[day][slot] += branch_strength_normalized[year_norm].get(branch, 0)
                
                day_slot_total_courses[day][slot] += 1
                common_assigned[course_code] = True
                break
    
    # -----------------------------
    # Step7: Assign Main Courses
    # -----------------------------
    # Create default branch slot allocation if not provided
    if branch_slot_allocation is None:
        default_branch_slot_allocation = {}
        for year in years:
            norm_year = normalize_year(year)
            default_branch_slot_allocation[norm_year] = {
                "Morning": branches.copy(),
                "Evening": branches.copy()
            }
        branch_slot_allocation = default_branch_slot_allocation
    
    # Validate against slot capacity
    for year in branch_slot_allocation:
        normalized_year = normalize_year(year)
        for slot in branch_slot_allocation[year]:
            branches_in_slot = branch_slot_allocation[year][slot]
            if normalized_year not in branch_strength_normalized:
                continue
            total_strength = sum(branch_strength_normalized[normalized_year].get(b, 0) for b in branches_in_slot)
            if total_strength > slot_max_students:
                print(f"[WARNING] {normalized_year} {slot} total ({total_strength}) exceeds slot capacity ({slot_max_students}).")
    
    branch_slot_allocation_day = {day: copy.deepcopy(branch_slot_allocation) for day in days}
    
    for day in days:
        for slot in slots:
            for year in years:
                normalized_year = normalize_year(year)
                allowed_branches = branch_slot_allocation_day[day].get(normalized_year, {}).get(slot, [])
                
                running_total = day_slot_total_students[day][slot]
                
                final_branches = []
                for b in allowed_branches:
                    b_strength = branch_strength_normalized[normalized_year].get(b, 0)
                    if running_total + b_strength <= slot_max_students:
                        final_branches.append(b)
                        running_total += b_strength
                    else:
                        print(f"Skipping branch {b} for {normalized_year} {slot} on {day} — would exceed slot capacity")
                
                branch_slot_allocation_day[day][normalized_year][slot] = final_branches
    
    def total_remaining():
        return sum(remaining_courses[y][b] for y in years for b in branches)
    
    # Find ENV day
    env_day = None
    env_code = common_course["course_code"]
    for d in days:
        if any(
            c.get("course_code") == env_code
            for s in slots
            for y in years
            for c in schedule[d][s][y]
        ):
            env_day = d
            break
    
    after_env_days = days[days.index(env_day)+1:] if env_day in days else days[:]
    
    for day in days:
        day_allocation = branch_slot_allocation_day.get(day, branch_slot_allocation)
        
        for slot in slots:
            blocked_branches = set()
            for year in years:
                for c in schedule[day][slot][year]:
                    if c.get("type") == "Common":
                        if isinstance(c["branch"], str):
                            for b in c["branch"].split(","):
                                blocked_branches.add(b.strip())
            
            for year in years:
                normalized_year = normalize_year(year)
                
                if any(c.get("type") == "Common" for c in schedule[day][slot][year]):
                    continue
                
                allowed_branches = day_allocation.get(normalized_year, {}).get(slot, [])
                
                candidates = [
                    b for b in allowed_branches
                    if remaining_courses[year].get(b, 0) > 0
                    and b not in blocked_branches
                ]
                
                # FIXED: Use normalized branch strength
                candidates.sort(key=lambda b: branch_strength_normalized[normalized_year].get(b, 0), reverse=True)
                
                for b in candidates:
                    if day_year_credits[day][year].get(b, 0) + credits_per_course > max_credits_per_day:
                        continue
                    if day_slot_total_students[day][slot] + branch_strength_normalized[normalized_year].get(b, 0) > slot_max_students:
                        continue
                    
                    course_index = courses_per_year.get(year, 0) - remaining_courses[year].get(b, 0)
                    year_key = normalize_year(year)
                    if course_index < len(branch_courses.get(year_key, {}).get(b, [])):
                        course_info = branch_courses[year_key][b][course_index]
                    else:
                        course_info = {"course_code": f"{b}{year[0]}X", "credits": credits_per_course}
                    
                    schedule[day][slot][year].append({
                        "course_code": course_info["course_code"],
                        "credits": course_info["credits"],
                        "branch": b,
                        "year": year,
                        "students": branch_strength_normalized[normalized_year].get(b, 0),
                        "type": "Main"
                    })
                    
                    remaining_courses[year][b] -= 1
                    day_year_credits[day][year][b] += course_info["credits"]
                    day_slot_total_students[day][slot] += branch_strength_normalized[normalized_year].get(b, 0)
                    day_slot_total_courses[day][slot] += 1
            
            # Fill empty slots after ENV day
            if day in after_env_days and total_remaining() > 0:
                slot_has_main = any(
                    len([c for c in schedule[day][slot][y] if c.get("type") != "Common"]) > 0
                    for y in years
                )
                if not slot_has_main:
                    placed = False
                    year_order = sorted(
                        years, key=lambda Y: sum(remaining_courses[Y].get(b, 0) for b in branches), reverse=True
                    )
                    for year in year_order:
                        normalized_year = normalize_year(year)
                        if any(c.get("type") == "Common" for c in schedule[day][slot][year]):
                            continue
                        allowed_branches = day_allocation.get(normalized_year, {}).get(slot, [])
                        for b in allowed_branches:
                            if remaining_courses[year].get(b, 0) <= 0:
                                continue
                            if day_year_credits[day][year].get(b, 0) + credits_per_course > max_credits_per_day:
                                continue
                            if day_slot_total_students[day][slot] + branch_strength_normalized[normalized_year].get(b, 0) > slot_max_students:
                                continue
                            
                            course_index = courses_per_year.get(year, 0) - remaining_courses[year].get(b, 0)
                            if course_index < len(branch_courses.get(normalized_year, {}).get(b, [])):
                                course_info = branch_courses[normalized_year][b][course_index]
                            else:
                                course_info = {"course_code": f"{b}{year[0]}X", "credits": credits_per_course}
                            
                            schedule[day][slot][year].append({
                                "course_code": course_info["course_code"],
                                "credits": course_info["credits"],
                                "branch": b,
                                "year": year,
                                "students": branch_strength_normalized[normalized_year].get(b, 0),
                                "type": "Main"
                            })
                            
                            remaining_courses[year][b] -= 1
                            day_year_credits[day][year][b] += course_info["credits"]
                            day_slot_total_students[day][slot] += branch_strength_normalized[normalized_year].get(b, 0)
                            day_slot_total_courses[day][slot] += 1
                            placed = True
                            break
                        if placed:
                            break
    
    # -----------------------------
    # Step 8: Prepare Exam Schedule DataFrame
    # -----------------------------
    rows = []
    for day in days:
        row = {"Day": day}
        for slot in slots:
            for year in years:
                course_list = [f"{c['course_code']} ({c['students']} students)" for c in schedule[day][slot][year]]
                row[f"{slot} - {year}"] = ", ".join(course_list) if course_list else "Empty"
        for year in years:
            for branch in branches:
                row[f"Credits {year}-{branch}"] = day_year_credits[day][year][branch]
        for slot in slots:
            row[f"Total Students - {slot}"] = day_slot_total_students[day][slot]
        rows.append(row)
    
    df_schedule = pd.DataFrame(rows)
    
    # -----------------------------
    # Step 9 & 10: FIXED Room Allocation with Proper Division Tracking
    # -----------------------------
    student_book = pd.read_excel(os.path.join(UPLOADS_FOLDER, "students.xlsx"), sheet_name=None)
    student_ids = {}
    
    for sheet_name, df in student_book.items():
        year_key = normalize_year(sheet_name.replace(" Year", ""))
        for branch in ["CSE", "DSAI", "ECE"]:
            ids = df[branch].dropna().astype(str).tolist()
            student_ids[(year_key, branch)] = ids
    
    room_names = df_room["Room"].dropna().astype(str).tolist()
    faculty_count = len(faculty_list)
    faculty_index = 0
    df_rooms_rows = []
    
    for day in days:
        used_faculty_today = set()
        for slot in slots:
            # Assign faculty to rooms
            assigned_faculty = []
            for _ in range(total_rooms):
                while faculty_list[faculty_index % faculty_count] in used_faculty_today:
                    faculty_index += 1
                assigned_faculty.append(faculty_list[faculty_index % faculty_count])
                used_faculty_today.add(faculty_list[faculty_index % faculty_count])
                faculty_index += 1
            room_faculty_mapping = {room_names[i]: assigned_faculty[i] for i in range(total_rooms)}
            
            # FIXED: Track room division usage properly
            # room_division_usage: {room_name: divisions_used}
            room_division_usage = {room: 0 for room in room_names}
            
            # Track which courses are already in which room to prevent duplicates
            room_course_map = {room: [] for room in room_names}
            
            # Collect all courses in this slot
            courses_in_slot = []
            for year in years:
                courses_in_slot.extend(schedule[day][slot][year])
            
            # Allocate each course to rooms
            for course in courses_in_slot:
                branch = course["branch"]
                year_label = normalize_year(course["year"])
                ids = student_ids.get((year_label, branch), [])
                
                if not ids:
                    # Handle courses without student IDs
                    target_room = None
                    for r in room_names:
                        # Check if room has space AND this course is not already in this room
                        if room_division_usage[r] < courses_per_room and course["course_code"] not in room_course_map[r]:
                            target_room = r
                            break
                    
                    if not target_room:
                        print(f"[WARNING] No available room divisions for {course['course_code']} on {day} {slot}")
                        continue
                    
                    division_num = room_division_usage[target_room] + 1
                    room_cap = room_capacity_map.get(target_room, room_capacity_per_course)
                    
                    if is_special_room(target_room):
                        room_display = f"{target_room} (Division {division_num}/{courses_per_room})"
                    else:
                        room_display = f"{target_room} (Section {division_num}/{courses_per_room})"
                    
                    df_rooms_rows.append({
                        "Day": day,
                        "Slot": slot,
                        "Course": course["course_code"],
                        "Branch": "".join([x for x in course["course_code"] if not x.isdigit()])[:4],
                        "Students": f"{course['students']} students",
                        "Rooms Assigned": room_display,
                        "Faculty": room_faculty_mapping[target_room],
                        "Room Capacity": room_cap
                    })
                    
                    room_division_usage[target_room] += 1
                    room_course_map[target_room].append(course["course_code"])
                else:
                    # Split students by room capacity
                    idx = 0
                    while idx < len(ids):
                        # Find next available room that doesn't already have this course
                        target_room = None
                        for r in room_names:
                            # Check if room has space AND this course is not already in this room
                            if room_division_usage[r] < courses_per_room and course["course_code"] not in room_course_map[r]:
                                target_room = r
                                break
                        
                        if not target_room:
                            print(f"[WARNING] No available room divisions for {course['course_code']} on {day} {slot}")
                            break
                        
                        # Get capacity for this specific room
                        room_cap = room_capacity_map.get(target_room, room_capacity_per_course)
                        
                        # Take students up to room capacity
                        end_idx = min(idx + room_cap, len(ids))
                        student_range = f"{ids[idx]}–{ids[end_idx - 1]}"
                        
                        # Calculate division number
                        division_num = room_division_usage[target_room] + 1
                        
                        # Determine room type for display
                        if is_special_room(target_room):
                            room_display = f"{target_room} (Division {division_num}/{courses_per_room})"
                        else:
                            room_display = f"{target_room} (Section {division_num}/{courses_per_room})"
                        
                        # Add to output
                        df_rooms_rows.append({
                            "Day": day,
                            "Slot": slot,
                            "Course": course["course_code"],
                            "Branch": "".join([x for x in course["course_code"] if not x.isdigit()])[:4],
                            "Students": student_range,
                            "Rooms Assigned": room_display,
                            "Faculty": room_faculty_mapping[target_room],
                            "Room Capacity": room_cap
                        })
                        
                        # Mark this division as used
                        room_division_usage[target_room] += 1
                        room_course_map[target_room].append(course["course_code"])
                        idx = end_idx
    
    df_rooms = pd.DataFrame(df_rooms_rows)
    
    # Merge courses in same room - UPDATED to preserve division info
    merged_rows = []
    room_dict = defaultdict(list)
    
    for row in df_rooms_rows:
        key = (row["Day"], row["Slot"], row["Rooms Assigned"], row["Faculty"])
        course_info = f"{row['Course']} ({row['Students']})"
        room_dict[key].append(course_info)
    
    for key, course_list in room_dict.items():
        day, slot, room, faculty = key
        merged_rows.append({
            "Day": day,
            "Slot": slot,
            "Rooms Assigned": room,
            "Faculty": faculty,
            "Courses + Students": ", ".join(course_list)
        })
    
    df_rooms_merged = pd.DataFrame(merged_rows)
    
    # -----------------------------
    # Step 10.5: Create Configuration Sheet
    # -----------------------------
    config_data = {
        "Setting": [
            "Start Date",
            "End Date", 
            "Courses Per Room",
            "Students Per Course",
            "Max Credits Per Day",
            "Total Rooms",
            "Students Per Slot (Max)"
        ],
        "Value": [
            start_date.strftime("%Y-%m-%d"),
            end_date.strftime("%Y-%m-%d"),
            courses_per_room,
            room_capacity_per_course,
            max_credits_per_day,
            total_rooms,
            slot_max_students
        ]
    }

    df_config = pd.DataFrame(config_data)

    # -----------------------------
    # Step 10.6: Course Allocation Verification
    # -----------------------------
    def verify_course_allocation(schedule, branch_courses, branch_strength_normalized, courses_per_year, years, branches):
        """
        Verify if all courses have been allocated according to branch strength
        """
        verification_results = {
            'total_courses_expected': {},
            'total_courses_allocated': {},
            'missing_courses': {},
            'extra_courses': {},
            'strength_mismatch': {},
            'courses_in_input': {},
            'summary': {}
        }
        
        # Calculate expected courses per branch
        for year in years:
            year_key = normalize_year(year)
            verification_results['total_courses_expected'][year] = {}
            verification_results['total_courses_allocated'][year] = {}
            verification_results['missing_courses'][year] = {}
            verification_results['extra_courses'][year] = {}
            verification_results['strength_mismatch'][year] = {}
            verification_results['courses_in_input'][year] = {}
            
            for branch in branches:
                # Count from CoursesPerYear.xlsx (what user expects)
                expected_from_settings = courses_per_year.get(year, 0)
                
                # Count actual courses in courselist.xlsx
                actual_courses_in_input = 0
                if year_key in branch_courses and branch in branch_courses[year_key]:
                    actual_courses_in_input = len(branch_courses[year_key][branch])
                
                # Store both counts for transparency
                verification_results['courses_in_input'][year][branch] = actual_courses_in_input
                
                # Use the minimum of expected and actual courses
                # This prevents false "missing" courses when input file is empty
                expected_count = min(expected_from_settings, actual_courses_in_input)
                verification_results['total_courses_expected'][year][branch] = expected_count
                
                # Count allocated courses
                allocated_count = 0
                allocated_course_codes = []
                student_counts = []
                
                for day in schedule:
                    for slot in schedule[day]:
                        for course in schedule[day][slot].get(year, []):
                            if course.get('branch') == branch and course.get('type') == 'Main':
                                allocated_count += 1
                                allocated_course_codes.append(course['course_code'])
                                student_counts.append(course['students'])
                
                verification_results['total_courses_allocated'][year][branch] = allocated_count
                
                # Find missing courses
                if year_key in branch_courses and branch in branch_courses[year_key]:
                    expected_courses = [c['course_code'] for c in branch_courses[year_key][branch]]
                    missing = [c for c in expected_courses if c not in allocated_course_codes]
                    verification_results['missing_courses'][year][branch] = missing
                    
                    # Find extra courses (allocated but not in expected list)
                    extra = [c for c in allocated_course_codes if c not in expected_courses]
                    verification_results['extra_courses'][year][branch] = extra
                
                # Check strength consistency
                expected_strength = branch_strength_normalized.get(year_key, {}).get(branch, 0)
                if student_counts:
                    # Check if all allocated courses have the same student count
                    if len(set(student_counts)) > 1:
                        verification_results['strength_mismatch'][year][branch] = {
                            'expected': expected_strength,
                            'found': student_counts
                        }
                    elif student_counts[0] != expected_strength:
                        verification_results['strength_mismatch'][year][branch] = {
                            'expected': expected_strength,
                            'found': [student_counts[0]]
                        }
        
        # Generate summary
        total_expected = sum(verification_results['total_courses_expected'][y][b] 
                            for y in years for b in branches)
        total_allocated = sum(verification_results['total_courses_allocated'][y][b] 
                             for y in years for b in branches)
        total_missing = sum(len(verification_results['missing_courses'][y][b]) 
                           for y in years for b in branches)
        total_extra = sum(len(verification_results['extra_courses'][y][b]) 
                         for y in years for b in branches)
        total_strength_issues = sum(1 for y in years for b in branches 
                                   if verification_results['strength_mismatch'][y].get(b))
        
        verification_results['summary'] = {
            'total_courses_expected': total_expected,
            'total_courses_allocated': total_allocated,
            'total_missing_courses': total_missing,
            'total_extra_courses': total_extra,
            'total_strength_issues': total_strength_issues,
            'is_complete': total_missing == 0 and total_extra == 0 and total_strength_issues == 0
        }
        
        return verification_results

    # Perform verification
    verification_results = verify_course_allocation(
        schedule, branch_courses, branch_strength_normalized, 
        courses_per_year, years, branches
    )

    # -----------------------------
    # Step 10.7: Create Verification Report Sheet
    # -----------------------------
    verification_rows = []

    # Summary section
    verification_rows.append({
        "Category": "SUMMARY",
        "Metric": "Total Courses Expected",
        "Value": verification_results['summary']['total_courses_expected'],
        "Status": "[OK]" if verification_results['summary']['total_courses_expected'] > 0 else "[WARNING]"
    })

    verification_rows.append({
        "Category": "SUMMARY",
        "Metric": "Total Courses Allocated",
        "Value": verification_results['summary']['total_courses_allocated'],
        "Status": "[OK]" if verification_results['summary']['total_courses_allocated'] > 0 else "[ERROR]"
    })

    verification_rows.append({
        "Category": "SUMMARY",
        "Metric": "Missing Courses",
        "Value": verification_results['summary']['total_missing_courses'],
        "Status": "[OK]" if verification_results['summary']['total_missing_courses'] == 0 else "[ERROR]"
    })

    verification_rows.append({
        "Category": "SUMMARY",
        "Metric": "Extra Courses",
        "Value": verification_results['summary']['total_extra_courses'],
        "Status": "[OK]" if verification_results['summary']['total_extra_courses'] == 0 else "[WARNING]"
    })

    verification_rows.append({
        "Category": "SUMMARY",
        "Metric": "Strength Mismatches",
        "Value": verification_results['summary']['total_strength_issues'],
        "Status": "[OK]" if verification_results['summary']['total_strength_issues'] == 0 else "[ERROR]"
    })

    verification_rows.append({
        "Category": "SUMMARY",
        "Metric": "Overall Status",
        "Value": "COMPLETE" if verification_results['summary']['is_complete'] else "INCOMPLETE",
        "Status": "[COMPLETE]" if verification_results['summary']['is_complete'] else "[INCOMPLETE]"
    })

    # Detailed breakdown
    for year in years:
        for branch in branches:
            expected = verification_results['total_courses_expected'][year][branch]
            allocated = verification_results['total_courses_allocated'][year][branch]
            in_input = verification_results['courses_in_input'][year][branch]
            from_settings = courses_per_year.get(year, 0)
            missing = verification_results['missing_courses'][year][branch]
            extra = verification_results['extra_courses'][year][branch]
            strength_issue = verification_results['strength_mismatch'][year].get(branch)
            
            # Year-Branch header
            verification_rows.append({
                "Category": f"{year} - {branch}",
                "Metric": "Expected in Settings",
                "Value": from_settings,
                "Status": "[INFO]"
            })
            
            verification_rows.append({
                "Category": f"{year} - {branch}",
                "Metric": "Found in Course List",
                "Value": in_input,
                "Status": "[OK]" if in_input > 0 else "[WARNING]"
            })
            
            verification_rows.append({
                "Category": f"{year} - {branch}",
                "Metric": "Allocated Courses",
                "Value": allocated,
                "Status": "[OK]" if allocated == expected else "[ERROR]"
            })
            
            if missing:
                verification_rows.append({
                    "Category": f"{year} - {branch}",
                    "Metric": "Missing Courses",
                    "Value": ", ".join(missing),
                    "Status": "[ERROR]"
                })
            
            if extra:
                verification_rows.append({
                    "Category": f"{year} - {branch}",
                    "Metric": "Extra Courses",
                    "Value": ", ".join(extra),
                    "Status": "[WARNING]"
                })
            
            if strength_issue:
                verification_rows.append({
                    "Category": f"{year} - {branch}",
                    "Metric": "Strength Issue",
                    "Value": f"Expected: {strength_issue['expected']}, Found: {strength_issue['found']}",
                    "Status": "[ERROR]"
                })
            
            # Add warning if no courses in input
            if in_input == 0:
                verification_rows.append({
                    "Category": f"{year} - {branch}",
                    "Metric": "[WARNING]",
                    "Value": "No courses found in courselist.xlsx",
                    "Status": "[WARNING]"
                })

    df_verification = pd.DataFrame(verification_rows)
    
    # -----------------------------
    # Step 11: Save to Excel (Updated)
    # -----------------------------
    output_file = "exam_schedule_with_rooms_faculty.xlsx"
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_schedule.to_excel(writer, sheet_name="Exam Schedule", index=False)
        df_config.to_excel(writer, sheet_name="Configuration", index=False)
        df_verification.to_excel(writer, sheet_name="Verification Report", index=False)
        df_rooms_merged.to_excel(writer, sheet_name="Room Allocation", index=False)
        for day in days:
            df_day = df_rooms[df_rooms["Day"]==day]
            df_day.to_excel(writer, sheet_name=f"Rooms-{day}", index=False)
    
    # -----------------------------
    # Step 12: Excel Formatting
    # -----------------------------
    wb = load_workbook(output_file)
    ws_main = wb["Exam Schedule"]
    
    for i, col in enumerate(ws_main.columns, start=1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws_main.column_dimensions[get_column_letter(i)].width = max_length + 5
    
    for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws_main.row_dimensions[cell.row].height = 30
    
    branch_colors = {
        "CSE": "FFC7CE",
        "DSAI": "C6EFCE",
        "ECE": "FFEB9C",
        "All": "BDD7EE"
    }
    
    for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row):
        for cell in row:
            for branch, color in branch_colors.items():
                if branch in str(cell.value):
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(bold=True)
    
    # Format Configuration sheet
    if "Configuration" in wb.sheetnames:
        ws_config = wb["Configuration"]
        for i, col in enumerate(ws_config.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws_config.column_dimensions[get_column_letter(i)].width = max_length + 5
        
        for row in ws_config.iter_rows(min_row=1, max_row=ws_config.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                if row[0].row == 1:  # Header row
                    cell.font = Font(bold=True)
    
    # Format Verification Report sheet
    if "Verification Report" in wb.sheetnames:
        ws_verify = wb["Verification Report"]
        for i, col in enumerate(ws_verify.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws_verify.column_dimensions[get_column_letter(i)].width = max_length + 5
        
        for row in ws_verify.iter_rows(min_row=1, max_row=ws_verify.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if row[0].row == 1:  # Header row
                    cell.font = Font(bold=True)
                elif str(cell.value) in ["[ERROR]", "[INCOMPLETE]"]:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif str(cell.value) in ["[OK]", "[COMPLETE]"]:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif str(cell.value) == "[WARNING]":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    
    wb.save(output_file)
    
    # -----------------------------
    # Step 13: Create Free Slot Sheet
    # -----------------------------
    free_rows = []
    
    for day in days:
        for slot in slots:
            # Extract base room names (without division info)
            rooms_filled = df_rooms[(df_rooms["Day"]==day) & (df_rooms["Slot"]==slot)]["Rooms Assigned"].apply(lambda x: x.split(' (')[0]).unique().tolist()
            available_rooms = [r for r in room_names if r not in rooms_filled]
            
            total_students_in_slot = day_slot_total_students[day][slot]
            remaining_capacity = max_students_per_slot - total_students_in_slot
            
            for year in years:
                for branch in branches:
                    courses_assigned = schedule[day][slot][year]
                    assigned_branches = [c["branch"] for c in courses_assigned]
                    has_common_course = any(c["branch"] == "All" for c in courses_assigned)
                    is_free = branch not in assigned_branches and not has_common_course
                    
                    free_rows.append({
                        "Day": day,
                        "Slot": slot,
                        "Year": year,
                        "Branch": branch,
                        "Status": "Free" if is_free else "Engaged",
                        "Available Rooms": ", ".join(available_rooms) if year==years[0] and branch==branches[0] else "",
                        "Remaining Capacity": remaining_capacity if year==years[0] and branch==branches[0] else ""
                    })
    
    df_free_slots = pd.DataFrame(free_rows)
    
    with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:
        df_free_slots.to_excel(writer, sheet_name="Free Slots", index=False)
    
    # Final formatting for all sheets
    wb = load_workbook(output_file)
    
    sheet_names = wb.sheetnames
    
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        
        for i, col in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(i)].width = max_length + 5
        
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row_cells:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.row_dimensions[cell.row].height = 30
        
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row_cells:
                cell_value_upper = str(cell.value).upper()
                for branch, color in branch_colors.items():
                    if branch in cell_value_upper:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        cell.font = Font(bold=True)
    
    wb.save(output_file)
    print(f"\n[SUCCESS] Exam timetable generated successfully: {output_file}")
    print(f"Configuration: {courses_per_room} courses per room")
    
    # -----------------------------
    # VERIFICATION OUTPUT
    # -----------------------------
    print("\n" + "="*70)
    print("ROOM ALLOCATION VERIFICATION")
    print("="*70)
    verification_df = df_rooms.copy()
    verification_df['Base_Room'] = verification_df['Rooms Assigned'].apply(lambda x: x.split(' (')[0])
    grouped = verification_df.groupby(['Day', 'Slot', 'Base_Room']).size()
    
    violations = []
    correct_count = 0
    for (day, slot, room), count in grouped.items():
        if count != courses_per_room:
            violations.append(f"{day} {slot} {room}: {count} courses (expected {courses_per_room})")
        else:
            correct_count += 1
    
    if not violations:
        print(f"[SUCCESS] Perfect! All {correct_count} room-slot combinations have exactly {courses_per_room} courses!")
    else:
        print(f"[WARNING] Found {len(violations)} room allocation issues:")
        for v in violations[:10]:  # Show first 10
            print(f"  - {v}")
        if len(violations) > 10:
            print(f"  ... and {len(violations) - 10} more")
    
    # Sample verification display
    print(f"\n[INFO] Sample room allocations:")
    sample = verification_df.head(9)
    for idx, row in sample.iterrows():
        print(f"  {row['Day']} {row['Slot']}: {row['Rooms Assigned']} - {row['Course']} ({row['Students']})")
    
    # -----------------------------
    # COURSE ALLOCATION VERIFICATION OUTPUT
    # -----------------------------
    print("\n" + "="*70)
    print("COURSE ALLOCATION VERIFICATION")
    print("="*70)

    summary = verification_results['summary']
    print(f"Total Courses Expected: {summary['total_courses_expected']}")
    print(f"Total Courses Allocated: {summary['total_courses_allocated']}")
    print(f"Missing Courses: {summary['total_missing_courses']}")
    print(f"Extra Courses: {summary['total_extra_courses']}")
    print(f"Strength Mismatches: {summary['total_strength_issues']}")
    print(f"Overall Status: {'[COMPLETE]' if summary['is_complete'] else '[INCOMPLETE]'}")

    if not summary['is_complete']:
        print("\n[WARNING] Issues Found:")
        for year in years:
            for branch in branches:
                from_settings = courses_per_year.get(year, 0)
                in_input = verification_results['courses_in_input'][year][branch]
                missing = verification_results['missing_courses'][year][branch]
                extra = verification_results['extra_courses'][year][branch]
                strength_issue = verification_results['strength_mismatch'][year].get(branch)
                
                if from_settings > 0 or in_input > 0 or missing or extra or strength_issue:
                    print(f"\n  {year} - {branch}:")
                    print(f"    Expected in Settings: {from_settings}")
                    print(f"    Found in Course List: {in_input}")
                    if missing:
                        print(f"    Missing: {', '.join(missing)}")
                    if extra:
                        print(f"    Extra: {', '.join(extra)}")
                    if strength_issue:
                        print(f"    Strength: Expected {strength_issue['expected']}, Found {strength_issue['found']}")
                    if in_input == 0:
                        print(f"    [WARNING] No courses found in courselist.xlsx")

    print("="*70)
    
    return output_file


# Main function to run when script is executed directly
if __name__ == "__main__":
    print("Exam Timetable Generator")
    print("="*50)
    
    # Check if uploads folder exists
    if not os.path.exists(UPLOADS_FOLDER):
        print(f"Creating uploads folder: {UPLOADS_FOLDER}")
        os.makedirs(UPLOADS_FOLDER)
    
    # 1. Input Dates
    print("\nPlease enter the date range for the exam timetable:")
    try:
        start_date_input = input("  Start Date (YYYY-MM-DD): ").strip()
        end_date_input = input("  End Date (YYYY-MM-DD): ").strip()
        
        # Validate date format
        datetime.strptime(start_date_input, "%Y-%m-%d")
        datetime.strptime(end_date_input, "%Y-%m-%d")
    except ValueError:
        print("\n[ERROR] Invalid date format. Please use YYYY-MM-DD (e.g., 2025-01-01). Exiting.")
        exit()

    # 2. Define Structure
    # Based on the prompt, we handle 3 years and 3 branches
    target_years = ["1st Year", "2nd Year", "3rd Year"]
    target_branches = ["CSE", "DSAI", "ECE"]
    
    # Create all combinations (9 total)
    all_combinations = []
    for y in target_years:
        for b in target_branches:
            all_combinations.append((y, b))
            
    # 3. Determine Allocation Method
    print("\nHow would you like to assign slots to the 9 Branch-Year combinations?")
    print("  1. Manual Input (Enter Morning/Evening for each of the 9)")
    print("  2. Random Auto-Assign (Randomly selects 4 for Morning, 5 for Evening)")
    
    method = input("\n  Enter choice (1 or 2): ").strip()
    
    branch_slot_allocation = {}
    
    # Initialize dict structure
    for y in target_years:
        norm_y = normalize_year(y)
        branch_slot_allocation[norm_y] = {"Morning": [], "Evening": []}

    if method == '1':
        # Manual Input Mode
        print("\n--- Manual Slot Assignment ---")
        print("Enter 'Morning' or 'Evening' (or M/E) for each:")
        
        for y in target_years:
            norm_y = normalize_year(y)
            for b in target_branches:
                while True:
                    user_input = input(f"  {y} - {b}: ").strip().lower()
                    if user_input in ['morning', 'm', 'evening', 'e']:
                        slot_name = "Morning" if user_input in ['morning', 'm'] else "Evening"
                        branch_slot_allocation[norm_y][slot_name].append(b)
                        break
                    else:
                        print("    [!] Invalid input. Please enter 'Morning' or 'Evening'.")
        
        # Verification for manual input
        total_morning = sum(len(v["Morning"]) for v in branch_slot_allocation.values())
        total_evening = sum(len(v["Evening"]) for v in branch_slot_allocation.values())
        print(f"\n  [INFO] Total allocated: {total_morning} in Morning, {total_evening} in Evening.")
        if total_morning not in [4, 5] or total_evening not in [4, 5]:
             print("  [WARNING] Allocation is not balanced (4 vs 5). Proceeding with user preference.")

    else:
        # Random Auto-Assign Mode (Default to 2 if invalid input)
        print("\n--- Generating Random Balanced Allocation ---")
        random.shuffle(all_combinations)
        
        # Requirement: Randomly take 4 branches from 9 in one slot and other 5 in other slot
        # We'll randomly choose which slot gets 4 or 5 to add variety
        if random.choice([True, False]):
            morning_set = all_combinations[:4]
            evening_set = all_combinations[4:]
        else:
            morning_set = all_combinations[:5]
            evening_set = all_combinations[5:]
            
        print("\n  Generated Allocation:")
        for y, b in morning_set:
            norm_y = normalize_year(y)
            branch_slot_allocation[norm_y]["Morning"].append(b)
            print(f"    {y} - {b}: Morning")
            
        for y, b in evening_set:
            norm_y = normalize_year(y)
            branch_slot_allocation[norm_y]["Evening"].append(b)
            print(f"    {y} - {b}: Evening")
            
    # Default parameters
    max_credits_per_day = 5
    courses_per_room = 2
    
    # Generate timetable
    try:
        output_file = generate_timetable(
            start_date=start_date_input,
            end_date=end_date_input,
            branch_slot_allocation=branch_slot_allocation,
            max_credits_per_day=max_credits_per_day,
            courses_per_room=courses_per_room
        )
        
        print(f"\n[SUCCESS] Timetable generation completed successfully!")
        print(f"Output file: {output_file}")
        print(f"You can find the generated timetable in the current directory.")
    except Exception as e:
        print(f"\n[ERROR] Error generating timetable: {e}")
        import traceback
        traceback.print_exc()