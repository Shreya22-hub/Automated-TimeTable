from flask import Flask, render_template, request, send_file, flash, redirect, url_for, jsonify, session
import os
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import ExamTimeTable as ett
import pandas as pd
from openpyxl import load_workbook
import json
from datetime import datetime
from functools import wraps

app = Flask(__name__,template_folder="app2")
app.secret_key = 'your-secret-key-here-change-this-in-production'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create uploads folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Admin credentials (In production, use a database!)
ADMIN_CREDENTIALS = {
    'username': 'admin',
    'password': generate_password_hash('admin123')  # Change this password!
}

REQUIRED_FILES = {
    'branch_strength': 'BranchStrength.xlsx',
    'courses_per_year': 'CoursesPerYear.xlsx',
    'common_course': 'CommonCourse.xlsx',
    'settings': 'Settings.xlsx',
    'faculty': 'FACULTY.csv',
    'rooms': 'rooms.xlsx',
    'courselist': 'courselist.xlsx',
    'students': 'students.xlsx'
}

# Decorator for admin-only routes
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('is_admin'):
            flash('Admin access required. Please login.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Decorator for routes that require timetable to exist
def timetable_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        output_file = session.get('output_file', 'exam_schedule_with_rooms_faculty.xlsx')
        if not os.path.exists(output_file):
            flash('No timetable found. Please generate one first or contact admin.', 'error')
            return redirect(url_for('public_home'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username == ADMIN_CREDENTIALS['username'] and check_password_hash(ADMIN_CREDENTIALS['password'], password):
            session['is_admin'] = True
            session['username'] = username
            flash('Login successful! Welcome Admin.', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid credentials. Please try again.', 'error')
            return redirect(url_for('login'))
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('is_admin', None)
    session.pop('username', None)
    flash('Logged out successfully.', 'success')
    return redirect(url_for('public_home'))

@app.route('/public')
def public_home():
    """Public home page for viewing timetable"""
    output_file = session.get('output_file', 'exam_schedule_with_rooms_faculty.xlsx')
    has_timetable = os.path.exists(output_file)
    return render_template('public_home.html', has_timetable=has_timetable)

@app.route('/', methods=['GET', 'POST'])
@admin_required
def index():
    if request.method == 'POST':
        try:
            # Get date inputs
            start_date = request.form.get('start_date')
            end_date = request.form.get('end_date')
            max_credits_per_day = request.form.get('max_credits_per_day', '5')
            
            if not start_date or not end_date:
                flash('Please provide both start and end dates', 'error')
                return redirect(url_for('index'))
            
            # Convert max_credits_per_day to integer
            try:
                max_credits = int(max_credits_per_day)
                if max_credits < 3 or max_credits > 10:
                    flash('Maximum credits per day must be between 3 and 10', 'error')
                    return redirect(url_for('index'))
            except ValueError:
                flash('Invalid maximum credits per day value', 'error')
                return redirect(url_for('index'))
            
            # Get branch slot allocation
            branch_allocation = {}
            years = ['1St Year', '2Nd Year', '3Rd Year', '4Th Year']
            slots = ['Morning', 'Evening']
            
            for year in years:
                branch_allocation[year] = {}
                for slot in slots:
                    field_name = f'allocation_{year.replace(" ", "_")}_{slot}'
                    branches_str = request.form.get(field_name, '')
                    branches = [b.strip().upper() for b in branches_str.split(',') if b.strip()]
                    branch_allocation[year][slot] = branches
            
            print("Branch Allocation Configuration:")
            print(branch_allocation)
            print(f"Max Credits Per Day: {max_credits}")
            
            # Save uploaded files
            saved_files = {}
            for field_name, filename in REQUIRED_FILES.items():
                if field_name not in request.files:
                    flash(f'Missing file: {filename}', 'error')
                    return redirect(url_for('index'))
                
                file = request.files[field_name]
                if file.filename == '':
                    flash(f'No file selected for: {filename}', 'error')
                    return redirect(url_for('index'))
                
                filepath = filename
                file.save(filepath)
                saved_files[field_name] = filepath
                print(f"Saved: {filepath}")
            
            # Store configuration in session for viewing later
            session['start_date'] = start_date
            session['end_date'] = end_date
            session['max_credits'] = max_credits
            session['branch_allocation'] = branch_allocation
            
            # Call the exam timetable generation function
            print("Calling generate_timetable function...")
            output_file = ett.generate_timetable(
                start_date=start_date,
                end_date=end_date,
                branch_slot_allocation=branch_allocation,
                max_credits_per_day=max_credits
            )
            
            print(f"Timetable generated: {output_file}")
            session['output_file'] = output_file
            flash('Exam timetable generated successfully! Redirecting to viewer...', 'success')
            
            # Automatically redirect to view page
            return redirect(url_for('view_timetable'))
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error occurred: {error_details}")
            flash(f'Error generating timetable: {str(e)}', 'error')
            return redirect(url_for('index'))
    
    # Check if there's an existing timetable
    has_existing = session.get('output_file') and os.path.exists(session.get('output_file', ''))
    return render_template('index.html', has_existing=has_existing, is_admin=True)

@app.route('/view')
@timetable_required
def view_timetable():
    """View the generated timetable online"""
    output_file = session.get('output_file', 'exam_schedule_with_rooms_faculty.xlsx')
    
    try:
        # Read all sheets from Excel
        excel_data = pd.read_excel(output_file, sheet_name=None)
        
        # Convert to JSON-friendly format
        sheets_data = {}
        for sheet_name, df in excel_data.items():
            # Replace NaN with empty string for JSON serialization
            df = df.fillna('')
            sheets_data[sheet_name] = {
                'columns': df.columns.tolist(),
                'data': df.values.tolist()
            }
        
        is_admin = session.get('is_admin', False)
        return render_template('view_timetable.html', 
                             sheets=sheets_data,
                             output_file=output_file,
                             is_admin=is_admin)
    
    except Exception as e:
        flash(f'Error reading timetable: {str(e)}', 'error')
        return redirect(url_for('public_home'))

@app.route('/download')
@timetable_required
def download_timetable():
    """Download the generated Excel file"""
    output_file = session.get('output_file', 'exam_schedule_with_rooms_faculty.xlsx')
    
    return send_file(
        output_file,
        as_attachment=True,
        download_name=f'exam_schedule_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/api/update_cell', methods=['POST'])
@admin_required
def update_cell():
    """Update a cell in the timetable - ADMIN ONLY"""
    try:
        data = request.json
        sheet_name = data.get('sheet')
        row_index = int(data.get('row'))
        col_index = int(data.get('col'))
        new_value = data.get('value')
        
        output_file = session.get('output_file', 'exam_schedule_with_rooms_faculty.xlsx')
        
        if not os.path.exists(output_file):
            return jsonify({'success': False, 'error': 'File not found'})
        
        # Load workbook
        wb = load_workbook(output_file)
        ws = wb[sheet_name]
        
        # Update cell (Excel is 1-indexed, add 2 for header)
        cell = ws.cell(row=row_index + 2, column=col_index + 1)
        cell.value = new_value
        
        # Save workbook
        wb.save(output_file)
        
        return jsonify({'success': True, 'message': 'Cell updated successfully'})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/get_sheet_data/<sheet_name>')
@timetable_required
def get_sheet_data(sheet_name):
    """Get data for a specific sheet"""
    try:
        output_file = session.get('output_file', 'exam_schedule_with_rooms_faculty.xlsx')
        
        df = pd.read_excel(output_file, sheet_name=sheet_name)
        df = df.fillna('')
        
        return jsonify({
            'success': True,
            'columns': df.columns.tolist(),
            'data': df.values.tolist()
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/health')
def health():
    return {'status': 'ok', 'message': 'Exam Timetable Generator is running'}

if __name__ == '__main__':
    print("Starting Exam Timetable Generator...")
    print(f"Upload folder: {app.config['UPLOAD_FOLDER']}")
    print("=" * 60)
    print("ADMIN LOGIN CREDENTIALS:")
    print("Username: admin")
    print("Password: admin123")
    print("=" * 60)
    print("IMPORTANT: Change the password in production!")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=5001)